"""
LLM Monitoring Routes Blueprint

Este módulo contiene todos los endpoints REST para el sistema Multi-LLM Brand Monitoring.
Se registra como Blueprint en app.py para mantener el código modular.

MODELO DE NEGOCIO:
    Los usuarios pagan una suscripción y el servicio incluye acceso a los LLMs.
    Las API keys son gestionadas globalmente por el dueño del servicio (variables de entorno).
    Los usuarios NO necesitan configurar sus propias API keys.

Endpoints:
    GET    /api/llm-monitoring/projects               - Listar proyectos
    POST   /api/llm-monitoring/projects               - Crear proyecto
    POST   /api/llm-monitoring/projects/:id/run-initial-analysis - Ejecutar primer análisis (una vez)
    GET    /api/llm-monitoring/projects/:id           - Obtener proyecto
    PUT    /api/llm-monitoring/projects/:id           - Actualizar proyecto
    DELETE /api/llm-monitoring/projects/:id           - Eliminar proyecto (soft delete)
    GET    /api/llm-monitoring/projects/:id/metrics   - Métricas detalladas
    GET    /api/llm-monitoring/projects/:id/comparison - Comparativa LLMs
    GET    /api/llm-monitoring/models                 - Listar modelos LLM disponibles
    PUT    /api/llm-monitoring/models/:id             - Actualizar modelo (admin)
    GET    /api/llm-monitoring/health                 - Health check del sistema
    
    NOTA: El endpoint POST /projects/:id/analyze fue ELIMINADO.
          El análisis ahora se ejecuta AUTOMÁTICAMENTE vía cron diario a las 4:00 AM.
"""

import logging
import os
import json
import re
import unicodedata
import threading
import secrets
from datetime import datetime, timedelta
from urllib.parse import urlparse
from flask import Blueprint, request, jsonify
from functools import wraps

# Importar sistema de autenticación
from auth import login_required, admin_required, get_current_user, cron_or_auth_required
from llm_monitoring_limits import (
    can_access_llm_monitoring,
    get_llm_plan_limits,
    count_user_active_projects,
    count_project_active_queries,
    get_llm_limits_summary,
    get_upgrade_options,
)
from services.project_access_service import (
    get_project_permissions,
    user_can_view_project,
    user_has_any_module_access,
)

# Importar servicios
from database import get_db_connection, acquire_analysis_lock, release_analysis_lock, get_latest_analysis_run
from services.llm_monitoring_service import MultiLLMMonitoringService, analyze_all_active_projects
from services.llm_monitoring_stats import LLMMonitoringStatsService

# Configurar logging
logger = logging.getLogger(__name__)

# Crear Blueprint
llm_monitoring_bp = Blueprint('llm_monitoring', __name__, url_prefix='/api/llm-monitoring')

# Estado en memoria para evitar dobles disparos del primer análisis
# (misma instancia de app). La fuente de verdad para "ya analizado" sigue en BD.
_INITIAL_ANALYSIS_RUNNING = set()
_INITIAL_ANALYSIS_RUNNING_LOCK = threading.Lock()


def _is_initial_analysis_running(project_id: int) -> bool:
    with _INITIAL_ANALYSIS_RUNNING_LOCK:
        return int(project_id) in _INITIAL_ANALYSIS_RUNNING


def _mark_initial_analysis_running(project_id: int) -> bool:
    """
    Marca un proyecto como "primer análisis en curso".
    Returns False si ya estaba en curso.
    """
    with _INITIAL_ANALYSIS_RUNNING_LOCK:
        normalized_id = int(project_id)
        if normalized_id in _INITIAL_ANALYSIS_RUNNING:
            return False
        _INITIAL_ANALYSIS_RUNNING.add(normalized_id)
        return True


def _clear_initial_analysis_running(project_id: int):
    with _INITIAL_ANALYSIS_RUNNING_LOCK:
        _INITIAL_ANALYSIS_RUNNING.discard(int(project_id))


# ============================================================================
# DECORADORES AUXILIARES
# ============================================================================

@llm_monitoring_bp.before_request
def enforce_llm_access():
    """
    Bloquea acceso si el usuario no tiene plan válido/billing activo.
    Excepciones: endpoints de cron y health.
    """
    try:
        path = request.path or ""
        if "/api/llm-monitoring/cron/" in path or path.endswith("/health"):
            return None
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Usuario no autenticado'}), 401
        if not can_access_llm_monitoring(user):
            # ✅ Permitir invitados con acceso explícito a proyectos compartidos
            if user_has_any_module_access(user['id'], 'llm_monitoring'):
                return None
            return jsonify({
                'error': 'paywall',
                'message': 'LLM Monitoring requires a paid plan',
                'upgrade_options': get_upgrade_options(user.get('plan', 'free')),
                'current_plan': user.get('plan', 'free')
            }), 402
    except Exception as e:
        logger.error(f"Error en enforce_llm_access: {e}", exc_info=True)
        return jsonify({'error': 'Error validando acceso'}), 500

def validate_project_ownership(f):
    """
    Decorador de acceso al proyecto.
    - Owner: acceso total (GET/POST/PUT/DELETE)
    - Colaborador viewer: solo acceso GET
    """
    @wraps(f)
    def decorated_function(project_id, *args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'error': 'Usuario no autenticado'}), 401
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Error de conexión a BD'}), 500
        
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT user_id FROM llm_monitoring_projects
                WHERE id = %s
            """, (project_id,))
            
            project = cur.fetchone()
            
            if not project:
                return jsonify({'error': 'Proyecto no encontrado'}), 404

            is_owner = project['user_id'] == user['id']
            if not is_owner:
                can_view_shared = request.method == 'GET' and user_can_view_project(
                    user['id'],
                    'llm_monitoring',
                    project_id
                )
                if not can_view_shared:
                    return jsonify({'error': 'No tienes permiso para acceder a este proyecto'}), 403
            
            return f(project_id, *args, **kwargs)
            
        except Exception as e:
            logger.error(f"Error validando ownership: {e}", exc_info=True)
            return jsonify({'error': f'Error validando permisos: {str(e)}'}), 500
        finally:
            cur.close()
            conn.close()
    
    return decorated_function


def _ensure_cron_token_or_admin():
    """
    Endurece endpoints sensibles de cron:
    - Permite CRON token válido
    - O usuario autenticado con rol admin
    - Bloquea usuarios autenticados no-admin
    """
    try:
        auth_header = request.headers.get('Authorization', '') or ''
        token = auth_header[7:].strip() if auth_header.lower().startswith('bearer ') else ''
        cron_secret = os.environ.get('CRON_TOKEN') or os.environ.get('CRON_SECRET')

        if cron_secret and token and secrets.compare_digest(token, cron_secret):
            return None

        user = get_current_user()
        if user and user.get('role') == 'admin':
            return None

        return jsonify({
            'success': False,
            'error': 'forbidden',
            'message': 'Se requiere token de cron o rol admin'
        }), 403
    except Exception as e:
        logger.error(f"Error validando acceso cron/admin: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Error validando permisos'}), 500


def _normalize_days_param(raw_days, default: int = 30, min_days: int = 1, max_days: int = 365) -> int:
    """
    Normaliza el parámetro days para evitar rangos no válidos o excesivos.
    """
    try:
        if raw_days is None or raw_days == '':
            return default
        days = int(raw_days)
    except (TypeError, ValueError):
        return default

    if days < min_days:
        return min_days
    if days > max_days:
        return max_days
    return days


def _remove_accents(text):
    """Remove accents for tolerant matching (e.g. 'café' matches 'cafe')."""
    nfkd = unicodedata.normalize('NFD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def classify_query_branded(query_text, brand_keywords):
    """
    Returns True if query_text contains any brand keyword.
    Uses case-insensitive, accent-insensitive, word-boundary matching.
    """
    if not brand_keywords or not query_text:
        return False
    text_lower = query_text.lower()
    text_no_accents = _remove_accents(text_lower)
    for kw in brand_keywords:
        kw_lower = kw.lower()
        kw_no_accents = _remove_accents(kw_lower)
        for pattern_text, text_to_search in [(kw_lower, text_lower), (kw_no_accents, text_no_accents)]:
            try:
                if re.search(r'\b' + re.escape(pattern_text) + r'\b', text_to_search, re.IGNORECASE):
                    return True
            except re.error:
                if pattern_text in text_to_search:
                    return True
    return False


def _compute_branded_metrics(results, brand_keywords):
    """
    Splits results into branded/non-branded and computes metrics for each.
    Each result must have: query_text, brand_mentioned, position_in_list, competitors_mentioned.
    Returns dict with 'branded' and 'non_branded' metric blocks.
    """
    branded = []
    non_branded = []
    for r in results:
        if classify_query_branded(r.get('query_text', ''), brand_keywords):
            branded.append(r)
        else:
            non_branded.append(r)

    def compute(subset, label):
        if not subset:
            return {
                'label': label,
                'total_queries': 0,
                'total_results': 0,
                'total_mentions': 0,
                'mention_rate': 0.0,
                'share_of_voice': 0.0,
                'avg_position': None
            }
        unique_queries = len(set(r.get('query_text', '') for r in subset))
        total = len(subset)
        mentions = sum(1 for r in subset if r.get('brand_mentioned'))
        mention_rate = round((mentions / total) * 100, 1) if total > 0 else 0.0

        # SOV: brand / (brand + competitors)
        total_comp = 0
        for r in subset:
            cm = r.get('competitors_mentioned')
            if cm and isinstance(cm, dict):
                total_comp += sum(cm.values())
            elif cm and isinstance(cm, str):
                try:
                    parsed = json.loads(cm)
                    total_comp += sum(parsed.values()) if isinstance(parsed, dict) else 0
                except (json.JSONDecodeError, TypeError):
                    pass
        total_all = mentions + total_comp
        sov = round((mentions / total_all) * 100, 1) if total_all > 0 else 0.0

        positions = [r['position_in_list'] for r in subset
                     if r.get('position_in_list') is not None]
        avg_pos = round(sum(positions) / len(positions), 1) if positions else None

        return {
            'label': label,
            'total_queries': unique_queries,
            'total_results': total,
            'total_mentions': mentions,
            'mention_rate': mention_rate,
            'share_of_voice': sov,
            'avg_position': avg_pos
        }

    return {
        'branded_metrics': compute(branded, 'Branded'),
        'non_branded_metrics': compute(non_branded, 'Non-Branded')
    }


def _calculate_trend(current, previous, has_previous_data):
    """
    Calcula tendencia: direction (up/down/stable) y change (%).
    Si no hay histórico suficiente en el período anterior, devuelve None.
    Reutilizable desde múltiples endpoints (/detail, /metrics, /comparison, exports).
    """
    if not has_previous_data:
        return None

    if previous == 0:
        if current > 0:
            return {'direction': 'up', 'change': 100, 'previous': 0}
        return {'direction': 'stable', 'change': 0, 'previous': 0}

    change = ((current - previous) / previous) * 100

    # Considerar "stable" si el cambio es menor al 2%
    if abs(change) < 2:
        direction = 'stable'
    else:
        direction = 'up' if change > 0 else 'down'

    return {
        'direction': direction,
        'change': round(abs(change), 1),
        'previous': round(previous, 1)
    }


def _get_effective_plan_limits(user: dict) -> dict:
    """
    Devuelve límites efectivos por usuario.
    Admin opera sin límites para soporte y validación interna.
    Enterprise: respeta custom_llm_prompts_limit y custom_llm_monthly_units_limit
    si están configurados por el admin; si no, opera sin límites (None).
    """
    limits = get_llm_plan_limits((user or {}).get('plan', 'free'))
    if user and user.get('role') == 'admin':
        limits = dict(limits)
        limits['max_projects'] = None
        limits['max_prompts_per_project'] = None
        limits['max_monthly_units'] = None
    elif user and (user or {}).get('plan') == 'enterprise':
        limits = dict(limits)
        # Aplicar custom limits si el admin los ha configurado para este usuario
        custom_prompts = user.get('custom_llm_prompts_limit')
        custom_units = user.get('custom_llm_monthly_units_limit')
        if custom_prompts is not None:
            limits['max_prompts_per_project'] = int(custom_prompts)
        if custom_units is not None:
            limits['max_monthly_units'] = int(custom_units)
    return limits


# ============================================================================
# ENDPOINTS: PROYECTOS
# ============================================================================

@llm_monitoring_bp.route('/projects', methods=['GET'])
@login_required
def get_projects():
    """
    Lista todos los proyectos de monitorización del usuario actual
    
    Returns:
        JSON con lista de proyectos y métricas básicas
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Usuario no autenticado'}), 401
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()

        try:
            # Owner projects + shared projects (viewer) when collaboration tables exist.
            cur.execute("""
                SELECT 
                    p.id,
                    p.user_id,
                    p.name,
                    p.brand_name,
                    p.brand_domain,
                    p.brand_keywords,
                    p.industry,
                    p.enabled_llms,
                    p.competitors,
                    p.competitor_domains,
                    p.competitor_keywords,
                    p.selected_competitors,
                    p.language,
                    p.country_code,
                    p.queries_per_llm,
                    p.is_active,
                    p.is_paused_by_quota,
                    p.paused_until,
                    p.paused_at,
                    p.paused_reason,
                    p.last_analysis_date,
                    p.created_at,
                    p.updated_at,
                    CASE WHEN p.user_id = %s THEN 'owner' ELSE 'viewer' END AS access_role,
                    (p.user_id = %s) AS is_owner,
                    (p.user_id = %s) AS can_edit,
                    (p.user_id = %s) AS can_manage_access,
                    (
                        SELECT COUNT(*)
                        FROM llm_monitoring_queries q
                        WHERE q.project_id = p.id AND q.is_active = TRUE
                    ) as total_queries,
                    COUNT(DISTINCT s.id) as total_snapshots,
                    MAX(s.snapshot_date) as last_snapshot_date
                FROM llm_monitoring_projects p
                LEFT JOIN llm_monitoring_snapshots s ON p.id = s.project_id
                WHERE (
                    p.user_id = %s
                    OR EXISTS (
                        SELECT 1
                        FROM project_collaborators c
                        WHERE c.module_name = 'llm_monitoring'
                          AND c.project_id = p.id
                          AND c.user_id = %s
                    )
                )
                GROUP BY p.id
                ORDER BY p.created_at DESC
            """, (
                user['id'],
                user['id'],
                user['id'],
                user['id'],
                user['id'],
                user['id'],
            ))
        except Exception as query_error:
            if "project_collaborators" not in str(query_error).lower():
                raise
            logger.warning(
                "project_collaborators table not available yet. Falling back to owner-only projects list."
            )
            conn.rollback()
            cur.execute("""
                SELECT 
                    p.id,
                    p.user_id,
                    p.name,
                    p.brand_name,
                    p.brand_domain,
                    p.brand_keywords,
                    p.industry,
                    p.enabled_llms,
                    p.competitors,
                    p.competitor_domains,
                    p.competitor_keywords,
                    p.selected_competitors,
                    p.language,
                    p.country_code,
                    p.queries_per_llm,
                    p.is_active,
                    p.is_paused_by_quota,
                    p.paused_until,
                    p.paused_at,
                    p.paused_reason,
                    p.last_analysis_date,
                    p.created_at,
                    p.updated_at,
                    'owner' AS access_role,
                    TRUE AS is_owner,
                    TRUE AS can_edit,
                    TRUE AS can_manage_access,
                    (
                        SELECT COUNT(*)
                        FROM llm_monitoring_queries q
                        WHERE q.project_id = p.id AND q.is_active = TRUE
                    ) as total_queries,
                    COUNT(DISTINCT s.id) as total_snapshots,
                    MAX(s.snapshot_date) as last_snapshot_date
                FROM llm_monitoring_projects p
                LEFT JOIN llm_monitoring_snapshots s ON p.id = s.project_id
                WHERE p.user_id = %s
                GROUP BY p.id
                ORDER BY p.created_at DESC
            """, (user['id'],))

        projects = cur.fetchall()
        
        # Formatear respuesta
        projects_list = []
        for project in projects:
            is_owner = bool(project.get('is_owner')) if project.get('is_owner') is not None else project.get('user_id') == user['id']
            can_edit = bool(project.get('can_edit')) if project.get('can_edit') is not None else is_owner
            can_manage_access = bool(project.get('can_manage_access')) if project.get('can_manage_access') is not None else is_owner
            projects_list.append({
                'id': project['id'],
                'name': project['name'],
                'brand_name': project['brand_name'],
                'brand_domain': project.get('brand_domain'),
                'brand_keywords': project.get('brand_keywords') or [],
                'industry': project['industry'],
                'enabled_llms': project['enabled_llms'],
                'competitors': project['competitors'],
                'competitor_domains': project.get('competitor_domains') or [],
                'competitor_keywords': project.get('competitor_keywords') or [],
                'selected_competitors': project.get('selected_competitors') or [],
                'language': project['language'],
                'country_code': project.get('country_code'),
                'queries_per_llm': project['queries_per_llm'],
                'total_queries': project.get('total_queries') or 0,
                'is_active': project['is_active'],
                'initial_analysis_in_progress': _is_initial_analysis_running(project['id']),
                'is_paused_by_quota': project.get('is_paused_by_quota', False),
                'paused_until': project['paused_until'].isoformat() if project.get('paused_until') else None,
                'paused_at': project['paused_at'].isoformat() if project.get('paused_at') else None,
                'paused_reason': project.get('paused_reason'),
                'last_analysis_date': project['last_analysis_date'].isoformat() if project['last_analysis_date'] else None,
                'created_at': project['created_at'].isoformat() if project['created_at'] else None,
                'updated_at': project['updated_at'].isoformat() if project['updated_at'] else None,
                'total_snapshots': project['total_snapshots'],
                'last_snapshot_date': project['last_snapshot_date'].isoformat() if project['last_snapshot_date'] else None,
                'access_role': project.get('access_role') or ('owner' if is_owner else 'viewer'),
                'is_owner': is_owner,
                'can_edit': can_edit,
                'can_manage_access': can_manage_access
            })
        
        limits_summary = get_llm_limits_summary(user)

        return jsonify({
            'success': True,
            'projects': projects_list,
            'total': len(projects_list),
            'limits': limits_summary
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo proyectos: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo proyectos: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/usage', methods=['GET'])
@login_required
def get_usage():
    """
    Devuelve límites y consumo del usuario para LLM Monitoring
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Usuario no autenticado'}), 401

    limits_summary = get_llm_limits_summary(user)
    return jsonify({'success': True, 'limits': limits_summary}), 200


@llm_monitoring_bp.route('/projects', methods=['POST'])
@login_required
def create_project():
    """
    Crea un nuevo proyecto de monitorización
    
    Body esperado:
    {
        "name": "Mi Marca SEO",
        "industry": "SEO tools",
        "brand_domain": "hmfertility.com",
        "brand_keywords": ["hmfertility", "hm", "fertility clinic"],
        "competitor_domains": ["competitor1.com"],
        "competitor_keywords": ["semrush", "ahrefs"],
        "language": "es",
        "enabled_llms": ["openai", "anthropic", "google", "perplexity"]
    }
    
    Returns:
        JSON con el proyecto creado
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Usuario no autenticado'}), 401
    
    plan_limits = _get_effective_plan_limits(user)
    max_projects = plan_limits.get('max_projects')

    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        return jsonify({'error': 'Body JSON inválido'}), 400
    
    # Validar campos requeridos
    required_fields = ['name', 'industry', 'brand_keywords']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'Campo requerido: {field}'}), 400
    
    # Valores por defecto y extracción de datos
    brand_domain = data.get('brand_domain')
    brand_keywords = data.get('brand_keywords', [])
    selected_competitors = data.get('selected_competitors', [])  # ✨ NEW
    language = str(data.get('language', 'es') or 'es').strip().lower() or 'es'
    country_code = str(data.get('country_code', 'ES') or 'ES').strip().upper()
    enabled_llms = data.get('enabled_llms', ['openai', 'anthropic', 'google', 'perplexity'])
    
    # Validaciones
    if not isinstance(brand_keywords, list) or len(brand_keywords) == 0:
        return jsonify({'error': 'brand_keywords debe ser un array con al menos 1 palabra clave'}), 400
    
    max_prompts = plan_limits.get('max_prompts_per_project')
    # queries_per_llm deja de ser configurable por usuario.
    # Se guarda internamente una capacidad derivada del plan para compatibilidad.
    # Enterprise (max_prompts=None) permite hasta 5000 prompts por proyecto.
    if max_prompts is None:
        configured_prompt_capacity = 5000  # Enterprise / Admin: sin límite efectivo
    elif isinstance(max_prompts, int):
        configured_prompt_capacity = max(5, min(5000, max_prompts))
    else:
        configured_prompt_capacity = 60
    
    if not isinstance(enabled_llms, list) or len(enabled_llms) == 0:
        return jsonify({'error': 'enabled_llms debe ser un array con al menos 1 LLM'}), 400
    
    valid_llms = ['openai', 'anthropic', 'google', 'perplexity']
    if not all(llm in valid_llms for llm in enabled_llms):
        return jsonify({'error': f'LLMs válidos: {valid_llms}'}), 400

    if not country_code or len(country_code) != 2 or not country_code.isalpha():
        return jsonify({'error': 'country_code debe ser un código ISO de 2 letras'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()

        # Lock del usuario para evitar carreras en creación de proyectos
        cur.execute("SELECT id FROM users WHERE id = %s FOR UPDATE", (user['id'],))

        if max_projects is not None:
            cur.execute("""
                SELECT COUNT(*) AS count
                FROM llm_monitoring_projects
                WHERE user_id = %s AND is_active = TRUE
            """, (user['id'],))
            row = cur.fetchone()
            active_projects = int(row['count']) if row else 0
            if active_projects >= max_projects:
                return jsonify({
                    'error': 'project_limit_reached',
                    'message': 'Has alcanzado el máximo de proyectos permitidos para tu plan',
                    'current_plan': user.get('plan', 'free'),
                    'upgrade_options': get_upgrade_options(user.get('plan', 'free')),
                    'limit': max_projects,
                    'current': active_projects
                }), 402
        
        # Verificar que no exista proyecto con ese nombre para el usuario
        cur.execute("""
            SELECT id FROM llm_monitoring_projects
            WHERE user_id = %s AND name = %s
        """, (user['id'], data['name']))
        
        if cur.fetchone():
            return jsonify({'error': 'Ya tienes un proyecto con ese nombre'}), 409
        
        # ✨ NEW: Extract legacy fields from selected_competitors for backward compatibility
        competitor_domains = []
        competitor_keywords = []
        if selected_competitors:
            for comp in selected_competitors:
                if comp.get('domain'):
                    competitor_domains.append(comp['domain'])
                if comp.get('keywords'):
                    competitor_keywords.extend(comp['keywords'])
        
        # Insertar proyecto con nuevos campos
        cur.execute("""
            INSERT INTO llm_monitoring_projects (
                user_id, name, industry,
                brand_domain, brand_keywords,
                selected_competitors,
                competitor_domains, competitor_keywords,
                enabled_llms, language, country_code, queries_per_llm,
                is_active, created_at, updated_at,
                brand_name, competitors
            ) VALUES (
                %s, %s, %s,
                %s, %s::jsonb,
                %s::jsonb,
                %s::jsonb, %s::jsonb,
                %s, %s, %s, %s,
                TRUE, NOW(), NOW(),
                %s, %s::jsonb
            )
            RETURNING id, created_at
        """, (
            user['id'],
            data['name'],
            data['industry'],
            brand_domain,
            json.dumps(brand_keywords),
            json.dumps(selected_competitors),  # ✨ NEW
            json.dumps(competitor_domains),  # Legacy
            json.dumps(competitor_keywords),  # Legacy
            enabled_llms,
            language,
            country_code,
            configured_prompt_capacity,
            # Campos legacy por compatibilidad
            brand_keywords[0] if brand_keywords else 'Brand',
            json.dumps(competitor_keywords)  # Usar keywords como legacy competitors
        ))
        
        result = cur.fetchone()
        project_id = result['id']
        created_at = result['created_at']
        
        logger.info(f"✅ Proyecto {project_id} creado. El usuario deberá añadir prompts manualmente.")
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Proyecto creado exitosamente. Ahora añade tus prompts manualmente.',
            'project': {
                'id': project_id,
                'name': data['name'],
                'industry': data['industry'],
                'brand_domain': brand_domain,
                'brand_keywords': brand_keywords,
                'competitor_domains': competitor_domains,
                'competitor_keywords': competitor_keywords,
                'enabled_llms': enabled_llms,
                'language': language,
                'country_code': country_code,
                'queries_per_llm': configured_prompt_capacity,
                'is_active': True,
                'created_at': created_at.isoformat(),
                'total_queries': 0
            }
        }), 201
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error creando proyecto: {e}", exc_info=True)
        return jsonify({'error': f'Error creando proyecto: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>', methods=['GET'])
@login_required
@validate_project_ownership
def get_project(project_id):
    """
    Obtiene detalles de un proyecto específico
    
    Returns:
        JSON con detalles del proyecto y estadísticas
    """
    logger.info(f"📊 GET /projects/{project_id} - Iniciando...")

    user = get_current_user()
    if not user:
        return jsonify({'error': 'Usuario no autenticado'}), 401
    
    conn = get_db_connection()
    if not conn:
        logger.error("❌ Error de conexión a BD")
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        logger.info(f"🔍 Consultando proyecto {project_id}...")
        
        # Obtener proyecto
        cur.execute("""
            SELECT 
                p.id,
                p.user_id,
                p.name,
                p.brand_name,
                p.brand_domain,
                p.brand_keywords,
                p.industry,
                p.enabled_llms,
                p.competitors,
                p.competitor_domains,
                p.competitor_keywords,
                p.selected_competitors,
                p.language,
                p.country_code,
                p.queries_per_llm,
                p.is_active,
                p.last_analysis_date,
                p.created_at,
                p.updated_at,
                COUNT(DISTINCT q.id) FILTER (WHERE q.is_active = TRUE) as total_queries,
                COUNT(DISTINCT s.id) as total_snapshots,
                MAX(s.snapshot_date) as last_snapshot_date
            FROM llm_monitoring_projects p
            LEFT JOIN llm_monitoring_queries q ON p.id = q.project_id
            LEFT JOIN llm_monitoring_snapshots s ON p.id = s.project_id
            WHERE p.id = %s
            GROUP BY p.id, p.user_id, p.name, p.brand_name, p.brand_domain, p.brand_keywords,
                     p.industry, p.enabled_llms, p.competitors, p.competitor_domains, 
                     p.competitor_keywords, p.selected_competitors, p.language, p.country_code, p.queries_per_llm,
                     p.is_active, p.last_analysis_date, p.created_at, p.updated_at
        """, (project_id,))
        
        project = cur.fetchone()
        logger.info(f"✅ Proyecto obtenido: {project['name'] if project else 'None'}")
        
        if not project:
            logger.warning(f"⚠️ Proyecto {project_id} no encontrado")
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        
        # 📊 CALCULAR SOV AGREGADO DE TODOS LOS DÍAS DISPONIBLES (últimos 30 días)
        # Método 2: SoV Agregado - Suma TODAS las menciones de TODOS los LLMs
        # Esto refleja el volumen REAL de menciones en el mercado
        # ✨ NUEVO: Soporte para rango de fechas global
        days = _normalize_days_param(request.args.get('days'), default=30)
        metric_type = request.args.get('metric', 'normal')
        if metric_type not in ['normal', 'weighted']:
            metric_type = 'normal'
        enabled_llms_filter = project.get('enabled_llms') or []
        logger.info(f"📈 Consultando métricas para proyecto {project_id} (últimos {days} días)...")

        # Banner UX: mostrar aviso solo cuando hay histórico en el rango
        # de LLMs actualmente desactivados.
        cur.execute("""
            SELECT DISTINCT llm_provider
            FROM llm_monitoring_snapshots
            WHERE project_id = %s
              AND snapshot_date >= CURRENT_DATE - (%s * INTERVAL '1 day')
        """, (project_id, days))
        historical_llms_in_range = sorted([
            row['llm_provider']
            for row in cur.fetchall()
            if row.get('llm_provider')
        ])
        active_llms_set = set(enabled_llms_filter)
        historical_llms_set = set(historical_llms_in_range)
        excluded_llms_with_data = sorted(list(historical_llms_set - active_llms_set))
        model_scope_notice = {
            'show': len(excluded_llms_with_data) > 0,
            'range_days': days,
            'active_llms': enabled_llms_filter,
            'excluded_llms_with_data': excluded_llms_with_data,
            'historical_llms_in_range': historical_llms_in_range
        }
        
        # Obtener todos los snapshots del rango seleccionado
        # ✨ NUEVO: Incluir campos Top3/5/10 para métricas de posición granulares
        snapshots_query = """
            SELECT 
                llm_provider,
                mention_rate,
                avg_position,
                share_of_voice,
                weighted_share_of_voice,
                positive_mentions,
                neutral_mentions,
                negative_mentions,
                total_mentions,
                total_queries,
                total_competitor_mentions,
                weighted_competitor_breakdown,
                appeared_in_top3,
                appeared_in_top5,
                appeared_in_top10,
                snapshot_date
            FROM llm_monitoring_snapshots
            WHERE project_id = %s
                AND snapshot_date >= CURRENT_DATE - (%s * INTERVAL '1 day')
        """
        snapshots_params = [project_id, days]
        if enabled_llms_filter:
            snapshots_query += " AND llm_provider = ANY(%s)"
            snapshots_params.append(enabled_llms_filter)
        snapshots_query += " ORDER BY snapshot_date DESC, llm_provider"
        cur.execute(snapshots_query, snapshots_params)
        
        all_snapshots = cur.fetchall()
        logger.info(f"📊 Métricas encontradas: {len(all_snapshots)} snapshots (últimos {days} días)")
        
        # ✨ NUEVO: Obtener snapshots del período ANTERIOR para calcular tendencias
        # Si el período actual es "últimos 30 días", el anterior es "hace 60-31 días"
        previous_snapshots_query = """
            SELECT 
                llm_provider,
                mention_rate,
                avg_position,
                share_of_voice,
                weighted_share_of_voice,
                positive_mentions,
                neutral_mentions,
                negative_mentions,
                total_mentions,
                total_queries,
                total_competitor_mentions,
                weighted_competitor_breakdown,
                appeared_in_top3,
                appeared_in_top5,
                appeared_in_top10,
                snapshot_date
            FROM llm_monitoring_snapshots
            WHERE project_id = %s
                AND snapshot_date >= CURRENT_DATE - (%s * INTERVAL '1 day')
                AND snapshot_date < CURRENT_DATE - (%s * INTERVAL '1 day')
        """
        previous_snapshots_params = [project_id, days * 2, days]
        if enabled_llms_filter:
            previous_snapshots_query += " AND llm_provider = ANY(%s)"
            previous_snapshots_params.append(enabled_llms_filter)
        previous_snapshots_query += " ORDER BY snapshot_date DESC, llm_provider"
        cur.execute(previous_snapshots_query, previous_snapshots_params)
        
        previous_snapshots = cur.fetchall()
        logger.info(f"📊 Snapshots período anterior: {len(previous_snapshots)} (para calcular tendencias)")
        
        # 🧮 Calcular MÉTRICAS AGREGADAS (volumen real)
        # En lugar de promediar SoV por LLM, sumamos TODAS las menciones
        
        # Totales agregados - PERÍODO ACTUAL
        total_brand_mentions = 0
        total_competitor_mentions = 0
        total_queries_all = 0
        total_positive = 0
        total_neutral = 0
        total_negative = 0
        all_positions = []
        
        # ✨ NUEVO: Totales ponderados para Share of Voice
        total_brand_mentions_weighted = 0.0
        total_competitor_mentions_weighted = 0.0
        
        # ✨ NUEVO: Métricas de posición granulares
        total_top3 = 0
        total_top5 = 0
        total_top10 = 0
        
        # Agrupar por LLM para cálculos individuales
        snapshots_by_llm = {}
        def _parse_weighted_breakdown(raw_value):
            if isinstance(raw_value, dict):
                return raw_value
            if isinstance(raw_value, str):
                try:
                    parsed = json.loads(raw_value)
                    return parsed if isinstance(parsed, dict) else {}
                except Exception:
                    return {}
            return {}
        
        def _accumulate_weighted_totals(snapshot):
            nonlocal total_brand_mentions_weighted, total_competitor_mentions_weighted
            
            weighted_sov_raw = snapshot.get('weighted_share_of_voice')
            try:
                weighted_sov = float(weighted_sov_raw) if weighted_sov_raw is not None else None
            except Exception:
                weighted_sov = None
            
            weighted_breakdown = _parse_weighted_breakdown(snapshot.get('weighted_competitor_breakdown'))
            
            if weighted_sov is not None and weighted_breakdown:
                try:
                    total_weighted_comp = sum(float(v) for v in weighted_breakdown.values() if v is not None)
                except Exception:
                    total_weighted_comp = 0.0
                
                if weighted_sov >= 100:
                    weighted_brand = total_weighted_comp if total_weighted_comp > 0 else 1.0
                elif weighted_sov > 0:
                    weighted_brand = (weighted_sov / (100 - weighted_sov)) * total_weighted_comp
                else:
                    weighted_brand = 0.0
                
                total_brand_mentions_weighted += weighted_brand
                total_competitor_mentions_weighted += total_weighted_comp
            else:
                # Fallback a métricas normales si no hay datos ponderados
                total_brand_mentions_weighted += (snapshot.get('total_mentions') or 0)
                total_competitor_mentions_weighted += (snapshot.get('total_competitor_mentions') or 0)
        
        for snapshot in all_snapshots:
            llm = snapshot['llm_provider']
            if llm not in snapshots_by_llm:
                snapshots_by_llm[llm] = []
            snapshots_by_llm[llm].append(snapshot)
            
            # Acumular totales agregados
            total_brand_mentions += (snapshot.get('total_mentions') or 0)
            total_competitor_mentions += (snapshot.get('total_competitor_mentions') or 0)
            total_queries_all += (snapshot.get('total_queries') or 0)
            total_positive += (snapshot.get('positive_mentions') or 0)
            total_neutral += (snapshot.get('neutral_mentions') or 0)
            total_negative += (snapshot.get('negative_mentions') or 0)
            
            # ✨ NUEVO: Acumular Top3/5/10
            total_top3 += (snapshot.get('appeared_in_top3') or 0)
            total_top5 += (snapshot.get('appeared_in_top5') or 0)
            total_top10 += (snapshot.get('appeared_in_top10') or 0)
            
            # Acumular posiciones
            if snapshot.get('avg_position') is not None:
                all_positions.append(float(snapshot['avg_position']))
            
            # ✨ NUEVO: Acumular métricas ponderadas para SoV
            _accumulate_weighted_totals(snapshot)
        
        # ✨ NUEVO: Calcular métricas del PERÍODO ANTERIOR para tendencias
        prev_brand_mentions = 0
        prev_competitor_mentions = 0
        prev_queries_all = 0
        prev_positive = 0
        prev_neutral = 0
        prev_negative = 0
        prev_positions = []
        
        # ✨ NUEVO: Totales ponderados del período anterior
        prev_brand_mentions_weighted = 0.0
        prev_competitor_mentions_weighted = 0.0
        
        def _accumulate_prev_weighted_totals(snapshot):
            nonlocal prev_brand_mentions_weighted, prev_competitor_mentions_weighted
            
            weighted_sov_raw = snapshot.get('weighted_share_of_voice')
            try:
                weighted_sov = float(weighted_sov_raw) if weighted_sov_raw is not None else None
            except Exception:
                weighted_sov = None
            
            weighted_breakdown = _parse_weighted_breakdown(snapshot.get('weighted_competitor_breakdown'))
            
            if weighted_sov is not None and weighted_breakdown:
                try:
                    total_weighted_comp = sum(float(v) for v in weighted_breakdown.values() if v is not None)
                except Exception:
                    total_weighted_comp = 0.0
                
                if weighted_sov >= 100:
                    weighted_brand = total_weighted_comp if total_weighted_comp > 0 else 1.0
                elif weighted_sov > 0:
                    weighted_brand = (weighted_sov / (100 - weighted_sov)) * total_weighted_comp
                else:
                    weighted_brand = 0.0
                
                prev_brand_mentions_weighted += weighted_brand
                prev_competitor_mentions_weighted += total_weighted_comp
            else:
                # Fallback a métricas normales si no hay datos ponderados
                prev_brand_mentions_weighted += (snapshot.get('total_mentions') or 0)
                prev_competitor_mentions_weighted += (snapshot.get('total_competitor_mentions') or 0)
        
        for snapshot in previous_snapshots:
            prev_brand_mentions += (snapshot.get('total_mentions') or 0)
            prev_competitor_mentions += (snapshot.get('total_competitor_mentions') or 0)
            prev_queries_all += (snapshot.get('total_queries') or 0)
            prev_positive += (snapshot.get('positive_mentions') or 0)
            prev_neutral += (snapshot.get('neutral_mentions') or 0)
            prev_negative += (snapshot.get('negative_mentions') or 0)
            if snapshot.get('avg_position') is not None:
                prev_positions.append(float(snapshot['avg_position']))
            
            # ✨ NUEVO: Acumular métricas ponderadas para SoV (periodo anterior)
            _accumulate_prev_weighted_totals(snapshot)
        
        # Métricas del período anterior
        prev_mention_rate = (prev_brand_mentions / prev_queries_all * 100) if prev_queries_all > 0 else 0
        if metric_type == 'weighted':
            prev_sov = (
                (prev_brand_mentions_weighted / (prev_brand_mentions_weighted + prev_competitor_mentions_weighted) * 100)
                if (prev_brand_mentions_weighted + prev_competitor_mentions_weighted) > 0 else 0
            )
        else:
            prev_sov = (
                (prev_brand_mentions / (prev_brand_mentions + prev_competitor_mentions) * 100)
                if (prev_brand_mentions + prev_competitor_mentions) > 0 else 0
            )
        prev_positive_pct = (prev_positive / prev_queries_all * 100) if prev_queries_all > 0 else 0
        
        # 📊 Calcular métricas agregadas globales
        aggregated_mention_rate = (total_brand_mentions / total_queries_all * 100) if total_queries_all > 0 else 0
        if metric_type == 'weighted':
            aggregated_sov = (
                (total_brand_mentions_weighted / (total_brand_mentions_weighted + total_competitor_mentions_weighted) * 100)
                if (total_brand_mentions_weighted + total_competitor_mentions_weighted) > 0 else 0
            )
        else:
            aggregated_sov = (
                (total_brand_mentions / (total_brand_mentions + total_competitor_mentions) * 100)
                if (total_brand_mentions + total_competitor_mentions) > 0 else 0
            )
        aggregated_avg_position = sum(all_positions) / len(all_positions) if all_positions else None
        
        # ✨ NUEVO: Métricas de posición desde resultados (consistencia con Responses Inspector)
        results_total_appearances = None
        results_avg_position = None
        results_top3 = 0
        results_top5 = 0
        results_top10 = 0
        
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=days)
        position_query = """
            SELECT 
                COUNT(*) FILTER (WHERE brand_mentioned) as total_mentions,
                AVG(position_in_list) FILTER (
                    WHERE brand_mentioned 
                      AND position_in_list IS NOT NULL 
                      AND position_in_list <= 30
                ) as avg_position,
                COUNT(*) FILTER (
                    WHERE brand_mentioned 
                      AND position_in_list IS NOT NULL 
                      AND position_in_list <= 3
                ) as top3,
                COUNT(*) FILTER (
                    WHERE brand_mentioned 
                      AND position_in_list IS NOT NULL 
                      AND position_in_list <= 5
                ) as top5,
                COUNT(*) FILTER (
                    WHERE brand_mentioned 
                      AND position_in_list IS NOT NULL 
                      AND position_in_list <= 10
                ) as top10
            FROM llm_monitoring_results
            WHERE project_id = %s
              AND analysis_date >= %s
              AND analysis_date <= %s
        """
        position_params = [project_id, start_date, end_date]
        if enabled_llms_filter:
            position_query += " AND llm_provider = ANY(%s)"
            position_params.append(enabled_llms_filter)
        cur.execute(position_query, position_params)
        
        position_row = cur.fetchone()
        if position_row:
            results_total_appearances = position_row.get('total_mentions') or 0
            results_avg_position = position_row.get('avg_position')
            results_top3 = position_row.get('top3') or 0
            results_top5 = position_row.get('top5') or 0
            results_top10 = position_row.get('top10') or 0
        
        # ✨ NUEVO: Calcular tasas de posición (% sobre total de menciones donde aparecimos)
        # Nota: Top3 incluye Top1, Top5 incluye Top3, etc.
        if results_total_appearances is not None:
            total_appearances = results_total_appearances
            total_top3 = results_top3
            total_top5 = results_top5
            total_top10 = results_top10
            if results_avg_position is not None:
                aggregated_avg_position = float(results_avg_position)
        else:
            total_appearances = total_brand_mentions  # Usar menciones como base
        top3_rate = (total_top3 / total_appearances * 100) if total_appearances > 0 else 0
        top5_rate = (total_top5 / total_appearances * 100) if total_appearances > 0 else 0
        top10_rate = (total_top10 / total_appearances * 100) if total_appearances > 0 else 0
        
        # Sentiment agregado
        aggregated_positive_pct = (total_positive / total_queries_all * 100) if total_queries_all > 0 else 0
        aggregated_neutral_pct = (total_neutral / total_queries_all * 100) if total_queries_all > 0 else 0
        aggregated_negative_pct = (total_negative / total_queries_all * 100) if total_queries_all > 0 else 0
        
        has_previous_period_data = len(previous_snapshots) > 0 and prev_queries_all > 0

        # ✨ NUEVO: Calcular TENDENCIAS (cambio vs período anterior)
        # Alias local → función de módulo (extraída para reusar en /metrics, /comparison, exports)
        calculate_trend = _calculate_trend
        
        # ✨ Calcular tendencia del SENTIMIENTO de forma categórica
        def get_dominant_sentiment(positive, neutral, negative):
            """Determina el sentimiento dominante y devuelve un valor numérico para comparación"""
            if positive >= neutral and positive >= negative:
                return ('positive', 3)  # 3 = mejor
            elif neutral >= positive and neutral >= negative:
                return ('neutral', 2)   # 2 = medio
            else:
                return ('negative', 1)  # 1 = peor
        
        current_sentiment, current_sentiment_value = get_dominant_sentiment(
            aggregated_positive_pct, aggregated_neutral_pct, aggregated_negative_pct
        )
        
        # Sentimiento del período anterior
        prev_positive_pct_calc = (prev_positive / prev_queries_all * 100) if prev_queries_all > 0 else 0
        prev_neutral_pct_calc = (prev_neutral / prev_queries_all * 100) if prev_queries_all > 0 else 0
        prev_negative_pct_calc = (prev_negative / prev_queries_all * 100) if prev_queries_all > 0 else 0
        
        prev_sentiment, prev_sentiment_value = get_dominant_sentiment(
            prev_positive_pct_calc, prev_neutral_pct_calc, prev_negative_pct_calc
        )
        
        # Determinar la tendencia categórica del sentimiento
        if not has_previous_period_data:
            sentiment_trend = None
        elif current_sentiment_value > prev_sentiment_value:
            sentiment_trend = {'direction': 'better', 'previous': prev_sentiment}
        elif current_sentiment_value < prev_sentiment_value:
            sentiment_trend = {'direction': 'worse', 'previous': prev_sentiment}
        else:
            sentiment_trend = {'direction': 'same', 'previous': prev_sentiment}

        trends = {
            'mention_rate': calculate_trend(aggregated_mention_rate, prev_mention_rate, has_previous_period_data),
            'share_of_voice': calculate_trend(aggregated_sov, prev_sov, has_previous_period_data),
            'sentiment': sentiment_trend  # ✨ Ahora es categórico: better/worse/same
        }
        
        # 🎯 Calcular métricas individuales por LLM (para compatibilidad con frontend)
        # El frontend espera datos por LLM, pero ahora usamos los valores agregados
        metrics_by_llm = {}
        
        for llm, llm_snapshots in snapshots_by_llm.items():
            if not llm_snapshots:
                continue
            
            # Último snapshot para fecha de referencia
            latest_snapshot = llm_snapshots[0]
            
            # ✨ IMPORTANTE: Usar métricas AGREGADAS para TODO (consistencia total)
            # Método 2: SoV Agregado aplicado a TODAS las métricas
            metrics_by_llm[llm] = {
                'mention_rate': round(aggregated_mention_rate, 2),  # ✨ AGREGADO (consistente)
                'avg_position': round(aggregated_avg_position, 2) if aggregated_avg_position else None,  # ✨ AGREGADO
                'share_of_voice': round(aggregated_sov, 2),  # ✨ AGREGADO
                'sentiment': {
                    'positive': round(aggregated_positive_pct, 2),  # ✨ AGREGADO
                    'neutral': round(aggregated_neutral_pct, 2),    # ✨ AGREGADO
                    'negative': round(aggregated_negative_pct, 2)   # ✨ AGREGADO
                },
                'total_queries': latest_snapshot.get('total_queries'),
                'date': latest_snapshot['snapshot_date'].isoformat() if latest_snapshot['snapshot_date'] else None,
                'snapshots_count': len(llm_snapshots)
            }
        
        logger.info(
            f"✅ Métricas agregadas calculadas ({metric_type}): "
            f"SoV={aggregated_sov:.2f}%, Mention Rate={aggregated_mention_rate:.2f}%"
        )
        
        # ✨ NUEVO: Calcular Quality Score
        # Componentes del Quality Score:
        # 1. Completeness: ¿Cada LLM analizó todas las queries esperadas?
        # 2. Freshness: ¿Cuánto hace del último análisis?
        # 3. Coverage: ¿Qué % de LLMs tienen datos?
        enabled_llms = project['enabled_llms'] or []
        llms_expected = len(enabled_llms)
        
        # Completeness (0-100): promedio de completitud por LLM (queries analizadas / esperadas)
        # La completitud se calcula contra prompts realmente configurados en el proyecto.
        expected_queries = project.get('total_queries') or 0
        llm_completeness = {}
        
        total_analyzed_queries = 0
        total_expected_queries = expected_queries * llms_expected if expected_queries else 0
        
        for llm_name in enabled_llms:
            snapshots = snapshots_by_llm.get(llm_name, [])
            latest_snapshot = None
            if snapshots:
                latest_snapshot = max(
                    snapshots,
                    key=lambda s: s.get('snapshot_date') or datetime.min
                )
            analyzed = (latest_snapshot or {}).get('total_queries') or 0
            total_analyzed_queries += analyzed
            if expected_queries and expected_queries > 0:
                pct = min(100.0, (analyzed / expected_queries) * 100)
            else:
                pct = 0.0
            llm_completeness[llm_name] = {
                'analyzed': analyzed,
                'expected': expected_queries,
                'completeness_pct': round(pct, 1)
            }
        if total_expected_queries > 0:
            completeness = min(100.0, (total_analyzed_queries / total_expected_queries) * 100)
        else:
            completeness = 0
        
        # Coverage (0-100): % de LLMs con al menos un snapshot
        llms_with_data = sum(1 for llm in enabled_llms if llm in snapshots_by_llm)
        coverage = (llms_with_data / llms_expected * 100) if llms_expected > 0 else 0
        
        # Freshness (0-100): 100 si datos de hoy, decrece con el tiempo
        from datetime import date as date_type
        last_snapshot = project['last_snapshot_date']
        if last_snapshot:
            days_since_update = (date_type.today() - last_snapshot).days
            freshness = max(0, 100 - (days_since_update * 10))  # -10% por día
        else:
            freshness = 0
        
        # Quality Score final (promedio ponderado)
        quality_score = round((completeness * 0.5 + freshness * 0.3 + coverage * 0.2), 1)
        
        quality_data = {
            'score': quality_score,
            'components': {
                'completeness': round(completeness, 1),
                'freshness': round(freshness, 1),
                'coverage': round(coverage, 1)
            },
            'details': {
                'llms_with_data': llms_with_data,
                'llms_expected': llms_expected,
                'total_analyzed_queries': total_analyzed_queries,
                'total_expected_queries': total_expected_queries,
                'queries_by_llm': llm_completeness,
                'days_since_update': (date_type.today() - last_snapshot).days if last_snapshot else None,
                'total_snapshots_in_period': len(all_snapshots)
            }
        }
        
        permissions = get_project_permissions(user['id'], 'llm_monitoring', project_id)
        logger.info(f"✅ Preparando respuesta para proyecto {project_id}")
        return jsonify({
            'success': True,
            'project': {
                'id': project['id'],
                'name': project['name'],
                'brand_name': project['brand_name'],
                'brand_domain': project.get('brand_domain'),
                'brand_keywords': project.get('brand_keywords', []),
                'industry': project['industry'],
                'enabled_llms': project['enabled_llms'],
                'competitors': project['competitors'],  # Legacy
                'competitor_domains': project.get('competitor_domains', []),  # Legacy
                'competitor_keywords': project.get('competitor_keywords', []),  # Legacy
                'selected_competitors': project.get('selected_competitors', []),  # ✨ NUEVO
                'language': project['language'],
                'country_code': project.get('country_code'),
                'queries_per_llm': project['queries_per_llm'],
                'is_active': project['is_active'],
                'last_analysis_date': project['last_analysis_date'].isoformat() if project['last_analysis_date'] else None,
                'created_at': project['created_at'].isoformat() if project['created_at'] else None,
                'updated_at': project['updated_at'].isoformat() if project['updated_at'] else None,
                'total_queries': project['total_queries'],
                'total_snapshots': project['total_snapshots'],
                'last_snapshot_date': project['last_snapshot_date'].isoformat() if project['last_snapshot_date'] else None,
                'access_role': permissions.get('access_role'),
                'is_owner': permissions.get('is_owner', False),
                'can_edit': permissions.get('can_edit', False),
                'can_manage_access': permissions.get('can_manage_access', False),
            },
            'latest_metrics': metrics_by_llm,
            # ✨ NUEVO: Tendencias (comparación con período anterior)
            'trends': trends,
            # ✨ NUEVO: Métricas de posición granulares
            'position_metrics': {
                'avg_position': round(aggregated_avg_position, 1) if aggregated_avg_position else None,
                'top3_rate': round(top3_rate, 1),
                'top5_rate': round(top5_rate, 1),
                'top10_rate': round(top10_rate, 1),
                'total_top3': total_top3,
                'total_top5': total_top5,
                'total_top10': total_top10,
                'total_appearances': total_appearances
            },
            # ✨ NUEVO: Quality Score del análisis
            'quality_score': quality_data,
            'model_scope_notice': model_scope_notice
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo proyecto: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo proyecto: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>', methods=['PUT'])
@login_required
@validate_project_ownership
def update_project(project_id):
    """
    Actualiza un proyecto existente
    
    Body (todos opcionales):
    {
        "name": "Nuevo Nombre",
        "industry": "Nueva Industria",
        "brand_domain": "newdomain.com",
        "brand_keywords": ["keyword1", "keyword2"],
        "competitor_domains": ["comp1.com"],
        "competitor_keywords": ["comp_kw1", "comp_kw2"],
        "is_active": false,
        "enabled_llms": ["openai", "google"]
    }
    
    Returns:
        JSON con el proyecto actualizado
    """
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'Body vacío'}), 400

    user = get_current_user()
    if not user:
        return jsonify({'error': 'Usuario no autenticado'}), 401
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT enabled_llms
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        current_project_row = cur.fetchone()
        if not current_project_row:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        previous_enabled_llms = current_project_row.get('enabled_llms') or []
        
        # Campos actualizables
        updates = []
        params = []
        
        if 'name' in data:
            updates.append("name = %s")
            params.append(data['name'])
        
        if 'industry' in data:
            updates.append("industry = %s")
            params.append(data['industry'])

        if 'language' in data:
            language = str(data.get('language') or '').strip().lower()
            if not language:
                return jsonify({'error': 'language no puede estar vacío'}), 400
            updates.append("language = %s")
            params.append(language)
        
        if 'brand_domain' in data:
            updates.append("brand_domain = %s")
            params.append(data['brand_domain'])
        
        if 'brand_keywords' in data:
            if not isinstance(data['brand_keywords'], list) or len(data['brand_keywords']) == 0:
                return jsonify({'error': 'brand_keywords debe ser un array con al menos 1 palabra clave'}), 400
            updates.append("brand_keywords = %s::jsonb")
            params.append(json.dumps(data['brand_keywords']))
            # Actualizar también brand_name legacy
            updates.append("brand_name = %s")
            params.append(data['brand_keywords'][0])
        
        # ✨ NEW: Handle selected_competitors (and extract legacy fields)
        if 'selected_competitors' in data:
            selected_competitors = data['selected_competitors']
            updates.append("selected_competitors = %s::jsonb")
            params.append(json.dumps(selected_competitors))
            
            # Extract legacy fields for backward compatibility
            competitor_domains = []
            competitor_keywords = []
            if selected_competitors:
                for comp in selected_competitors:
                    if comp.get('domain'):
                        competitor_domains.append(comp['domain'])
                    if comp.get('keywords'):
                        competitor_keywords.extend(comp['keywords'])
            
            # Update legacy fields
            updates.append("competitor_domains = %s::jsonb")
            params.append(json.dumps(competitor_domains))
            updates.append("competitor_keywords = %s::jsonb")
            params.append(json.dumps(competitor_keywords))
            updates.append("competitors = %s::jsonb")
            params.append(json.dumps(competitor_keywords))
        
        if 'is_active' in data:
            updates.append("is_active = %s")
            params.append(data['is_active'])
        
        if 'enabled_llms' in data:
            # Validar LLMs
            valid_llms = ['openai', 'anthropic', 'google', 'perplexity']
            if not isinstance(data['enabled_llms'], list) or len(data['enabled_llms']) == 0:
                return jsonify({'error': 'enabled_llms debe ser un array con al menos 1 LLM'}), 400
            if not all(llm in valid_llms for llm in data['enabled_llms']):
                return jsonify({'error': f'LLMs válidos: {valid_llms}'}), 400
            updates.append("enabled_llms = %s")
            params.append(data['enabled_llms'])
        
        if 'country_code' in data:
            country_code = str(data.get('country_code') or '').strip().upper()
            if not country_code or len(country_code) != 2 or not country_code.isalpha():
                return jsonify({'error': 'country_code debe ser un código ISO de 2 letras'}), 400
            updates.append("country_code = %s")
            params.append(country_code)
        
        if not updates:
            return jsonify({'error': 'No hay campos para actualizar'}), 400
        
        # Actualizar proyecto
        updates.append("updated_at = NOW()")
        params.append(project_id)
        
        query = f"""
            UPDATE llm_monitoring_projects
            SET {', '.join(updates)}
            WHERE id = %s
            RETURNING *
        """
        
        cur.execute(query, params)
        project = cur.fetchone()
        
        conn.commit()

        updated_enabled_llms = project.get('enabled_llms') or []
        llm_changes = {
            'previous_enabled_llms': previous_enabled_llms,
            'current_enabled_llms': updated_enabled_llms,
            'added_llms': sorted(list(set(updated_enabled_llms) - set(previous_enabled_llms))),
            'removed_llms': sorted(list(set(previous_enabled_llms) - set(updated_enabled_llms)))
        }
        model_selection_changed = (
            'enabled_llms' in data and
            (len(llm_changes['added_llms']) > 0 or len(llm_changes['removed_llms']) > 0)
        )
        
        return jsonify({
            'success': True,
            'model_selection_changed': model_selection_changed,
            'llm_changes': llm_changes if model_selection_changed else None,
            'project': {
                'id': project['id'],
                'name': project['name'],
                'brand_name': project['brand_name'],
                'brand_domain': project['brand_domain'],
                'brand_keywords': project['brand_keywords'],
                'industry': project['industry'],
                'enabled_llms': project['enabled_llms'],
                'competitors': project['competitors'],
                'competitor_domains': project['competitor_domains'],
                'competitor_keywords': project['competitor_keywords'],
                'language': project['language'],
                'country_code': project['country_code'],
                'queries_per_llm': project['queries_per_llm'],
                'is_active': project['is_active'],
                'updated_at': project['updated_at'].isoformat() if project['updated_at'] else None
            }
        }), 200
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error actualizando proyecto: {e}", exc_info=True)
        return jsonify({'error': f'Error actualizando proyecto: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>/deactivate', methods=['PUT'])
@login_required
@validate_project_ownership
def deactivate_project(project_id):
    """
    Desactiva un proyecto (marca is_active = false)
    El proyecto deja de ejecutarse en el CRON diario pero mantiene sus datos
    
    Returns:
        JSON con confirmación
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        # Marcar como inactivo
        cur.execute("""
            UPDATE llm_monitoring_projects
            SET is_active = FALSE, updated_at = NOW()
            WHERE id = %s AND is_active = TRUE
            RETURNING id, name
        """, (project_id,))
        
        project = cur.fetchone()
        
        if not project:
            return jsonify({'error': 'Proyecto no encontrado o ya está inactivo'}), 404
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': f'Proyecto "{project["name"]}" desactivado. Ya no se ejecutará en análisis automáticos.',
            'project_id': project['id']
        }), 200
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error desactivando proyecto: {e}", exc_info=True)
        return jsonify({'error': f'Error desactivando proyecto: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>/activate', methods=['PUT'])
@login_required
@validate_project_ownership
def activate_project(project_id):
    """
    Reactiva un proyecto inactivo (marca is_active = true)
    
    Returns:
        JSON con confirmación
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Usuario no autenticado'}), 401
    plan_limits = _get_effective_plan_limits(user)
    max_projects = plan_limits.get('max_projects')
    active_projects = count_user_active_projects(user['id'])
    if max_projects is not None and active_projects >= max_projects:
        return jsonify({
            'error': 'project_limit_reached',
            'message': 'Has alcanzado el máximo de proyectos permitidos para tu plan',
            'current_plan': user.get('plan', 'free'),
            'upgrade_options': get_upgrade_options(user.get('plan', 'free')),
            'limit': max_projects,
            'current': active_projects
        }), 402

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        # Marcar como activo
        cur.execute("""
            UPDATE llm_monitoring_projects
            SET is_active = TRUE, updated_at = NOW()
            WHERE id = %s AND is_active = FALSE
            RETURNING id, name
        """, (project_id,))
        
        project = cur.fetchone()
        
        if not project:
            return jsonify({'error': 'Proyecto no encontrado o ya está activo'}), 404
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': f'Proyecto "{project["name"]}" reactivado. Se incluirá en próximos análisis automáticos.',
            'project_id': project['id']
        }), 200
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error reactivando proyecto: {e}", exc_info=True)
        return jsonify({'error': f'Error reactivando proyecto: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>', methods=['DELETE'])
@login_required
@validate_project_ownership
def delete_project(project_id):
    """
    Elimina DEFINITIVAMENTE un proyecto (hard delete)
    SOLO funciona si el proyecto está INACTIVO
    
    Elimina:
    - El proyecto
    - Todas sus queries
    - Todos los resultados de análisis
    - Todos los snapshots
    
    Returns:
        JSON con confirmación
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        # Verificar que el proyecto esté inactivo
        cur.execute("""
            SELECT id, name, is_active 
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        
        project = cur.fetchone()
        
        if not project:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        
        if project['is_active']:
            return jsonify({
                'error': 'No se puede eliminar un proyecto activo. Desactívalo primero.',
                'action_required': 'deactivate_first'
            }), 400
        
        project_name = project['name']
        
        # Eliminar en cascada (orden importante para foreign keys)
        # 1. Snapshots
        cur.execute("DELETE FROM llm_monitoring_snapshots WHERE project_id = %s", (project_id,))
        snapshots_deleted = cur.rowcount
        
        # 2. Resultados
        cur.execute("DELETE FROM llm_monitoring_results WHERE project_id = %s", (project_id,))
        results_deleted = cur.rowcount
        
        # 3. Queries
        cur.execute("DELETE FROM llm_monitoring_queries WHERE project_id = %s", (project_id,))
        queries_deleted = cur.rowcount
        
        # 4. Proyecto
        cur.execute("DELETE FROM llm_monitoring_projects WHERE id = %s", (project_id,))
        
        conn.commit()
        
        logger.info(f"🗑️ Proyecto '{project_name}' eliminado definitivamente:")
        logger.info(f"   - Queries: {queries_deleted}")
        logger.info(f"   - Resultados: {results_deleted}")
        logger.info(f"   - Snapshots: {snapshots_deleted}")
        
        return jsonify({
            'success': True,
            'message': f'Proyecto "{project_name}" eliminado definitivamente',
            'project_id': project_id,
            'stats': {
                'queries_deleted': queries_deleted,
                'results_deleted': results_deleted,
                'snapshots_deleted': snapshots_deleted
            }
        }), 200
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error eliminando proyecto: {e}", exc_info=True)
        return jsonify({'error': f'Error eliminando proyecto: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


# ============================================================================
# ENDPOINTS: PROMPTS/QUERIES (Manual Management)
# ============================================================================
@llm_monitoring_bp.route('/projects/<int:project_id>/queries', methods=['POST'])
@login_required
@validate_project_ownership
def add_queries_to_project(project_id):
    """
    Añade queries/prompts manualmente a un proyecto
    
    Body esperado:
    {
        "queries": ["¿Qué es X?", "¿Cómo funciona Y?", ...],
        "language": "es" (opcional, default del proyecto),
        "query_type": "manual" (opcional, default: "manual")
    }
    
    Returns:
        JSON con resultado de la operación
    """
    user = get_current_user()
    
    data = request.get_json()
    queries_list = data.get('queries', [])
    language = data.get('language')
    query_type = data.get('query_type', 'manual')
    
    if not queries_list:
        return jsonify({'error': 'No se proporcionaron queries'}), 400
    
    if not isinstance(queries_list, list):
        return jsonify({'error': 'queries debe ser una lista'}), 400
    
    # Validar límites por plan (prompts por proyecto)
    plan_limits = _get_effective_plan_limits(user)
    max_prompts = plan_limits.get('max_prompts_per_project')

    # Obtener configuración del proyecto si no se especificó idioma
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()

        # Lock del proyecto para evitar carreras al añadir prompts
        cur.execute("SELECT id FROM llm_monitoring_projects WHERE id = %s FOR UPDATE", (project_id,))

        if max_prompts is not None:
            cur.execute("""
                SELECT COUNT(*) AS count
                FROM llm_monitoring_queries
                WHERE project_id = %s AND is_active = TRUE
            """, (project_id,))
            row = cur.fetchone()
            current_count = int(row['count']) if row else 0
            incoming_count = len([q for q in queries_list if isinstance(q, str) and q.strip()])
            if current_count + incoming_count > max_prompts:
                return jsonify({
                    'error': 'prompt_limit_exceeded',
                    'message': 'Has alcanzado el máximo de prompts permitidos para este proyecto',
                    'current_plan': user.get('plan', 'free'),
                    'upgrade_options': get_upgrade_options(user.get('plan', 'free')),
                    'limit': max_prompts,
                    'current': current_count,
                    'requested': incoming_count
                }), 402
        
        # Si no se especificó idioma, usar el del proyecto
        if not language:
            cur.execute("SELECT language FROM llm_monitoring_projects WHERE id = %s", (project_id,))
            project = cur.fetchone()
            if project:
                language = project['language']
            else:
                language = 'es'
        
        added_count = 0
        duplicate_count = 0
        error_count = 0
        
        for query_text in queries_list:
            query_text = query_text.strip()
            if not query_text:
                error_count += 1
                continue
            
            try:
                cur.execute("""
                    INSERT INTO llm_monitoring_queries (
                        project_id, query_text, language, query_type, is_active, added_at
                    ) VALUES (%s, %s, %s, %s, TRUE, NOW())
                    ON CONFLICT (project_id, query_text) DO NOTHING
                """, (project_id, query_text, language, query_type))
                
                if cur.rowcount > 0:
                    added_count += 1
                else:
                    duplicate_count += 1
                    
            except Exception as e:
                logger.warning(f"Error añadiendo query '{query_text}': {e}")
                error_count += 1
                continue
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'added_count': added_count,
            'duplicate_count': duplicate_count,
            'error_count': error_count,
            'message': f'{added_count} queries añadidas exitosamente'
        }), 200
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error añadiendo queries: {e}", exc_info=True)
        return jsonify({'error': f'Error añadiendo queries: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>/queries/<int:query_id>', methods=['DELETE'])
@login_required
@validate_project_ownership
def delete_query(project_id, query_id):
    """
    Elimina una query de un proyecto (soft delete: marca is_active = false)
    
    Returns:
        JSON con confirmación
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        # Verificar que la query pertenece al proyecto
        cur.execute("""
            SELECT id, query_text
            FROM llm_monitoring_queries
            WHERE id = %s AND project_id = %s
        """, (query_id, project_id))
        
        query = cur.fetchone()
        
        if not query:
            return jsonify({'error': 'Query no encontrada'}), 404
        
        # Soft delete
        cur.execute("""
            UPDATE llm_monitoring_queries
            SET is_active = FALSE
            WHERE id = %s AND project_id = %s
            RETURNING id
        """, (query_id, project_id))
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Query eliminada exitosamente',
            'query_id': query_id
        }), 200
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error eliminando query: {e}", exc_info=True)
        return jsonify({'error': f'Error eliminando query: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>/queries/<int:query_id>/history', methods=['GET'])
@login_required
@validate_project_ownership
def get_query_history(project_id, query_id):
    """
    ✨ NUEVO: Obtiene el historial de visibilidad de un prompt/query específico
    para mostrar la evolución temporal en una gráfica.
    
    Query params:
        - days: Número de días de historial (default: 30, usa el time range global)
    
    Returns:
        JSON con historial de menciones por fecha y LLM
    """
    # ✨ Obtener parámetro days del time range global (normalizado como el resto)
    days = _normalize_days_param(request.args.get('days'), default=30)
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    cur = None
    try:
        cur = conn.cursor()
        
        # Verificar que la query pertenece al proyecto
        cur.execute("""
            SELECT q.id, q.query_text, p.enabled_llms
            FROM llm_monitoring_queries q
            JOIN llm_monitoring_projects p ON q.project_id = p.id
            WHERE q.id = %s AND q.project_id = %s
        """, (query_id, project_id))
        
        query = cur.fetchone()
        
        if not query:
            return jsonify({'error': 'Query no encontrada', 'success': False}), 404
        
        logger.info(f"📊 Fetching history for query {query_id} ('{query['query_text'][:30]}...') - last {days} days")
        
        # Obtener historial de resultados para esta query
        enabled_llms_filter = query.get('enabled_llms') or []
        history_query = """
            SELECT
                analysis_date,
                llm_provider,
                brand_mentioned,
                position_in_list,
                sentiment
            FROM llm_monitoring_results
            WHERE query_id = %s AND project_id = %s
                AND analysis_date >= CURRENT_DATE - (%s * INTERVAL '1 day')
        """
        history_params = [query_id, project_id, days]
        if enabled_llms_filter:
            history_query += " AND llm_provider = ANY(%s)"
            history_params.append(enabled_llms_filter)
        history_query += " ORDER BY analysis_date ASC, llm_provider"
        cur.execute(history_query, history_params)
        
        results = cur.fetchall()
        
        logger.info(f"   → Found {len(results)} result records")
        
        # Si no hay resultados, retornar lista vacía con éxito
        if not results:
            return jsonify({
                'success': True,
                'query_id': query_id,
                'query_text': query['query_text'],
                'history': [],
                'llm_providers': [],
                'total_data_points': 0,
                'days': days,
                'message': 'No historical data found for this query in the selected period'
            }), 200
        
        # Agrupar resultados por fecha
        history_by_date = {}
        llm_providers_set = set()
        
        for row in results:
            date_str = row['analysis_date'].isoformat() if row['analysis_date'] else None
            if not date_str:
                continue
                
            llm = row['llm_provider']
            llm_providers_set.add(llm)
            
            if date_str not in history_by_date:
                history_by_date[date_str] = {
                    'date': date_str,
                    'total_llms': 0,
                    'llms_mentioned': 0,
                    'visibility_rate': 0,
                    'by_llm': {}
                }
            
            history_by_date[date_str]['total_llms'] += 1
            if row['brand_mentioned']:
                history_by_date[date_str]['llms_mentioned'] += 1
            
            history_by_date[date_str]['by_llm'][llm] = {
                'mentioned': row['brand_mentioned'] or False,
                'position': row['position_in_list'],
                'sentiment': row['sentiment']
            }
        
        # Calcular visibility_rate por fecha
        for date_str, data in history_by_date.items():
            if data['total_llms'] > 0:
                data['visibility_rate'] = round(
                    (data['llms_mentioned'] / data['total_llms']) * 100, 1
                )
        
        # Convertir a lista ordenada por fecha
        history_list = sorted(
            list(history_by_date.values()),
            key=lambda x: x['date']
        )
        
        # Obtener lista de LLMs únicos para la leyenda del gráfico
        llm_providers = sorted(list(llm_providers_set))
        
        logger.info(f"   ✅ Returning {len(history_list)} data points for {len(llm_providers)} LLMs")
        
        return jsonify({
            'success': True,
            'query_id': query_id,
            'query_text': query['query_text'],
            'history': history_list,
            'llm_providers': llm_providers,
            'total_data_points': len(history_list),
            'days': days
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo historial de query: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo historial: {str(e)}', 'success': False}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>/queries/suggest', methods=['POST'])
@login_required
@validate_project_ownership
def suggest_queries(project_id):
    """
    Genera sugerencias de queries usando IA (Gemini Flash)
    
    Analiza los prompts existentes del proyecto y el contexto (marca, industria)
    para sugerir prompts adicionales relevantes usando Gemini Flash.
    
    Body opcional:
    {
        "count": 10  (número de sugerencias, default: 10, max: 20)
    }
    
    Returns:
        JSON con lista de sugerencias generadas por IA
    """
    user = get_current_user()
    
    data = request.get_json() or {}
    count = min(data.get('count', 10), 20)  # Máximo 20 sugerencias
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        # Obtener datos del proyecto
        cur.execute("""
            SELECT name, brand_name, industry, language, competitors
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        
        project = cur.fetchone()
        
        if not project:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        
        # Obtener queries existentes
        cur.execute("""
            SELECT query_text
            FROM llm_monitoring_queries
            WHERE project_id = %s AND is_active = TRUE
            ORDER BY added_at DESC
            LIMIT 20
        """, (project_id,))
        
        existing_queries = cur.fetchall()
        existing_queries_list = [q['query_text'] for q in existing_queries]
        
        cur.close()
        conn.close()
        
        # Generar sugerencias usando IA
        from services.llm_monitoring_service import generate_query_suggestions_with_ai
        
        logger.info(f"🤖 Generando sugerencias para proyecto {project_id}: {project['brand_name']}")
        logger.info(f"   - Industria: {project['industry']}")
        logger.info(f"   - Queries existentes: {len(existing_queries_list)}")
        logger.info(f"   - Competidores: {project['competitors']}")
        
        suggestions = generate_query_suggestions_with_ai(
            brand_name=project['brand_name'],
            industry=project['industry'],
            language=project['language'],
            existing_queries=existing_queries_list,
            competitors=project['competitors'] or [],
            count=count
        )
        
        if not suggestions:
            logger.warning(f"⚠️ No se generaron sugerencias para proyecto {project_id}")
            # Verificar si es por falta de API key
            import os
            if not os.getenv('GOOGLE_API_KEY'):
                return jsonify({
                    'success': False,
                    'error': 'GOOGLE_API_KEY no está configurada en el servidor',
                    'hint': 'Contacta al administrador para configurar las API keys'
                }), 500
            else:
                return jsonify({
                    'success': False,
                    'error': 'No se pudieron generar sugerencias. Es posible que Gemini API esté teniendo problemas.',
                    'hint': 'Intenta de nuevo en unos momentos'
                }), 500
        
        return jsonify({
            'success': True,
            'suggestions': suggestions,
            'count': len(suggestions),
            'message': f'{len(suggestions)} sugerencias generadas por IA'
        }), 200
        
    except Exception as e:
        logger.error(f"Error generando sugerencias: {e}", exc_info=True)
        return jsonify({
            'error': f'Error generando sugerencias: {str(e)}',
            'hint': 'Verifica que GOOGLE_API_KEY esté configurada'
        }), 500


@llm_monitoring_bp.route('/projects/<int:project_id>/queries/suggest-variations', methods=['POST'])
@login_required
@validate_project_ownership
def suggest_query_variations(project_id):
    """
    ✨ NUEVO: Genera variaciones rápidas de prompts existentes usando IA
    
    Body:
    {
        "existing_prompts": ["prompt1", "prompt2"],
        "count": 6
    }
    
    Returns:
        JSON con lista de variaciones sugeridas
    """
    data = request.get_json() or {}
    existing_prompts = data.get('existing_prompts', [])
    count = min(data.get('count', 6), 10)
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        # Obtener datos del proyecto
        cur.execute("""
            SELECT brand_name, industry, language, competitors
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        
        project = cur.fetchone()
        
        if not project:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        
        cur.close()
        conn.close()
        
        brand_name = project['brand_name']
        industry = project['industry'] or 'general'
        competitors = project['competitors'] or []
        
        language = (project['language'] or 'en').lower()

        # Intentar generar con IA usando el provider configurado en BD (sin modelo hardcodeado)
        try:
            from services.llm_monitoring_service import generate_query_suggestions_with_ai

            suggestions = generate_query_suggestions_with_ai(
                brand_name=brand_name,
                industry=industry,
                language=language,
                existing_queries=existing_prompts,
                competitors=competitors,
                count=count
            )
            if suggestions:
                return jsonify({
                    'success': True,
                    'suggestions': suggestions,
                    'source': 'ai'
                }), 200

        except Exception as ai_error:
            logger.warning(f"AI generation failed, using fallback: {ai_error}")
        
        # Fallback: Generate simple variations locally with randomization
        import random
        
        competitor_fallback = {
            'es': 'competidores',
            'it': 'concorrenti',
            'fr': 'concurrents',
            'de': 'Wettbewerber',
            'pt': 'concorrentes',
        }
        comp_name = competitors[0] if competitors else competitor_fallback.get(language, 'competitors')
        
        if language == 'es':
            all_variations = [
                f"¿Qué es {brand_name} y cómo funciona?",
                f"Mejores herramientas de {industry}",
                f"{brand_name} vs {comp_name} comparativa",
                f"¿Vale la pena {brand_name}? Opiniones",
                f"Alternativas a {brand_name}",
                f"Cómo empezar con {brand_name}",
                f"Precios y planes de {brand_name}",
                f"Las mejores soluciones de {industry}",
                f"Opiniones sobre {brand_name}",
                f"¿Qué opinan de {brand_name}?",
                f"Ventajas y desventajas de {brand_name}",
                f"¿Recomiendan {brand_name}?",
                f"Tutorial de {brand_name}",
                f"Características de {brand_name}",
                f"¿Es bueno {brand_name}?",
                f"Empresas que usan {brand_name}",
                f"Comparativa de {industry}",
                f"Top {industry} en 2024",
                f"¿Cuál es mejor {brand_name} o {comp_name}?",
                f"Experiencias con {brand_name}"
            ]
        elif language == 'it':
            all_variations = [
                f"Cos'è {brand_name} e come funziona?",
                f"Migliori strumenti di {industry}",
                f"Confronto {brand_name} vs {comp_name}",
                f"{brand_name} vale la pena? Recensioni",
                f"Alternative a {brand_name}",
                f"Come iniziare con {brand_name}",
                f"Prezzi e piani di {brand_name}",
                f"Opinioni su {brand_name}",
                f"Pro e contro di {brand_name}",
                f"{brand_name} per principianti"
            ]
        elif language == 'fr':
            all_variations = [
                f"Qu'est-ce que {brand_name} et comment ça marche ?",
                f"Meilleurs outils de {industry}",
                f"Comparatif {brand_name} vs {comp_name}",
                f"{brand_name} vaut-il le coup ? Avis",
                f"Alternatives à {brand_name}",
                f"Comment démarrer avec {brand_name}",
                f"Tarifs et offres de {brand_name}",
                f"Avis sur {brand_name}",
                f"Avantages et inconvénients de {brand_name}",
                f"{brand_name} pour débutants"
            ]
        elif language == 'de':
            all_variations = [
                f"Was ist {brand_name} und wie funktioniert es?",
                f"Beste {industry}-Tools",
                f"{brand_name} vs {comp_name} Vergleich",
                f"Lohnt sich {brand_name}? Erfahrungen",
                f"Alternativen zu {brand_name}",
                f"Wie startet man mit {brand_name}?",
                f"Preise und Pakete von {brand_name}",
                f"Bewertungen zu {brand_name}",
                f"Vor- und Nachteile von {brand_name}",
                f"{brand_name} für Einsteiger"
            ]
        elif language == 'pt':
            all_variations = [
                f"O que é {brand_name} e como funciona?",
                f"Melhores ferramentas de {industry}",
                f"Comparativo {brand_name} vs {comp_name}",
                f"{brand_name} vale a pena? Avaliações",
                f"Alternativas ao {brand_name}",
                f"Como começar com {brand_name}",
                f"Preços e planos do {brand_name}",
                f"Opiniões sobre {brand_name}",
                f"Prós e contras do {brand_name}",
                f"{brand_name} para iniciantes"
            ]
        else:
            all_variations = [
                f"What is {brand_name} and how does it work?",
                f"Best {industry} tools and platforms",
                f"{brand_name} vs {comp_name} comparison",
                f"Is {brand_name} worth it? Reviews",
                f"Alternatives to {brand_name}",
                f"How to get started with {brand_name}",
                f"{brand_name} pricing and plans",
                f"Top rated {industry} solutions",
                f"{brand_name} reviews and opinions",
                f"Pros and cons of {brand_name}",
                f"Would you recommend {brand_name}?",
                f"{brand_name} tutorial",
                f"{brand_name} features",
                f"Is {brand_name} good?",
                f"Companies using {brand_name}",
                f"Best {industry} comparison",
                f"Top {industry} in 2024",
                f"Which is better {brand_name} or {comp_name}?",
                f"User experiences with {brand_name}",
                f"{brand_name} for beginners"
            ]
        
        # Shuffle and pick random variations
        random.shuffle(all_variations)
        variations = all_variations
        
        # Filter out existing prompts
        existing_lower = [p.lower() for p in existing_prompts]
        suggestions = [v for v in variations if v.lower() not in existing_lower][:count]
        
        return jsonify({
            'success': True,
            'suggestions': suggestions,
            'source': 'fallback'
        }), 200
        
    except Exception as e:
        logger.error(f"Error generating variations: {e}", exc_info=True)
        return jsonify({'error': str(e), 'success': False}), 500
    finally:
        pass  # Connection already closed


# ============================================================================
# ENDPOINTS: ANÁLISIS
# ============================================================================

# REMOVED: Manual analysis endpoint
#
# Razón: El sistema ahora funciona EXCLUSIVAMENTE con cron diario (4:00 AM).
# 
# El análisis manual fue eliminado porque:
# - Toma 15-30 minutos (timeout en navegador)
# - Prioriza completitud sobre velocidad
# - Requiere sistema robusto de reintentos
# - El cron diario garantiza 100% de completitud
#
# El endpoint /projects/<int:project_id>/analyze ya NO está disponible.
# Solo se permite /projects/<id>/run-initial-analysis (una vez) para evitar
# que un proyecto nuevo quede vacío hasta el siguiente cron diario.
#
# Para ejecutar análisis manual (admin/debugging):
# - Usar: python3 fix_openai_incomplete_analysis.py
# - O ejecutar manualmente: python3 daily_llm_monitoring_cron.py


@llm_monitoring_bp.route('/projects/<int:project_id>/run-initial-analysis', methods=['POST'])
@login_required
@validate_project_ownership
def run_initial_analysis(project_id):
    """
    Ejecuta el PRIMER análisis de un proyecto en background.
    Solo disponible para proyectos sin análisis previo.
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Usuario no autenticado'}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500

    cur = None
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT
                p.id,
                p.name,
                p.is_active,
                p.last_analysis_date,
                p.enabled_llms,
                p.queries_per_llm,
                (
                    SELECT COUNT(*)
                    FROM llm_monitoring_queries q
                    WHERE q.project_id = p.id AND q.is_active = TRUE
                ) as total_queries
            FROM llm_monitoring_projects p
            WHERE p.id = %s
        """, (project_id,))
        project = cur.fetchone()
        if not project:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        if not project.get('is_active'):
            return jsonify({
                'error': 'project_inactive',
                'message': 'Activa el proyecto antes de ejecutar el primer análisis'
            }), 400

        # "Solo primera vez": si ya hay snapshot o fecha de análisis, no permitir.
        cur.execute("""
            SELECT COUNT(*) AS count
            FROM llm_monitoring_snapshots
            WHERE project_id = %s
        """, (project_id,))
        row = cur.fetchone()
        has_snapshots = int(row['count']) > 0 if row else False

        if project.get('last_analysis_date') or has_snapshots:
            return jsonify({
                'error': 'initial_analysis_already_completed',
                'message': 'Este proyecto ya tiene su primer análisis completado'
            }), 409

        configured_queries = int(project.get('total_queries') or 0)
        if configured_queries <= 0:
            return jsonify({
                'error': 'no_active_queries',
                'message': 'Añade al menos un prompt activo antes de ejecutar el primer análisis'
            }), 400

        # Evitar doble click / doble thread del mismo proyecto.
        if not _mark_initial_analysis_running(project_id):
            return jsonify({
                'error': 'initial_analysis_in_progress',
                'message': 'El primer análisis ya está en curso para este proyecto'
            }), 409

        enabled_llms = project.get('enabled_llms') or []
        llm_count = max(1, len(enabled_llms))
        estimated_units = llm_count * configured_queries
        estimated_min_minutes = max(2, min(12, int(round(estimated_units / 24.0))))
        estimated_max_minutes = max(5, min(35, int(round(estimated_units / 8.0))))

        project_name = project.get('name') or f'Project #{project_id}'

        def run_first_analysis_in_background():
            try:
                logger.info(
                    f"🚀 Initial analysis started for project {project_id} ({project_name}) by user {user.get('id')}"
                )
                service = MultiLLMMonitoringService(api_keys=None)
                result = service.analyze_project(project_id=project_id, max_workers=8)

                if result.get('success'):
                    logger.info(
                        f"✅ Initial analysis finished for project {project_id}: "
                        f"{result.get('total_queries_executed', 0)} queries"
                    )
                else:
                    logger.warning(
                        f"⚠️ Initial analysis ended with warning for project {project_id}: "
                        f"{result.get('error', 'unknown_error')}"
                    )
            except Exception as e:
                logger.error(f"❌ Error running initial analysis for project {project_id}: {e}", exc_info=True)
            finally:
                _clear_initial_analysis_running(project_id)

        thread = threading.Thread(target=run_first_analysis_in_background, daemon=True)
        thread.start()

        return jsonify({
            'success': True,
            'project_id': project_id,
            'message': 'Initial analysis started in background',
            'estimated_minutes_min': estimated_min_minutes,
            'estimated_minutes_max': estimated_max_minutes
        }), 202

    except Exception as e:
        # Limpieza defensiva si algo falló antes de lanzar correctamente.
        _clear_initial_analysis_running(project_id)
        logger.error(f"Error triggering initial analysis for project {project_id}: {e}", exc_info=True)
        return jsonify({'error': f'Error iniciando análisis inicial: {str(e)}'}), 500
    finally:
        if cur:
            cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>/metrics', methods=['GET'])
@login_required
@validate_project_ownership
def get_project_metrics(project_id):
    """
    Obtiene métricas detalladas de un proyecto
    
    Query params opcionales:
        start_date: Fecha inicio (YYYY-MM-DD, default: último mes)
        end_date: Fecha fin (YYYY-MM-DD, default: hoy)
        llm_provider: Filtrar por LLM (openai, anthropic, google, perplexity)
    
    Returns:
        JSON con snapshots y métricas agregadas
    """
    # Parámetros de fecha
    raw_days = request.args.get('days')
    days = _normalize_days_param(raw_days, default=30)
    
    if raw_days is not None and raw_days != '':
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    else:
        end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
        start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        
    llm_provider = request.args.get('llm_provider')

    # ── Previous period dates (for period-over-period comparison) ──
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d') if isinstance(start_date, str) else start_date
    prev_end = start_date  # previous period ends where current starts
    prev_start = (start_date_obj - timedelta(days=days)).strftime('%Y-%m-%d')

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500

    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT enabled_llms, brand_keywords
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        project_row = cur.fetchone()
        if not project_row:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        enabled_llms_filter = project_row.get('enabled_llms') or []
        brand_keywords = project_row.get('brand_keywords') or []

        # Query base
        query = """
            SELECT
                s.*,
                p.brand_name,
                p.competitors
            FROM llm_monitoring_snapshots s
            JOIN llm_monitoring_projects p ON s.project_id = p.id
            WHERE s.project_id = %s
            AND s.snapshot_date >= %s
            AND s.snapshot_date <= %s
        """
        
        params = [project_id, start_date, end_date]
        
        # Filtro opcional por LLM
        if llm_provider:
            query += " AND s.llm_provider = %s"
            params.append(llm_provider)
        elif enabled_llms_filter:
            query += " AND s.llm_provider = ANY(%s)"
            params.append(enabled_llms_filter)
        
        query += " ORDER BY s.snapshot_date DESC, s.llm_provider"
        
        cur.execute(query, params)
        snapshots = cur.fetchall()
        
        # Formatear snapshots
        snapshots_list = []
        for s in snapshots:
            total_mentions = (s.get('total_mentions') or 0)
            pos_mentions = (s.get('positive_mentions') or 0)
            neu_mentions = (s.get('neutral_mentions') or 0)
            neg_mentions = (s.get('negative_mentions') or 0)

            positive_pct = (pos_mentions / total_mentions * 100) if total_mentions else 0
            neutral_pct = (neu_mentions / total_mentions * 100) if total_mentions else 0
            negative_pct = (neg_mentions / total_mentions * 100) if total_mentions else 0

            snapshots_list.append({
                'id': s['id'],
                'llm_provider': s['llm_provider'],
                'snapshot_date': s['snapshot_date'].isoformat() if s['snapshot_date'] else None,
                'mention_rate': float(s['mention_rate']) if s['mention_rate'] is not None else 0,
                'avg_position': float(s['avg_position']) if s['avg_position'] is not None else None,
                'top3_count': s.get('appeared_in_top3'),
                'top5_count': s.get('appeared_in_top5'),
                'top10_count': s.get('appeared_in_top10'),
                'share_of_voice': float(s['share_of_voice']) if s['share_of_voice'] is not None else 0,
                # ✨ NUEVO: Share of Voice ponderado por posición
                'weighted_share_of_voice': float(s.get('weighted_share_of_voice') or s.get('share_of_voice') or 0),
                # ➕ Exponer datos de competidores para el frontend (gráfico SOV)
                'total_competitor_mentions': int(s.get('total_competitor_mentions') or 0),
                'competitor_breakdown': s.get('competitor_breakdown') or {},
                'weighted_competitor_breakdown': s.get('weighted_competitor_breakdown') or {},
                'sentiment': {
                    'positive': float(positive_pct),
                    'neutral': float(neutral_pct),
                    'negative': float(negative_pct)
                },
                'total_queries': s.get('total_queries') or 0,
                'total_cost': float(s['total_cost_usd']) if s['total_cost_usd'] is not None else 0
            })
        
        # Calcular métricas agregadas
        total_cost = sum(s['total_cost'] for s in snapshots_list)
        total_queries = sum(s['total_queries'] for s in snapshots_list)
        
        # Promedio de mention_rate por LLM
        metrics_by_llm = {}
        for llm in ['openai', 'anthropic', 'google', 'perplexity']:
            llm_snapshots = [s for s in snapshots_list if s['llm_provider'] == llm]
            if llm_snapshots:
                metrics_by_llm[llm] = {
                    'avg_mention_rate': sum(s['mention_rate'] for s in llm_snapshots) / len(llm_snapshots),
                    'avg_position': sum(s['avg_position'] for s in llm_snapshots if s['avg_position']) / len([s for s in llm_snapshots if s['avg_position']]) if any(s['avg_position'] for s in llm_snapshots) else None,
                    'avg_share_of_voice': sum(s['share_of_voice'] for s in llm_snapshots) / len(llm_snapshots),
                    'total_snapshots': len(llm_snapshots)
                }
        
        # ── Branded vs Non-Branded breakdown ──
        branded_data = {'branded_metrics': None, 'non_branded_metrics': None}
        try:
            if brand_keywords:
                bq = """
                    SELECT r.llm_provider, q.query_text, r.brand_mentioned,
                           r.position_in_list, r.competitors_mentioned
                    FROM llm_monitoring_results r
                    JOIN llm_monitoring_queries q ON r.query_id = q.id
                    WHERE r.project_id = %s
                      AND r.analysis_date >= %s AND r.analysis_date <= %s
                """
                bparams = [project_id, start_date, end_date]
                if llm_provider:
                    bq += " AND r.llm_provider = %s"
                    bparams.append(llm_provider)
                elif enabled_llms_filter:
                    bq += " AND r.llm_provider = ANY(%s)"
                    bparams.append(enabled_llms_filter)
                cur.execute(bq, bparams)
                branded_data = _compute_branded_metrics(cur.fetchall(), brand_keywords)
        except Exception as be:
            logger.warning(f"Could not compute branded/non-branded metrics: {be}")

        # ── Previous period: metrics_by_llm ──
        previous_metrics_by_llm = {}
        try:
            prev_snap_query = """
                SELECT llm_provider, AVG(mention_rate) as avg_mention_rate
                FROM llm_monitoring_snapshots
                WHERE project_id = %s
                  AND snapshot_date >= %s AND snapshot_date < %s
            """
            prev_snap_params = [project_id, prev_start, prev_end]
            if llm_provider:
                prev_snap_query += " AND llm_provider = %s"
                prev_snap_params.append(llm_provider)
            elif enabled_llms_filter:
                prev_snap_query += " AND llm_provider = ANY(%s)"
                prev_snap_params.append(enabled_llms_filter)
            prev_snap_query += " GROUP BY llm_provider"
            cur.execute(prev_snap_query, prev_snap_params)
            for row in cur.fetchall():
                previous_metrics_by_llm[row['llm_provider']] = {
                    'avg_mention_rate': float(row['avg_mention_rate']) if row['avg_mention_rate'] is not None else 0
                }
        except Exception as prev_err:
            logger.warning(f"Could not compute previous metrics_by_llm: {prev_err}")

        # ── Previous period: branded / non-branded trends ──
        branded_trends = {}
        non_branded_trends = {}
        try:
            if brand_keywords:
                prev_bq = """
                    SELECT r.llm_provider, q.query_text, r.brand_mentioned,
                           r.position_in_list, r.competitors_mentioned
                    FROM llm_monitoring_results r
                    JOIN llm_monitoring_queries q ON r.query_id = q.id
                    WHERE r.project_id = %s
                      AND r.analysis_date >= %s AND r.analysis_date < %s
                """
                prev_bparams = [project_id, prev_start, prev_end]
                if llm_provider:
                    prev_bq += " AND r.llm_provider = %s"
                    prev_bparams.append(llm_provider)
                elif enabled_llms_filter:
                    prev_bq += " AND r.llm_provider = ANY(%s)"
                    prev_bparams.append(enabled_llms_filter)
                cur.execute(prev_bq, prev_bparams)
                prev_branded_data = _compute_branded_metrics(cur.fetchall(), brand_keywords)

                current_branded = branded_data.get('branded_metrics') or {}
                current_non_branded = branded_data.get('non_branded_metrics') or {}
                prev_branded = prev_branded_data.get('branded_metrics') or {}
                prev_non_branded = prev_branded_data.get('non_branded_metrics') or {}

                has_prev_branded = (prev_branded.get('total_results', 0) > 0)
                has_prev_non_branded = (prev_non_branded.get('total_results', 0) > 0)

                branded_trends = {
                    'mention_rate': _calculate_trend(
                        current_branded.get('mention_rate', 0),
                        prev_branded.get('mention_rate', 0),
                        has_prev_branded
                    ),
                    'share_of_voice': _calculate_trend(
                        current_branded.get('share_of_voice', 0),
                        prev_branded.get('share_of_voice', 0),
                        has_prev_branded
                    )
                }
                non_branded_trends = {
                    'mention_rate': _calculate_trend(
                        current_non_branded.get('mention_rate', 0),
                        prev_non_branded.get('mention_rate', 0),
                        has_prev_non_branded
                    ),
                    'share_of_voice': _calculate_trend(
                        current_non_branded.get('share_of_voice', 0),
                        prev_non_branded.get('share_of_voice', 0),
                        has_prev_non_branded
                    )
                }
        except Exception as bt_err:
            logger.warning(f"Could not compute branded trends: {bt_err}")

        return jsonify({
            'success': True,
            'project_id': project_id,
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'snapshots': snapshots_list,
            'aggregated': {
                'total_snapshots': len(snapshots_list),
                'total_cost_usd': total_cost,
                'total_queries_analyzed': total_queries,
                'metrics_by_llm': metrics_by_llm
            },
            'branded_metrics': branded_data.get('branded_metrics'),
            'non_branded_metrics': branded_data.get('non_branded_metrics'),
            'branded_trends': branded_trends,
            'non_branded_trends': non_branded_trends,
            'previous_metrics_by_llm': previous_metrics_by_llm
        }), 200

    except Exception as e:
        logger.error(f"Error obteniendo métricas: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo métricas: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/projects/<int:project_id>/urls-ranking', methods=['GET'])
@login_required
@validate_project_ownership
def get_urls_ranking(project_id):
    """
    Obtiene el ranking de URLs más mencionadas por los LLMs
    
    Query params opcionales:
        days: Número de días (default: 30)
        llm_provider: Filtrar por LLM específico ('openai', 'anthropic', 'google', 'perplexity')
        limit: Número máximo de URLs (default: 50). Usa 0 para "sin límite"
    
    Returns:
        JSON con ranking de URLs citadas por los LLMs
    """
    days = _normalize_days_param(request.args.get('days'), default=30)
    llm_provider = request.args.get('llm_provider')
    raw_limit = request.args.get('limit')
    limit = 50
    if raw_limit is not None and str(raw_limit).strip() != '':
        try:
            limit = int(raw_limit)
        except (TypeError, ValueError):
            return jsonify({'error': 'limit debe ser un número entero'}), 400
    
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'Error de conexión a BD'}), 500

        cur = conn.cursor()
        cur.execute("""
            SELECT enabled_llms
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        project = cur.fetchone()
        cur.close()
        conn.close()

        enabled_llms_filter = (project or {}).get('enabled_llms') or []
        if llm_provider and enabled_llms_filter and llm_provider not in enabled_llms_filter:
            return jsonify({
                'error': 'llm_provider no habilitado para este proyecto'
            }), 400

        if not llm_provider and not enabled_llms_filter:
            enabled_llms_filter = None

        urls_ranking = LLMMonitoringStatsService.get_project_urls_ranking(
            project_id=project_id,
            days=days,
            llm_provider=llm_provider,
            enabled_llms=enabled_llms_filter,
            limit=limit
        )
        
        return jsonify({
            'success': True,
            'project_id': project_id,
            'filters': {
                'days': days,
                'llm_provider': llm_provider or 'all',
                'limit': limit
            },
            'urls': urls_ranking,
            'total_urls': len(urls_ranking)
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo ranking de URLs: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo ranking de URLs: {str(e)}'}), 500


@llm_monitoring_bp.route('/projects/<int:project_id>/comparison', methods=['GET'])
@login_required
@validate_project_ownership
def get_llm_comparison(project_id):
    """
    Comparativa entre LLMs para un proyecto (usa vista llm_visibility_comparison)
    
    Query params opcionales:
        metric: 'weighted' o 'normal' (default: 'weighted')
    
    Returns:
        JSON con comparativa de métricas entre LLMs
    """
    # ✨ NUEVO: Parámetro de métrica para Share of Voice
    metric_type = request.args.get('metric', 'weighted')
    
    # ✨ NUEVO: Parámetro de días
    days = _normalize_days_param(request.args.get('days'), default=30)
    start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT enabled_llms, brand_domain
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        project_row = cur.fetchone()
        if not project_row:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        enabled_llms_filter = project_row.get('enabled_llms') or []
        brand_domain_raw = project_row.get('brand_domain') or ''

        def _normalize_domain(value):
            raw_value = str(value or '').strip().lower()
            if not raw_value:
                return ''
            if not raw_value.startswith(('http://', 'https://')):
                raw_value = f"https://{raw_value}"
            try:
                parsed = urlparse(raw_value)
            except Exception:
                return ''
            host = (parsed.netloc or parsed.path or '').split('/')[0].split(':')[0].strip().lower()
            if host.startswith('www.'):
                host = host[4:]
            return host

        def _extract_source_host(raw_url):
            raw_value = str(raw_url or '').strip()
            if not raw_value:
                return ''
            if not raw_value.startswith(('http://', 'https://')):
                raw_value = f"https://{raw_value}"
            try:
                parsed = urlparse(raw_value)
            except Exception:
                return ''
            host = (parsed.netloc or parsed.path or '').split('/')[0].split(':')[0].strip().lower()
            if host.startswith('www.'):
                host = host[4:]
            return host

        normalized_brand_domain = _normalize_domain(brand_domain_raw)

        def _to_date_key(value):
            if value is None:
                return None
            if isinstance(value, datetime):
                return value.date().isoformat()
            if hasattr(value, 'isoformat'):
                return value.isoformat()
            return str(value)
        
        # Traer filas por LLM directamente desde snapshots
        # ✨ NUEVO: Incluir weighted_share_of_voice
        comparison_query = """
            SELECT 
                llm_provider,
                snapshot_date,
                mention_rate,
                total_mentions,
                avg_position,
                share_of_voice,
                weighted_share_of_voice,
                avg_sentiment_score,
                positive_mentions,
                neutral_mentions,
                negative_mentions,
                total_queries
            FROM llm_monitoring_snapshots
            WHERE project_id = %s
            AND snapshot_date >= %s
        """
        comparison_params = [project_id, start_date]
        if enabled_llms_filter:
            comparison_query += " AND llm_provider = ANY(%s)"
            comparison_params.append(enabled_llms_filter)
        comparison_query += " ORDER BY snapshot_date DESC, llm_provider"
        cur.execute(comparison_query, comparison_params)
        
        rows = cur.fetchall()
        
        # Recalcular position_source con criterio estricto de dominio/subdominio para enlaces.
        position_source_rows_query = """
            SELECT
                llm_provider,
                analysis_date,
                brand_mentioned,
                mention_contexts,
                sources
            FROM llm_monitoring_results
            WHERE project_id = %s
            AND analysis_date >= %s
        """
        position_source_rows_params = [project_id, start_date]
        if enabled_llms_filter:
            position_source_rows_query += " AND llm_provider = ANY(%s)"
            position_source_rows_params.append(enabled_llms_filter)
        cur.execute(position_source_rows_query, position_source_rows_params)

        position_rows = cur.fetchall()

        # Crear lookup dict para position_source por (llm_provider, date)
        position_source_map = {}
        for row in position_rows:
            key = (row['llm_provider'], _to_date_key(row.get('analysis_date')))
            bucket = position_source_map.setdefault(key, {'text_count': 0, 'link_count': 0, 'both_count': 0})

            raw_contexts = row.get('mention_contexts') or []
            if isinstance(raw_contexts, str):
                try:
                    raw_contexts = json.loads(raw_contexts)
                except Exception:
                    raw_contexts = []
            if not isinstance(raw_contexts, list):
                raw_contexts = []

            raw_sources = row.get('sources') or []
            if isinstance(raw_sources, str):
                try:
                    raw_sources = json.loads(raw_sources)
                except Exception:
                    raw_sources = []
            if not isinstance(raw_sources, list):
                raw_sources = []

            has_text_context = False
            has_link_context = False
            for context in raw_contexts:
                ctx = str(context or '').strip()
                if not ctx:
                    continue
                if ctx.startswith('🔗'):
                    has_link_context = True
                else:
                    has_text_context = True

            brand_in_link = False
            if normalized_brand_domain:
                for source in raw_sources:
                    if not isinstance(source, dict):
                        continue
                    source_host = _extract_source_host(source.get('url'))
                    if source_host and (
                        source_host == normalized_brand_domain or
                        source_host.endswith(f".{normalized_brand_domain}")
                    ):
                        brand_in_link = True
                        break

            # Fallback defensivo para registros antiguos sin contextos.
            brand_mentioned = bool(row.get('brand_mentioned'))
            if has_text_context:
                brand_in_text = True
            elif brand_in_link:
                brand_in_text = False
            elif brand_mentioned and not has_link_context:
                brand_in_text = True
            else:
                brand_in_text = False

            if not brand_in_text and not brand_in_link:
                continue

            if brand_in_text and brand_in_link:
                bucket['both_count'] += 1
            elif brand_in_text:
                bucket['text_count'] += 1
            else:
                bucket['link_count'] += 1

        for key, counts in position_source_map.items():
            text_count = counts['text_count']
            link_count = counts['link_count']
            both_count = counts['both_count']
            
            # Determinar badge predominante
            if both_count > 0 or (text_count > 0 and link_count > 0):
                dominant_source = 'both'
            elif text_count >= link_count:
                dominant_source = 'text'
            else:
                dominant_source = 'link'
            
            position_source_map[key] = {
                'dominant': dominant_source,
                'text_count': text_count,
                'link_count': link_count,
                'both_count': both_count
            }
        
        # Formatear datos
        comparison_list = []
        for c in rows:
            # 🔧 FIX: Calcular sentiment basándose en los contadores, igual que en el KPI
            positive_mentions = c.get('positive_mentions') or 0
            neutral_mentions = c.get('neutral_mentions') or 0
            negative_mentions = c.get('negative_mentions') or 0
            total_queries_row = c.get('total_queries') or 0
            
            positive_pct = (positive_mentions / total_queries_row * 100) if total_queries_row else 0
            neutral_pct = (neutral_mentions / total_queries_row * 100) if total_queries_row else 0
            negative_pct = (negative_mentions / total_queries_row * 100) if total_queries_row else 0
            
            # ✨ NUEVO: Seleccionar Share of Voice según métrica
            if metric_type == 'weighted':
                sov = c.get('weighted_share_of_voice') or c.get('share_of_voice') or 0
            else:
                sov = c.get('share_of_voice') or 0
            
            # ✨ NUEVO: Obtener position_source info para este snapshot
            snapshot_key = (c['llm_provider'], _to_date_key(c.get('snapshot_date')))
            position_source_info = position_source_map.get(snapshot_key, {
                'dominant': None,
                'text_count': 0,
                'link_count': 0,
                'both_count': 0
            })
            
            comparison_list.append({
                'llm_provider': c['llm_provider'],
                'snapshot_date': c['snapshot_date'].isoformat() if c['snapshot_date'] else None,
                'mention_rate': float(c['mention_rate']) if c['mention_rate'] is not None else 0,
                'total_mentions': c.get('total_mentions') or 0,  # 🔧 FIX: Campo faltante para Grid.js
                'avg_position': float(c['avg_position']) if c['avg_position'] is not None else None,
                'position_source': position_source_info['dominant'],  # ✨ NUEVO: 'text', 'link', 'both'
                'position_source_details': position_source_info,  # ✨ NUEVO: Detalles para tooltip
                'share_of_voice': float(sov) if sov is not None else 0,  # ✨ MODIFICADO: Usar métrica seleccionada
                'sentiment_score': float(c['avg_sentiment_score']) if c['avg_sentiment_score'] is not None else 0,
                'sentiment': {
                    'positive': float(positive_pct),
                    'neutral': float(neutral_pct),
                    'negative': float(negative_pct)
                },
                'total_queries': total_queries_row
            })
        
        # Agrupar por fecha para comparación lado a lado
        by_date = {}
        for c in comparison_list:
            date = c['snapshot_date']
            if date not in by_date:
                by_date[date] = {}
            by_date[date][c['llm_provider']] = c

        # ── Previous period aggregated metrics (period-over-period) ──
        previous_period = {}
        try:
            prev_end_comp = start_date  # previous period ends where current starts
            prev_start_comp = (datetime.now() - timedelta(days=days * 2)).strftime('%Y-%m-%d')

            prev_comp_query = """
                SELECT llm_provider,
                       AVG(mention_rate) as avg_mention_rate,
                       AVG(share_of_voice) as avg_sov,
                       AVG(weighted_share_of_voice) as avg_weighted_sov,
                       AVG(avg_sentiment_score) as avg_sentiment
                FROM llm_monitoring_snapshots
                WHERE project_id = %s
                  AND snapshot_date >= %s AND snapshot_date < %s
            """
            prev_comp_params = [project_id, prev_start_comp, prev_end_comp]
            if enabled_llms_filter:
                prev_comp_query += " AND llm_provider = ANY(%s)"
                prev_comp_params.append(enabled_llms_filter)
            prev_comp_query += " GROUP BY llm_provider"
            cur.execute(prev_comp_query, prev_comp_params)

            for row in cur.fetchall():
                previous_period[row['llm_provider']] = {
                    'avg_mention_rate': float(row['avg_mention_rate']) if row['avg_mention_rate'] is not None else 0,
                    'avg_sov': float(row['avg_sov']) if row['avg_sov'] is not None else 0,
                    'avg_weighted_sov': float(row['avg_weighted_sov']) if row['avg_weighted_sov'] is not None else 0,
                    'avg_sentiment': float(row['avg_sentiment']) if row['avg_sentiment'] is not None else 0
                }
        except Exception as prev_comp_err:
            logger.warning(f"Could not compute previous period comparison: {prev_comp_err}")

        return jsonify({
            'success': True,
            'project_id': project_id,
            'comparison': comparison_list,
            'by_date': by_date,
            'previous_period': previous_period
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo comparativa: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo comparativa: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


# ============================================================================
# ENDPOINTS: CONFIGURACIÓN
# ============================================================================
# 
# NOTA: Los endpoints de configuración de API keys y presupuesto por usuario
# han sido ELIMINADOS porque en este modelo de negocio, los usuarios NO 
# configuran sus propias API keys. El servicio usa API keys globales
# gestionadas por el dueño del servicio en variables de entorno.
#
# Si en el futuro se necesita un modelo "Enterprise" donde clientes grandes
# usen sus propias APIs, se pueden restaurar estos endpoints.
# ============================================================================


# ============================================================================
# ENDPOINTS: QUERIES DETALLADAS
# ============================================================================

@llm_monitoring_bp.route('/projects/<int:project_id>/queries', methods=['GET'])
@login_required
@validate_project_ownership
def get_project_queries(project_id):
    """
    Obtener tabla detallada de queries/prompts con métricas agregadas
    
    Devuelve información similar a la tabla de LLM Pulse:
    - Query text
    - Country & Language
    - Total responses (cuántos LLMs respondieron)
    - Total mentions
    - Visibility % (promedio)
    - Average position
    - Last update
    - Creation date
    
    Query params:
        days: Días hacia atrás para filtrar resultados (default: 30)
    
    Returns:
        JSON con lista de queries y sus métricas
    """
    user = get_current_user()
    days = _normalize_days_param(request.args.get('days'), default=30)
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        # Obtener información del proyecto (para country)
        cur.execute("""
            SELECT name, language, country_code, enabled_llms
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        
        project = cur.fetchone()
        if not project:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        enabled_llms_filter = project.get('enabled_llms') or []
        
        # Calcular rango de fechas
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=days)
        
        # Obtener información del proyecto para filtrar URLs de marca
        cur.execute("""
            SELECT brand_domain FROM llm_monitoring_projects WHERE id = %s
        """, (project_id,))
        
        project_info = cur.fetchone()
        brand_domain = project_info['brand_domain'] if project_info else None
        
        # Obtener queries con métricas agregadas
        # ✨ MEJORADO: Contar menciones en texto + menciones en URLs
        query_metrics_sql = """
            WITH query_metrics AS (
                SELECT 
                    q.id,
                    q.query_text,
                    q.language,
                    q.query_type,
                    q.added_at as created_at,
                    COUNT(DISTINCT r.llm_provider) as total_responses,
                    COUNT(DISTINCT r.id) as total_results,
                    -- ✨ NUEVO: Contar menciones en texto (brand_mentioned)
                    SUM(CASE WHEN r.brand_mentioned THEN 1 ELSE 0 END) as text_mentions,
                    -- ✨ NUEVO: Contar URLs de marca en sources (requiere jsonb_array_elements)
                    SUM(
                        CASE 
                            WHEN r.sources IS NOT NULL AND r.sources::text != '[]' 
                            THEN (
                                SELECT COUNT(*) 
                                FROM jsonb_array_elements(r.sources::jsonb) AS source
                                WHERE source->>'url' ILIKE %s
                            )
                            ELSE 0
                        END
                    ) as url_citations,
                    AVG(CASE WHEN r.brand_mentioned THEN 100 ELSE 0 END) as visibility_pct,
                    AVG(r.position_in_list) FILTER (WHERE r.position_in_list IS NOT NULL) as avg_position,
                    MAX(r.analysis_date) as last_analysis_date,
                    MAX(r.created_at) as last_update
                FROM llm_monitoring_queries q
                LEFT JOIN llm_monitoring_results r ON q.id = r.query_id 
                    AND r.analysis_date >= %s 
                    AND r.analysis_date <= %s
                    {llm_filter}
                WHERE q.project_id = %s AND q.is_active = TRUE
                GROUP BY q.id, q.query_text, q.language, q.query_type, q.added_at
            )
            SELECT 
                id,
                query_text,
                language,
                query_type,
                created_at,
                total_responses,
                total_results,
                -- ✨ NUEVO: Total de menciones = menciones en texto + citaciones en URLs
                (COALESCE(text_mentions, 0) + COALESCE(url_citations, 0)) as total_mentions,
                text_mentions,
                url_citations,
                ROUND(visibility_pct::numeric, 1) as visibility_pct,
                ROUND(avg_position::numeric, 1) as avg_position,
                last_analysis_date,
                last_update
            FROM query_metrics
            ORDER BY last_update DESC NULLS LAST, created_at DESC
        """.format(
            llm_filter="AND r.llm_provider = ANY(%s)" if enabled_llms_filter else ""
        )
        query_metrics_params = [f'%{brand_domain}%' if brand_domain else '%', start_date, end_date]
        if enabled_llms_filter:
            query_metrics_params.append(enabled_llms_filter)
        query_metrics_params.append(project_id)
        cur.execute(query_metrics_sql, query_metrics_params)
        
        queries_raw = cur.fetchall()
        
        # ✨ NUEVO: Obtener menciones detalladas por LLM para cada query (para acordeón expandible)
        # Esto nos permite mostrar qué LLMs mencionaron la marca y los competidores
        # Usamos DISTINCT ON para obtener solo el resultado más reciente por (query_id, llm_provider)
        # 🔧 FIX: Incluir información sobre menciones en URLs también
        mentions_query_sql = """
            SELECT DISTINCT ON (r.query_id, r.llm_provider)
                r.query_id,
                r.llm_provider,
                r.brand_mentioned,
                r.position_in_list,
                r.competitors_mentioned,
                r.sources,
                r.analysis_date
            FROM llm_monitoring_results r
            WHERE r.query_id = ANY(%s)
                AND r.analysis_date >= %s
                AND r.analysis_date <= %s
                {llm_filter}
            ORDER BY r.query_id, r.llm_provider, r.analysis_date DESC
        """.format(
            llm_filter="AND r.llm_provider = ANY(%s)" if enabled_llms_filter else ""
        )
        mentions_query_params = [[q['id'] for q in queries_raw], start_date, end_date]
        if enabled_llms_filter:
            mentions_query_params.append(enabled_llms_filter)
        cur.execute(mentions_query_sql, mentions_query_params)
        
        mentions_by_query = {}
        for row in cur.fetchall():
            query_id = row['query_id']
            llm = row['llm_provider']
            
            if query_id not in mentions_by_query:
                mentions_by_query[query_id] = {}
            
            # 🔧 FIX: Detectar menciones en URLs también
            brand_in_text = row['brand_mentioned'] or False
            brand_in_urls = False
            
            # Verificar si la marca aparece en las URLs citadas
            if brand_domain and row['sources']:
                sources = row['sources']
                # sources puede ser una string JSON o ya un dict/list
                if isinstance(sources, str):
                    import json
                    try:
                        sources = json.loads(sources)
                    except:
                        sources = []
                
                if isinstance(sources, list):
                    for source in sources:
                        if isinstance(source, dict):
                            url = source.get('url', '').lower()
                            if brand_domain.lower() in url:
                                brand_in_urls = True
                                break
            
            # La marca fue mencionada si apareció en texto O en URLs
            brand_mentioned_total = brand_in_text or brand_in_urls
            
            # Ahora cada fila es única por (query_id, llm_provider) - el más reciente
            mentions_by_query[query_id][llm] = {
                'brand_mentioned': brand_mentioned_total,  # 🔧 FIX: Total incluyendo URLs
                'brand_mentioned_in_text': brand_in_text,  # ✨ NUEVO: Desglose
                'brand_mentioned_in_urls': brand_in_urls,  # ✨ NUEVO: Desglose
                'position': row['position_in_list'],
                'competitors': row['competitors_mentioned'] or {}
            }
        
        cur.close()
        conn.close()
        
        # Formatear datos para el frontend
        queries_list = []
        for q in queries_raw:
            query_id = q['id']
            
            # ✨ NUEVO: Añadir información de menciones por LLM
            mentions_detail = mentions_by_query.get(query_id, {})
            
            queries_list.append({
                'id': query_id,
                'prompt': q['query_text'],
                'country': 'Global',  # Por ahora global, se puede añadir por query
                'language': q['language'] or project['language'] or 'en',
                'query_type': q['query_type'],
                'total_responses': q['total_responses'] or 0,
                'total_mentions': q['total_mentions'] or 0,
                'visibility_pct': float(q['visibility_pct']) if q['visibility_pct'] else 0,
                'avg_position': float(q['avg_position']) if q['avg_position'] else None,
                'last_update': q['last_update'].isoformat() if q['last_update'] else None,
                'last_analysis_date': q['last_analysis_date'].isoformat() if q['last_analysis_date'] else None,
                'created_at': q['created_at'].isoformat() if q['created_at'] else None,
                'mentions_by_llm': mentions_detail  # ✨ NUEVO: Detalles para acordeón
            })
        
        return jsonify({
            'success': True,
            'queries': queries_list,
            'total': len(queries_list),
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'days': days
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo queries: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo queries: {str(e)}'}), 500


# ============================================================================
# ENDPOINTS: SHARE OF VOICE HISTÓRICO
# ============================================================================

@llm_monitoring_bp.route('/projects/<int:project_id>/share-of-voice-history', methods=['GET'])
@login_required
@validate_project_ownership
def get_share_of_voice_history(project_id):
    """
    Obtener datos históricos de Share of Voice para gráfico de líneas temporal
    
    Similar a los gráficos comparativos de Manual AI, muestra la evolución
    del Share of Voice de la marca vs competidores a lo largo del tiempo.
    
    Query params:
        days: Número de días hacia atrás (default: 30)
        metric: 'normal' o 'weighted' (default: 'weighted') - tipo de Share of Voice
    
    Returns:
        JSON con datos para gráfico de líneas:
        {
            'dates': ['2025-01-01', '2025-01-02', ...],
            'datasets': [
                {
                    'label': 'Tu Marca',
                    'data': [45.2, 48.1, ...],
                    'borderColor': '#3b82f6',
                    ...
                },
                {
                    'label': 'Competidor 1',
                    'data': [30.5, 28.3, ...],
                    'borderColor': '#ef4444',
                    ...
                }
            ]
        }
    """
    user = get_current_user()
    days = _normalize_days_param(request.args.get('days'), default=30)
    metric_type = request.args.get('metric', 'weighted')  # 'normal' o 'weighted'
    query_scope = request.args.get('query_scope', 'all')  # 'all', 'branded', 'non_branded'

    # Validar metric_type
    if metric_type not in ['normal', 'weighted']:
        metric_type = 'weighted'
    if query_scope not in ['all', 'branded', 'non_branded']:
        query_scope = 'all'

    logger.info(f"📊 Share of Voice history requested - Type: {metric_type}, Days: {days}, Scope: {query_scope}")

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500

    try:
        cur = conn.cursor()

        # Obtener proyecto
        cur.execute("""
            SELECT
                id, user_id, name, brand_name, industry,
                brand_domain, brand_keywords,
                competitors, selected_competitors,
                language, country_code,
                queries_per_llm, enabled_llms,
                is_active, created_at, last_analysis_date
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))

        project = cur.fetchone()

        if not project:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        # Calcular fechas de inicio y fin
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        enabled_llms_filter = project.get('enabled_llms') or []
        brand_keywords_sov = project.get('brand_keywords') or []

        # ── Branded/Non-Branded scope: compute from individual results ──
        if query_scope != 'all' and brand_keywords_sov:
            sov_result_query = """
                SELECT r.analysis_date, r.llm_provider, q.query_text,
                       r.brand_mentioned, r.competitors_mentioned
                FROM llm_monitoring_results r
                JOIN llm_monitoring_queries q ON r.query_id = q.id
                WHERE r.project_id = %s AND r.analysis_date >= %s
            """
            sov_result_params = [project_id, start_date]
            if enabled_llms_filter:
                sov_result_query += " AND r.llm_provider = ANY(%s)"
                sov_result_params.append(enabled_llms_filter)
            sov_result_query += " ORDER BY r.analysis_date"
            cur.execute(sov_result_query, sov_result_params)
            all_results = cur.fetchall()

            # Filter by branded/non-branded
            filtered = []
            for r in all_results:
                is_branded = classify_query_branded(r.get('query_text', ''), brand_keywords_sov)
                if (query_scope == 'branded' and is_branded) or \
                   (query_scope == 'non_branded' and not is_branded):
                    filtered.append(r)

            # Group by date and compute SOV
            from collections import defaultdict
            scope_by_date = defaultdict(lambda: {'brand': 0, 'comp': 0})
            for r in filtered:
                d = r['analysis_date'].isoformat() if hasattr(r['analysis_date'], 'isoformat') else str(r['analysis_date'])
                if r.get('brand_mentioned'):
                    scope_by_date[d]['brand'] += 1
                cm = r.get('competitors_mentioned')
                if cm:
                    if isinstance(cm, str):
                        try:
                            cm = json.loads(cm)
                        except Exception:
                            cm = {}
                    if isinstance(cm, dict):
                        scope_by_date[d]['comp'] += sum(cm.values())

            dates_sorted = sorted(scope_by_date.keys())
            brand_data = []
            for d in dates_sorted:
                b = scope_by_date[d]['brand']
                c = scope_by_date[d]['comp']
                total = b + c
                brand_data.append(round((b / total) * 100, 1) if total > 0 else 0)

            cur.close()
            conn.close()

            scope_label = 'Non-Branded' if query_scope == 'non_branded' else 'Branded'
            return jsonify({
                'success': True,
                'query_scope': query_scope,
                'dates': dates_sorted,
                'datasets': [{
                    'label': f'Your Brand ({scope_label})',
                    'data': brand_data,
                    'borderColor': '#3b82f6',
                    'backgroundColor': 'rgba(59, 130, 246, 0.1)',
                    'fill': True,
                    'tension': 0.3
                }]
            }), 200

        # Obtener todos los snapshots del período agrupados por fecha (incluir métricas ponderadas)
        sov_history_query = """
            SELECT 
                snapshot_date,
                llm_provider,
                share_of_voice,
                weighted_share_of_voice,
                total_mentions,
                total_competitor_mentions,
                competitor_breakdown,
                weighted_competitor_breakdown
            FROM llm_monitoring_snapshots
            WHERE project_id = %s 
                AND snapshot_date >= %s 
        """
        sov_history_params = [project_id, start_date]
        if enabled_llms_filter:
            sov_history_query += " AND llm_provider = ANY(%s)"
            sov_history_params.append(enabled_llms_filter)
        sov_history_query += " ORDER BY snapshot_date, llm_provider"
        cur.execute(sov_history_query, sov_history_params)
        
        snapshots = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # Si no hay datos, devolver estructura vacía
        if not snapshots:
            return jsonify({
                'success': True,
                'dates': [],
                'datasets': []
            }), 200
        
        # ✨ NEW: Use selected_competitors structure for clear attribution
        # =======================================================================
        # MAPEO DIRECTO: Cada dominio tiene sus keywords asociadas
        # =======================================================================
        
        selected_competitors = project.get('selected_competitors') or []
        
        # Crear mapeo: variante_detectada -> dominio_competidor
        competitor_mapping = {}
        competitor_display_names = {}  # dominio -> nombre_display
        
        def normalize_variant(variant):
            """Normaliza una variante para matching"""
            v = variant.lower().strip()
            # Quitar extensiones de dominio comunes
            v = v.replace('.com', '').replace('.es', '').replace('.net', '').replace('.org', '')
            v = v.replace('.mx', '').replace('.ar', '').replace('.cl', '').replace('.pe', '')
            v = v.replace('www.', '')
            return v
        
        def get_display_name(domain):
            """Obtiene nombre display limpio del dominio"""
            if not domain:
                return 'Unknown Competitor'
            # Quitar www. y extensión
            name = domain.replace('www.', '')
            # Tomar solo el nombre antes del TLD
            name_parts = name.split('.')
            if len(name_parts) > 0:
                return name_parts[0].upper()
            return domain.upper()
        
        # ✨ NEW: Mapear directamente desde selected_competitors
        for comp in selected_competitors:
            domain = comp.get('domain', '').strip()
            keywords = comp.get('keywords', [])
            
            if not domain:
                continue
            
            # Usar el dominio como identificador único
            domain_lower = domain.lower()
            display_name = get_display_name(domain)
            competitor_display_names[domain_lower] = display_name
            
            # Mapear el dominio a sí mismo
            competitor_mapping[domain_lower] = domain_lower
            
            # Mapear todas las keywords asociadas a este dominio
            for keyword in keywords:
                keyword_lower = keyword.lower().strip()
                competitor_mapping[keyword_lower] = domain_lower
                
                # También mapear variante normalizada
                keyword_normalized = normalize_variant(keyword)
                if keyword_normalized != keyword_lower:
                    competitor_mapping[keyword_normalized] = domain_lower
        
        # Agrupar datos por fecha Y agrupar menciones por dominio de competidor
        from collections import defaultdict
        data_by_date = defaultdict(lambda: {
            'brand_mentions': 0,
            'competitor_mentions': defaultdict(int),  # Agrupado por dominio
            'llm_count': 0
        })
        
        # ✨ NUEVO: Elegir la métrica correcta según el parámetro
        for snapshot in snapshots:
            date_str = snapshot['snapshot_date'].isoformat()
            
            # ✨ MEJORADO: Usar menciones ponderadas o normales según el parámetro
            if metric_type == 'weighted':
                # Si hay weighted_share_of_voice, usarlo para inferir menciones ponderadas
                try:
                    weighted_sov_raw = snapshot.get('weighted_share_of_voice')
                    # Convertir a float de forma segura
                    try:
                        weighted_sov = float(weighted_sov_raw) if weighted_sov_raw is not None else None
                    except Exception:
                        weighted_sov = None
                    
                    weighted_breakdown_raw = snapshot.get('weighted_competitor_breakdown')
                    # Asegurar que sea un dict; si viene en string JSON, parsear
                    if isinstance(weighted_breakdown_raw, dict):
                        weighted_breakdown = weighted_breakdown_raw
                    elif isinstance(weighted_breakdown_raw, str):
                        try:
                            parsed = json.loads(weighted_breakdown_raw)
                            weighted_breakdown = parsed if isinstance(parsed, dict) else {}
                        except Exception:
                            weighted_breakdown = {}
                    else:
                        weighted_breakdown = {}
                    
                    # Si tenemos datos ponderados, usarlos
                    if weighted_sov is not None and weighted_breakdown:
                        # Calcular menciones ponderadas de competidores
                        try:
                            total_weighted_comp = sum(float(v) for v in weighted_breakdown.values() if v is not None)
                        except Exception:
                            total_weighted_comp = 0.0
                        
                        if weighted_sov and weighted_sov > 0:
                            # SoV = brand / (brand + comp) * 100
                            # brand = (SoV / (100 - SoV)) * comp
                            if weighted_sov >= 100:
                                # Si SoV es 100%, la marca tiene todas las menciones
                                weighted_brand = total_weighted_comp if total_weighted_comp > 0 else 1.0
                            else:
                                weighted_brand = (weighted_sov / (100 - weighted_sov)) * total_weighted_comp
                        else:
                            weighted_brand = 0.0
                        
                        data_by_date[date_str]['brand_mentions'] += weighted_brand
                        breakdown = weighted_breakdown
                    else:
                        # Fallback a métricas normales si no hay datos ponderados
                        logger.warning(f"⚠️ No weighted data for {date_str}, falling back to normal metrics")
                        data_by_date[date_str]['brand_mentions'] += (snapshot['total_mentions'] or 0)
                        breakdown = snapshot.get('competitor_breakdown') or {}
                except Exception as e_weighted:
                    # Cualquier error en la ruta ponderada no debe romper el endpoint
                    logger.warning(f"⚠️ Weighted calc error on {date_str}: {e_weighted}. Using normal metrics as fallback.")
                    data_by_date[date_str]['brand_mentions'] += (snapshot['total_mentions'] or 0)
                    breakdown = snapshot.get('competitor_breakdown') or {}
            else:
                # Modo normal: usar métricas estándar
                data_by_date[date_str]['brand_mentions'] += (snapshot['total_mentions'] or 0)
                breakdown = snapshot.get('competitor_breakdown') or {}
            
            data_by_date[date_str]['llm_count'] += 1
            
            # 🔧 FIX: Asegurar que breakdown es un dict antes de iterar
            if isinstance(breakdown, str):
                try:
                    breakdown = json.loads(breakdown)
                except (json.JSONDecodeError, TypeError):
                    logger.warning(f"⚠️ Could not parse breakdown JSON for {date_str}, skipping")
                    breakdown = {}
            
            if not isinstance(breakdown, dict):
                logger.warning(f"⚠️ Breakdown is not a dict for {date_str}, skipping")
                breakdown = {}
            
            for detected_variant, mentions in breakdown.items():
                variant_lower = detected_variant.lower().strip()
                variant_normalized = normalize_variant(detected_variant)
                
                # Buscar en el mapeo directo
                competitor_domain = None
                if variant_lower in competitor_mapping:
                    competitor_domain = competitor_mapping[variant_lower]
                elif variant_normalized in competitor_mapping:
                    competitor_domain = competitor_mapping[variant_normalized]
                else:
                    # Buscar por coincidencia parcial (más robustez)
                    for mapped_variant, domain in competitor_mapping.items():
                        if (mapped_variant in variant_lower or 
                            variant_lower in mapped_variant or
                            normalize_variant(mapped_variant) == variant_normalized):
                            competitor_domain = domain
                            break
                
                # Si encontramos el dominio, acumular menciones
                if competitor_domain:
                    data_by_date[date_str]['competitor_mentions'][competitor_domain] += mentions
                # Si no, ignorar (no es un competidor configurado)
        
        # Ordenar fechas
        dates = sorted(data_by_date.keys())
        
        # Preparar datasets
        datasets = []
        
        # Dataset para la marca principal (TODAS las variantes ya están sumadas en total_mentions)
        brand_name = project['brand_name'] or 'Your Brand'
        brand_data = []
        
        for date_str in dates:
            day_data = data_by_date[date_str]
            brand_mentions = day_data['brand_mentions']
            total_comp_mentions = sum(day_data['competitor_mentions'].values())
            total_mentions = brand_mentions + total_comp_mentions
            
            # Calcular Share of Voice de la marca
            sov = (brand_mentions / total_mentions * 100) if total_mentions > 0 else 0
            brand_data.append(round(sov, 2))
        
        datasets.append({
            'label': brand_name,
            'data': brand_data,
            'borderColor': '#3b82f6',  # Blue para la marca
            'backgroundColor': 'rgba(59, 130, 246, 0.1)',
            'borderWidth': 3,
            'tension': 0.4,
            'fill': True,
            'pointRadius': 4,
            'pointHoverRadius': 6
        })
        
        # Datasets para competidores (ahora agrupados)
        competitor_colors = [
            '#ef4444',  # Red
            '#f97316',  # Orange
            '#22c55e',  # Green
            '#a855f7',  # Purple
            '#ec4899'   # Pink
        ]
        
        # ✨ NEW: Obtener lista única de dominios de competidores
        all_competitor_domains = set()
        for day_data in data_by_date.values():
            all_competitor_domains.update(day_data['competitor_mentions'].keys())
        
        for idx, competitor_domain in enumerate(sorted(all_competitor_domains)):
            comp_data = []
            
            # ✨ NEW: Obtener nombre display del dominio
            display_name = competitor_display_names.get(competitor_domain, get_display_name(competitor_domain))
            
            for date_str in dates:
                day_data = data_by_date[date_str]
                brand_mentions = day_data['brand_mentions']
                comp_mentions = day_data['competitor_mentions'].get(competitor_domain, 0)
                total_comp_mentions = sum(day_data['competitor_mentions'].values())
                total_mentions = brand_mentions + total_comp_mentions
                
                # Calcular Share of Voice del competidor
                sov = (comp_mentions / total_mentions * 100) if total_mentions > 0 else 0
                comp_data.append(round(sov, 2))
            
            color = competitor_colors[idx % len(competitor_colors)]
            datasets.append({
                'label': display_name,
                'data': comp_data,
                'borderColor': color,
                'backgroundColor': color.replace('rgb', 'rgba').replace(')', ', 0.1)') if 'rgb' in color else f'{color}20',
                'borderWidth': 2,
                'tension': 0.4,
                'fill': False,
                'pointRadius': 3,
                'pointHoverRadius': 5
            })
        
        # =======================================================================
        # DATOS ADICIONALES: Menciones absolutas + Donut data
        # =======================================================================
        
        # 1. Datasets para gráfico de menciones (números absolutos)
        mentions_datasets = []
        
        # Dataset de menciones de marca
        brand_mentions_data = []
        for date_str in dates:
            day_data = data_by_date[date_str]
            brand_mentions_data.append(day_data['brand_mentions'])
        
        mentions_datasets.append({
            'label': brand_name,
            'data': brand_mentions_data,
            'borderColor': '#3b82f6',
            'backgroundColor': 'rgba(59, 130, 246, 0.1)',
            'borderWidth': 3,
            'tension': 0.4,
            'fill': True,
            'pointRadius': 4,
            'pointHoverRadius': 6
        })
        
        # Datasets de menciones de competidores
        for idx, competitor_domain in enumerate(sorted(all_competitor_domains)):
            comp_mentions_data = []
            display_name = competitor_display_names.get(competitor_domain, get_display_name(competitor_domain))
            
            for date_str in dates:
                day_data = data_by_date[date_str]
                comp_mentions_data.append(day_data['competitor_mentions'].get(competitor_domain, 0))
            
            color = competitor_colors[idx % len(competitor_colors)]
            mentions_datasets.append({
                'label': display_name,
                'data': comp_mentions_data,
                'borderColor': color,
                'backgroundColor': color.replace('rgb', 'rgba').replace(')', ', 0.1)') if 'rgb' in color else f'{color}20',
                'borderWidth': 2,
                'tension': 0.4,
                'fill': False,
                'pointRadius': 3,
                'pointHoverRadius': 5
            })
        
        # 2. Datos para gráfico de rosco (promedio del período)
        total_brand_mentions_period = sum(data_by_date[date_str]['brand_mentions'] for date_str in dates)
        total_competitor_mentions_period = defaultdict(int)
        
        for date_str in dates:
            for comp_domain, mentions in data_by_date[date_str]['competitor_mentions'].items():
                total_competitor_mentions_period[comp_domain] += mentions
        
        # Calcular totales
        grand_total = total_brand_mentions_period + sum(total_competitor_mentions_period.values())
        
        # Preparar datos del donut
        donut_labels = [brand_name]
        donut_values = [round(total_brand_mentions_period / grand_total * 100, 2) if grand_total > 0 else 0]
        donut_colors = ['#3b82f6']
        
        for idx, competitor_domain in enumerate(sorted(all_competitor_domains)):
            display_name = competitor_display_names.get(competitor_domain, get_display_name(competitor_domain))
            comp_total = total_competitor_mentions_period.get(competitor_domain, 0)
            comp_percentage = round(comp_total / grand_total * 100, 2) if grand_total > 0 else 0
            
            if comp_percentage > 0:  # Solo incluir si tiene menciones
                donut_labels.append(display_name)
                donut_values.append(comp_percentage)
                donut_colors.append(competitor_colors[idx % len(competitor_colors)])
        
        return jsonify({
            'success': True,
            'metric_type': metric_type,  # ✨ NUEVO: Indicar qué métrica se está usando
            'dates': dates,
            'datasets': datasets,  # Share of Voice (%)
            'mentions_datasets': mentions_datasets,  # Menciones absolutas
            'donut_data': {
                'labels': donut_labels,
                'values': donut_values,
                'colors': donut_colors
            },
            'period': {
                'start_date': start_date,
                'end_date': end_date,
                'days': days
            }
        }), 200
        
    except Exception as e:
        logger.error(f"❌ Error obteniendo histórico de Share of Voice: {e}", exc_info=True)
        logger.error(f"   Project ID: {project_id}, Days: {days}, Metric: {metric_type}")
        return jsonify({
            'error': f'Error obteniendo datos: {str(e)}',
            'details': 'Check server logs for more information'
        }), 500


# ============================================================================
# ENDPOINTS: MODELOS
# ============================================================================

@llm_monitoring_bp.route('/models', methods=['GET'])
@login_required
def get_models():
    """
    Lista todos los modelos LLM disponibles
    
    Query params opcionales:
        provider: Filtrar por proveedor (openai, anthropic, google, perplexity)
        is_current: Filtrar solo modelos actuales (true/false)
    
    Returns:
        JSON con lista de modelos y sus precios
    """
    provider = request.args.get('provider')
    is_current = request.args.get('is_current')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        query = "SELECT * FROM llm_model_registry WHERE 1=1"
        params = []
        
        if provider:
            query += " AND llm_provider = %s"
            params.append(provider)
        
        if is_current is not None:
            query += " AND is_current = %s"
            params.append(is_current.lower() == 'true')
        
        query += " ORDER BY llm_provider, is_current DESC, model_display_name"
        
        cur.execute(query, params)
        models = cur.fetchall()
        
        # Formatear modelos
        models_list = []
        for m in models:
            models_list.append({
                'id': m['id'],
                'llm_provider': m['llm_provider'],
                'model_id': m['model_id'],
                'model_display_name': m['model_display_name'],
                'cost_per_1m_input_tokens': float(m['cost_per_1m_input_tokens']) if m['cost_per_1m_input_tokens'] else 0,
                'cost_per_1m_output_tokens': float(m['cost_per_1m_output_tokens']) if m['cost_per_1m_output_tokens'] else 0,
                'max_tokens': m['max_tokens'],
                'is_current': m['is_current'],
                'is_available': m['is_available'],
                'created_at': m['created_at'].isoformat() if m['created_at'] else None,
                'updated_at': m['updated_at'].isoformat() if m['updated_at'] else None
            })
        
        return jsonify({
            'success': True,
            'models': models_list,
            'total': len(models_list)
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo modelos: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo modelos: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/models/current', methods=['GET'])
@login_required
def get_current_models():
    """
    Obtiene los modelos actuales (is_current=TRUE) de cada proveedor
    
    Returns:
        JSON con los modelos actuales por proveedor
        
    Example response:
        {
            "success": true,
            "models": {
                "openai": {"model_id": "gpt-5.4", "display_name": "GPT-5.4"},
                "anthropic": {"model_id": "claude-sonnet-4-6", "display_name": "Claude Sonnet 4.6"},
                "google": {"model_id": "gemini-3-flash-preview", "display_name": "Gemini 3 Flash"},
                "perplexity": {"model_id": "sonar-pro", "display_name": "Perplexity Sonar Pro"}
            }
        }
    """
    # Fallbacks por defecto (Model IDs correctos según docs oficiales)
    fallbacks = {
        'openai': {'model_id': 'gpt-5.4', 'display_name': 'GPT-5.4'},
        'anthropic': {'model_id': 'claude-sonnet-4-6', 'display_name': 'Claude Sonnet 4.6'},
        'google': {'model_id': 'gemini-3-flash-preview', 'display_name': 'Gemini 3 Flash'},
        'perplexity': {'model_id': 'sonar-pro', 'display_name': 'Perplexity Sonar Pro'}
    }
    
    conn = get_db_connection()
    if not conn:
        # Si no hay BD, devolver fallbacks
        return jsonify({
            'success': True,
            'models': fallbacks,
            'source': 'fallback'
        }), 200
    
    try:
        cur = conn.cursor()
        
        # Query con columnas incluyendo knowledge_cutoff
        cur.execute("""
            SELECT llm_provider, model_id, model_display_name,
                   knowledge_cutoff, knowledge_cutoff_date
            FROM llm_model_registry
            WHERE is_current = TRUE AND is_available = TRUE
            ORDER BY llm_provider
        """)

        models = cur.fetchall()

        models_dict = {}
        for m in models:
            models_dict[m['llm_provider']] = {
                'model_id': m['model_id'],
                'display_name': m['model_display_name'] or m['model_id'],
                'knowledge_cutoff': m.get('knowledge_cutoff'),
                'knowledge_cutoff_date': m['knowledge_cutoff_date'].isoformat() if m.get('knowledge_cutoff_date') else None,
            }
        
        # Aplicar fallbacks para providers que no tienen modelo en BD
        for provider, fallback in fallbacks.items():
            if provider not in models_dict:
                models_dict[provider] = fallback
        
        return jsonify({
            'success': True,
            'models': models_dict,
            'source': 'database' if models else 'fallback'
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo modelos actuales: {e}", exc_info=True)
        # En caso de error, devolver fallbacks
        return jsonify({
            'success': True,
            'models': fallbacks,
            'source': 'fallback_error'
        }), 200
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/models/<int:model_id>', methods=['PUT'])
@admin_required
def update_model(model_id):
    """
    Actualiza un modelo LLM (solo admin)
    
    Body:
    {
        "cost_per_1m_input_tokens": 15.0,
        "cost_per_1m_output_tokens": 45.0,
        "is_current": true,
        "is_available": true
    }
    
    Returns:
        JSON con el modelo actualizado
    """
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'Body vacío'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()
        
        updates = []
        params = []
        
        if 'cost_per_1m_input_tokens' in data:
            updates.append("cost_per_1m_input_tokens = %s")
            params.append(data['cost_per_1m_input_tokens'])
        
        if 'cost_per_1m_output_tokens' in data:
            updates.append("cost_per_1m_output_tokens = %s")
            params.append(data['cost_per_1m_output_tokens'])
        
        if 'is_current' in data:
            updates.append("is_current = %s")
            params.append(data['is_current'])
        
        if 'is_available' in data:
            updates.append("is_available = %s")
            params.append(data['is_available'])
        
        if not updates:
            return jsonify({'error': 'No hay campos para actualizar'}), 400
        
        updates.append("updated_at = NOW()")
        params.append(model_id)
        
        query = f"""
            UPDATE llm_model_registry
            SET {', '.join(updates)}
            WHERE id = %s
            RETURNING *
        """
        
        cur.execute(query, params)
        model = cur.fetchone()
        
        if not model:
            return jsonify({'error': 'Modelo no encontrado'}), 404
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'model': {
                'id': model['id'],
                'llm_provider': model['llm_provider'],
                'model_id': model['model_id'],
                'model_display_name': model['model_display_name'],
                'cost_per_1m_input_tokens': float(model['cost_per_1m_input_tokens']),
                'cost_per_1m_output_tokens': float(model['cost_per_1m_output_tokens']),
                'is_current': model['is_current'],
                'is_available': model['is_available'],
                'updated_at': model['updated_at'].isoformat() if model['updated_at'] else None
            }
        }), 200
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Error actualizando modelo: {e}", exc_info=True)
        return jsonify({'error': f'Error actualizando modelo: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


# ============================================================================
# CRON JOBS
# ============================================================================

@llm_monitoring_bp.route('/cron/daily-analysis', methods=['POST'])
@cron_or_auth_required
def trigger_daily_analysis():
    """
    Trigger para análisis diario automático (cron job)

    Query params:
        async (int): Si es 1, ejecuta en background y responde inmediatamente con 202
        project_id (int): Si se proporciona, analiza SOLO ese proyecto (útil para re-runs)

    Returns:
        JSON con resultado de la ejecución del cron
    """
    auth_error = _ensure_cron_token_or_admin()
    if auth_error:
        return auth_error

    try:
        # Verificar si se solicita ejecución asíncrona
        is_async = request.args.get('async', '0') == '1'
        triggered_by = request.args.get('triggered_by', 'cron')
        single_project_id = request.args.get('project_id', type=int)

        # --- Modo proyecto individual (sin lock global) ---
        if single_project_id:
            def run_single_project(pid):
                service = MultiLLMMonitoringService(api_keys=None)
                return service.analyze_project(project_id=pid, max_workers=8)

            if is_async:
                def run_single_bg():
                    try:
                        logger.info(f"🚀 LLM Monitoring: Single project analysis started (project_id={single_project_id})")
                        result = run_single_project(single_project_id)
                        logger.info(f"✅ LLM Monitoring: Single project analysis done: {result.get('total_queries_executed', 0)} queries")
                    except Exception as e:
                        logger.error(f"💥 LLM Monitoring: Single project analysis error: {e}")

                thread = threading.Thread(target=run_single_bg, daemon=True)
                thread.start()
                return jsonify({
                    'success': True,
                    'message': f'Analysis for project {single_project_id} triggered in background',
                    'async': True,
                    'project_id': single_project_id
                }), 202

            result = run_single_project(single_project_id)
            return jsonify({
                'success': result.get('success', result.get('total_queries_executed', 0) > 0),
                'project_id': single_project_id,
                'result': result
            }), 200

        # --- Modo todos los proyectos (comportamiento original) ---
        if is_async:
            # ✅ NUEVO: Intentar adquirir lock ANTES de lanzar thread
            run_id = acquire_analysis_lock(triggered_by=triggered_by)
            if run_id is None:
                # Ya hay un análisis en curso
                latest_run = get_latest_analysis_run()
                return jsonify({
                    'success': False,
                    'error': 'Analysis already running',
                    'message': 'Hay un análisis en curso. Espera a que termine antes de lanzar otro.',
                    'latest_run': latest_run
                }), 409  # Conflict

            # Ejecutar en background con lock y state tracking
            def run_analysis_in_background(run_id):
                try:
                    logger.info(f"🚀 LLM Monitoring: Starting daily analysis in background (run_id={run_id})")
                    results = analyze_all_active_projects(
                        api_keys=None,  # Usar variables de entorno
                        max_workers=10
                    )

                    # Procesar resultados
                    successful = sum(1 for r in results if 'error' not in r)
                    failed = sum(1 for r in results if 'error' in r)
                    total_queries = sum(r.get('total_queries_executed', 0) for r in results if 'error' not in r)

                    logger.info(f"✅ LLM Monitoring: Daily analysis completed - {len(results)} projects processed")
                    logger.info(f"   Successful: {successful}, Failed: {failed}, Total queries: {total_queries}")

                    # ✅ Liberar lock y registrar resultados
                    release_analysis_lock(
                        run_id=run_id,
                        total_projects=len(results),
                        successful=successful,
                        failed=failed,
                        total_queries=total_queries
                    )
                except Exception as e:
                    logger.error(f"💥 LLM Monitoring: Background analysis error: {e}")
                    # ✅ Liberar lock incluso en caso de error
                    release_analysis_lock(run_id=run_id, error_message=str(e))

            # Iniciar thread en background
            thread = threading.Thread(target=run_analysis_in_background, args=(run_id,), daemon=True)
            thread.start()

            logger.info(f"📤 LLM Monitoring: Daily analysis triggered (async mode, run_id={run_id})")
            return jsonify({
                'success': True,
                'message': 'Daily analysis triggered in background',
                'async': True,
                'run_id': run_id
            }), 202

        # Modo síncrono (comportamiento original) — también con lock
        run_id = acquire_analysis_lock(triggered_by=triggered_by)
        if run_id is None:
            latest_run = get_latest_analysis_run()
            return jsonify({
                'success': False,
                'error': 'Analysis already running',
                'message': 'Hay un análisis en curso. Espera a que termine antes de lanzar otro.',
                'latest_run': latest_run
            }), 409

        try:
            results = analyze_all_active_projects(
                api_keys=None,
                max_workers=10
            )

            successful = sum(1 for r in results if 'error' not in r)
            failed = sum(1 for r in results if 'error' in r)
            total_queries = sum(r.get('total_queries_executed', 0) for r in results if 'error' not in r)

            release_analysis_lock(
                run_id=run_id,
                total_projects=len(results),
                successful=successful,
                failed=failed,
                total_queries=total_queries
            )

            return jsonify({
                'success': True,
                'total_projects': len(results),
                'successful': successful,
                'failed': failed,
                'total_queries': total_queries,
                'run_id': run_id,
                'results': results
            }), 200
        except Exception as e:
            release_analysis_lock(run_id=run_id, error_message=str(e))
            raise

    except Exception as e:
        logger.error(f"Error in daily analysis cron trigger: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# HEALTH CHECK
# ============================================================================

@llm_monitoring_bp.route('/health', methods=['GET'])
def health_check():
    """
    Health check del sistema LLM Monitoring
    
    Returns:
        JSON con estado del sistema
    """
    try:
        # Verificar conexión a BD
        conn = get_db_connection()
        if not conn:
            return jsonify({
                'status': 'error',
                'database': 'disconnected'
            }), 503
        
        cur = conn.cursor()
        
        # Contar proyectos activos
        cur.execute("SELECT COUNT(*) as count FROM llm_monitoring_projects WHERE is_active = TRUE")
        active_projects = cur.fetchone()['count']
        
        # Verificar API keys disponibles
        api_keys_available = {
            'openai': bool(os.getenv('OPENAI_API_KEY')),
            'anthropic': bool(os.getenv('ANTHROPIC_API_KEY')),
            'google': bool(os.getenv('GOOGLE_API_KEY')),
            'perplexity': bool(os.getenv('PERPLEXITY_API_KEY'))
        }
        
        cur.close()
        conn.close()
        
        return jsonify({
            'status': 'ok',
            'database': 'connected',
            'active_projects': active_projects,
            'api_keys_configured': sum(api_keys_available.values()),
            'api_keys': api_keys_available
        }), 200
        
    except Exception as e:
        logger.error(f"Health check failed: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500


@llm_monitoring_bp.route('/projects/<int:project_id>/responses', methods=['GET'])
@login_required
@validate_project_ownership
def get_project_responses(project_id):
    """
    Obtener respuestas detalladas de LLMs para inspección manual
    
    Query params:
        query_id: ID de query específica (opcional)
        llm_provider: Filtrar por proveedor (opcional)
        days: Días hacia atrás (default: 7)
    
    Returns:
        JSON con respuestas completas de cada LLM
    """
    query_id = request.args.get('query_id', type=int)
    llm_provider = request.args.get('llm_provider')
    days = _normalize_days_param(request.args.get('days'), default=7)
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500
    
    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT enabled_llms, brand_keywords
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        project_row = cur.fetchone()
        if not project_row:
            return jsonify({'error': 'Proyecto no encontrado'}), 404
        enabled_llms_filter = project_row.get('enabled_llms') or []
        resp_brand_keywords = project_row.get('brand_keywords') or []
        if llm_provider and enabled_llms_filter and llm_provider not in enabled_llms_filter:
            cur.close()
            conn.close()
            return jsonify({
                'error': 'llm_provider no habilitado para este proyecto'
            }), 400

        # Calcular rango de fechas
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=days)

        # Query base
        query = """
            SELECT
                r.id,
                r.query_id,
                q.query_text,
                r.llm_provider,
                r.model_used,
                r.brand_mentioned,
                r.position_in_list,
                r.sentiment,
                r.mention_contexts,
                r.competitors_mentioned,
                r.full_response,
                r.response_length,
                r.sources,
                r.analysis_date,
                r.created_at
            FROM llm_monitoring_results r
            JOIN llm_monitoring_queries q ON r.query_id = q.id
            WHERE r.project_id = %s
                AND r.analysis_date >= %s
                AND r.analysis_date <= %s
        """
        
        params = [project_id, start_date, end_date]
        
        # Filtros opcionales
        if query_id:
            query += " AND r.query_id = %s"
            params.append(query_id)
        
        if llm_provider:
            query += " AND r.llm_provider = %s"
            params.append(llm_provider)
        elif enabled_llms_filter:
            query += " AND r.llm_provider = ANY(%s)"
            params.append(enabled_llms_filter)
        
        query += " ORDER BY r.analysis_date DESC, q.query_text, r.llm_provider"
        
        cur.execute(query, params)
        results = cur.fetchall()
        
        # Formatear resultados
        responses = []
        for r in results:
            responses.append({
                'id': r['id'],
                'query_id': r['query_id'],
                'query_text': r['query_text'],
                'llm_provider': r['llm_provider'],
                'model_used': r['model_used'],
                'brand_mentioned': r['brand_mentioned'],
                'position_in_list': r['position_in_list'],
                'sentiment': r['sentiment'],
                'mention_contexts': r['mention_contexts'] or [],
                'competitors_mentioned': r['competitors_mentioned'] or {},
                'full_response': r['full_response'],
                'response_length': r['response_length'],
                'sources': r['sources'] or [],
                'is_branded_query': classify_query_branded(r['query_text'], resp_brand_keywords),
                'analysis_date': r['analysis_date'].isoformat() if r['analysis_date'] else None,
                'created_at': r['created_at'].isoformat() if r['created_at'] else None
            })
        
        cur.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'responses': responses,
            'total': len(responses),
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'days': days
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error obteniendo respuestas: {e}", exc_info=True)
        return jsonify({'error': f'Error obteniendo respuestas: {str(e)}'}), 500


# ============================================================================
# EXPORTACIÓN DE DATOS - Excel y PDF
# ============================================================================

@llm_monitoring_bp.route('/projects/<int:project_id>/export/excel', methods=['GET'])
@login_required
@validate_project_ownership
def export_project_excel(project_id):
    """
    Exportar datos COMPLETOS del proyecto a Excel (todas las métricas visibles en UI)

    Query params:
        days: int - Período de días (default: 30)

    Sheets generadas:
        1. Project Summary - Resumen del proyecto y métricas globales
        2. Share of Voice - SOV de marca y competidores por día
        3. LLM Comparison - Métricas comparativas por proveedor LLM
        4. Daily Metrics - Snapshots diarios detallados por LLM
        5. Prompts & Queries - Performance por prompt/query
        6. URL Rankings - URLs más citadas por los LLMs
        7. Sentiment Analysis - Distribución de sentimiento por LLM
        8. Detailed Results - Datos individuales por query × LLM × día
    """
    from io import BytesIO
    from flask import send_file
    json_module = json  # alias to avoid shadowing in nested scope

    logger.info(f"📥 Starting comprehensive Excel export for project {project_id}")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        logger.error(f"openpyxl no está instalado: {e}")
        return jsonify({'error': 'Excel export not available. Missing openpyxl library.'}), 500

    days = _normalize_days_param(request.args.get('days'), default=30)

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500

    try:
        cur = conn.cursor()

        # ──────────────────────────────────────────────────
        # FETCH ALL DATA
        # ──────────────────────────────────────────────────

        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        # 1. Project info
        cur.execute("""
            SELECT name, brand_name, industry, brand_domain, brand_keywords,
                   competitor_domains, selected_competitors,
                   language, country_code, enabled_llms, queries_per_llm,
                   is_active, created_at, last_analysis_date
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        project = cur.fetchone()

        if not project:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        enabled_llms_filter = project.get('enabled_llms') or []
        brand_keywords = project.get('brand_keywords') or []

        # Helper for LLM filter in queries
        def _add_llm_filter(query_str, params_list, column='llm_provider'):
            if enabled_llms_filter:
                query_str += f" AND {column} = ANY(%s)"
                params_list.append(enabled_llms_filter)
            return query_str, params_list

        # 2. Snapshots (daily metrics per LLM) — source for multiple sheets
        snap_query = """
            SELECT
                snapshot_date, llm_provider,
                total_queries, total_mentions, mention_rate,
                avg_position,
                appeared_in_top3, appeared_in_top5, appeared_in_top10,
                share_of_voice, weighted_share_of_voice,
                total_competitor_mentions, competitor_breakdown, weighted_competitor_breakdown,
                positive_mentions, neutral_mentions, negative_mentions, avg_sentiment_score,
                avg_response_time_ms, total_cost_usd, total_tokens
            FROM llm_monitoring_snapshots
            WHERE project_id = %s AND snapshot_date >= %s
        """
        snap_params = [project_id, start_date_str]
        snap_query, snap_params = _add_llm_filter(snap_query, snap_params)
        snap_query += " ORDER BY snapshot_date DESC, llm_provider"
        cur.execute(snap_query, snap_params)
        snapshots = cur.fetchall()

        # 3. Aggregated LLM metrics (from results table)
        metrics_query = """
            SELECT
                llm_provider,
                COUNT(DISTINCT query_id) as total_queries,
                COUNT(*) as total_results,
                SUM(CASE WHEN brand_mentioned THEN 1 ELSE 0 END) as total_mentions,
                ROUND(AVG(CASE WHEN brand_mentioned THEN 100.0 ELSE 0 END), 1) as mention_rate_pct,
                AVG(position_in_list) FILTER (WHERE position_in_list IS NOT NULL) as avg_position,
                SUM(CASE WHEN sentiment = 'positive' THEN 1 ELSE 0 END) as positive_count,
                SUM(CASE WHEN sentiment = 'neutral' THEN 1 ELSE 0 END) as neutral_count,
                SUM(CASE WHEN sentiment = 'negative' THEN 1 ELSE 0 END) as negative_count,
                AVG(sentiment_score) as avg_sentiment_score,
                SUM(cost_usd) as total_cost,
                SUM(tokens_used) as total_tokens,
                AVG(response_time_ms) as avg_response_time
            FROM llm_monitoring_results
            WHERE project_id = %s AND analysis_date >= %s AND analysis_date <= %s
        """
        metrics_params = [project_id, start_date, end_date]
        metrics_query, metrics_params = _add_llm_filter(metrics_query, metrics_params)
        metrics_query += " GROUP BY llm_provider ORDER BY llm_provider"
        cur.execute(metrics_query, metrics_params)
        llm_metrics = cur.fetchall()

        # 4. Query/Prompt level data
        queries_query = """
            SELECT
                q.query_text AS prompt,
                q.language,
                q.query_type,
                COUNT(DISTINCT r.llm_provider) as llms_analyzed,
                COUNT(r.id) as total_results,
                SUM(CASE WHEN r.brand_mentioned THEN 1 ELSE 0 END) as total_mentions,
                SUM(CASE WHEN r.position_source = 'link' OR r.position_source = 'both' THEN 1 ELSE 0 END) as url_citations,
                SUM(CASE WHEN r.position_source = 'text' THEN 1 ELSE 0 END) as text_mentions,
                ROUND(AVG(CASE WHEN r.brand_mentioned THEN 100.0 ELSE 0 END), 1) as visibility_pct,
                AVG(r.position_in_list) FILTER (WHERE r.position_in_list IS NOT NULL) as avg_position,
                MAX(r.analysis_date) as last_analysis
            FROM llm_monitoring_queries q
            LEFT JOIN llm_monitoring_results r ON q.id = r.query_id
                AND r.analysis_date >= %s AND r.analysis_date <= %s
                {llm_filter}
            WHERE q.project_id = %s AND q.is_active = TRUE
            GROUP BY q.id, q.query_text, q.language, q.query_type
            ORDER BY total_mentions DESC, visibility_pct DESC
        """.format(
            llm_filter="AND r.llm_provider = ANY(%s)" if enabled_llms_filter else ""
        )
        queries_params = [start_date, end_date]
        if enabled_llms_filter:
            queries_params.append(enabled_llms_filter)
        queries_params.append(project_id)
        cur.execute(queries_query, queries_params)
        queries = cur.fetchall()

        # 5. Detailed results (per query × LLM × date)
        detail_query = """
            SELECT
                r.analysis_date,
                r.llm_provider,
                r.model_used,
                q.query_text,
                r.brand_mentioned,
                r.mention_count,
                r.position_in_list,
                r.total_items_in_list,
                r.position_source,
                r.sentiment,
                r.sentiment_score,
                r.competitors_mentioned,
                r.sources,
                r.tokens_used,
                r.cost_usd,
                r.response_time_ms,
                r.response_length
            FROM llm_monitoring_results r
            JOIN llm_monitoring_queries q ON r.query_id = q.id
            WHERE r.project_id = %s AND r.analysis_date >= %s AND r.analysis_date <= %s
        """
        detail_params = [project_id, start_date, end_date]
        detail_query, detail_params = _add_llm_filter(detail_query, detail_params, 'r.llm_provider')
        detail_query += " ORDER BY r.analysis_date DESC, q.query_text, r.llm_provider"
        cur.execute(detail_query, detail_params)
        detailed_results = cur.fetchall()

        # 6b. Previous period data (for period-over-period comparison in export)
        prev_start_date = start_date - timedelta(days=days)
        prev_start_date_str = prev_start_date.strftime('%Y-%m-%d')

        # Previous branded/non-branded metrics
        prev_branded_export = {'branded_metrics': None, 'non_branded_metrics': None}
        try:
            if brand_keywords:
                prev_bq_export = """
                    SELECT r.llm_provider, q.query_text, r.brand_mentioned,
                           r.position_in_list, r.competitors_mentioned
                    FROM llm_monitoring_results r
                    JOIN llm_monitoring_queries q ON r.query_id = q.id
                    WHERE r.project_id = %s
                      AND r.analysis_date >= %s AND r.analysis_date < %s
                """
                prev_bparams_export = [project_id, prev_start_date_str, start_date_str]
                prev_bq_export, prev_bparams_export = _add_llm_filter(prev_bq_export, prev_bparams_export, 'r.llm_provider')
                cur.execute(prev_bq_export, prev_bparams_export)
                prev_branded_export = _compute_branded_metrics(cur.fetchall(), brand_keywords)
        except Exception as prev_b_err:
            logger.warning(f"Could not compute prev branded for export: {prev_b_err}")

        # Previous LLM metrics (avg mention_rate per provider)
        prev_llm_mr_export = {}
        try:
            prev_mr_query = """
                SELECT llm_provider, AVG(mention_rate) as avg_mention_rate
                FROM llm_monitoring_snapshots
                WHERE project_id = %s
                  AND snapshot_date >= %s AND snapshot_date < %s
            """
            prev_mr_params = [project_id, prev_start_date_str, start_date_str]
            prev_mr_query, prev_mr_params = _add_llm_filter(prev_mr_query, prev_mr_params)
            prev_mr_query += " GROUP BY llm_provider"
            cur.execute(prev_mr_query, prev_mr_params)
            for row in cur.fetchall():
                prev_llm_mr_export[row['llm_provider']] = float(row['avg_mention_rate']) if row['avg_mention_rate'] is not None else 0
        except Exception as prev_mr_err:
            logger.warning(f"Could not compute prev LLM MR for export: {prev_mr_err}")

        cur.close()
        conn.close()

        # 6. URL Rankings (via service)
        try:
            urls_ranking = LLMMonitoringStatsService.get_project_urls_ranking(
                project_id=project_id,
                days=days,
                enabled_llms=enabled_llms_filter if enabled_llms_filter else None,
                limit=100
            )
        except Exception as url_err:
            logger.warning(f"⚠️ Could not fetch URL rankings: {url_err}")
            urls_ranking = []

        # ──────────────────────────────────────────────────
        # BUILD EXCEL WORKBOOK
        # ──────────────────────────────────────────────────

        wb = openpyxl.Workbook()

        # — Shared styles —
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="161616", end_color="161616", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        subheader_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        title_font = Font(bold=True, size=14, color="161616")
        section_font = Font(bold=True, size=12, color="161616")
        brand_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        positive_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        negative_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(wrap_text=True, vertical='top')

        def write_header_row(ws, row, headers_list, col_start=1):
            """Write styled header row"""
            for col_idx, h in enumerate(headers_list, col_start):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

        def write_data_cell(ws, row, col, value, fmt=None):
            """Write data cell with border"""
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if fmt:
                cell.number_format = fmt
            return cell

        def auto_width(ws, min_width=10, max_width=60):
            """Auto-adjust column widths"""
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                adjusted = min(max(max_len + 2, min_width), max_width)
                ws.column_dimensions[col_letter].width = adjusted

        # ════════════════════════════════════════════════
        # SHEET 1: PROJECT SUMMARY
        # ════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Project Summary"

        ws1['A1'] = "LLM Monitoring Report"
        ws1['A1'].font = title_font
        ws1['A2'] = f"Project: {project['name']}"
        ws1['A2'].font = Font(bold=True, size=12)
        ws1['A3'] = f"Period: Last {days} days ({start_date.strftime('%Y-%m-%d')} → {end_date.strftime('%Y-%m-%d')})"
        ws1['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Project details
        ws1['A6'] = "Project Configuration"
        ws1['A6'].font = section_font
        details = [
            ("Brand Name", project.get('brand_name') or project['name']),
            ("Industry", project['industry'] or 'N/A'),
            ("Brand Domain", project.get('brand_domain') or 'N/A'),
            ("Brand Keywords", ', '.join(project.get('brand_keywords') or []) or 'N/A'),
            ("Language", project['language'] or 'N/A'),
            ("Country", project['country_code'] or 'N/A'),
            ("Enabled LLMs", ', '.join([llm.upper() for llm in (project.get('enabled_llms') or [])]) or 'All'),
            ("Queries per LLM", project.get('queries_per_llm') or 'N/A'),
            ("Status", "Active" if project.get('is_active') else "Paused"),
            ("Last Analysis", str(project.get('last_analysis_date') or 'Never')),
        ]
        for i, (label, value) in enumerate(details, 7):
            ws1[f'A{i}'] = label
            ws1[f'A{i}'].font = Font(bold=True)
            ws1[f'B{i}'] = str(value)

        # Competitors
        selected_competitors = project.get('selected_competitors') or []
        row_offset = 7 + len(details) + 1
        ws1[f'A{row_offset}'] = "Competitors"
        ws1[f'A{row_offset}'].font = section_font
        row_offset += 1
        if selected_competitors:
            ws1[f'A{row_offset}'] = "Domain"
            ws1[f'A{row_offset}'].font = Font(bold=True)
            ws1[f'B{row_offset}'] = "Keywords"
            ws1[f'B{row_offset}'].font = Font(bold=True)
            row_offset += 1
            for comp in selected_competitors:
                ws1[f'A{row_offset}'] = comp.get('domain', 'N/A')
                ws1[f'B{row_offset}'] = ', '.join(comp.get('keywords', []))
                row_offset += 1
        else:
            ws1[f'A{row_offset}'] = "No competitors configured"
            row_offset += 1

        # Global aggregated metrics summary
        row_offset += 1
        ws1[f'A{row_offset}'] = "Global Metrics Summary"
        ws1[f'A{row_offset}'].font = section_font
        row_offset += 1

        total_mentions_all = sum(m.get('total_mentions') or 0 for m in llm_metrics)
        total_results_all = sum(m.get('total_results') or 0 for m in llm_metrics)
        total_cost_all = sum(float(m.get('total_cost') or 0) for m in llm_metrics)
        total_tokens_all = sum(m.get('total_tokens') or 0 for m in llm_metrics)

        global_summary = [
            ("Total LLM Responses", total_results_all),
            ("Total Brand Mentions", total_mentions_all),
            ("Overall Mention Rate", f"{round(total_mentions_all / total_results_all * 100, 1)}%" if total_results_all > 0 else "N/A"),
            ("Total Cost (USD)", f"${round(total_cost_all, 4)}"),
            ("Total Tokens Used", total_tokens_all),
            ("Active Queries", len(queries)),
            ("URLs Cited", len(urls_ranking)),
        ]
        for label, value in global_summary:
            ws1[f'A{row_offset}'] = label
            ws1[f'A{row_offset}'].font = Font(bold=True)
            ws1[f'B{row_offset}'] = str(value)
            row_offset += 1

        # Period Comparison section (branded MR / SOV deltas)
        row_offset += 1
        ws1[f'A{row_offset}'] = "Period Comparison (vs Previous Period)"
        ws1[f'A{row_offset}'].font = section_font
        row_offset += 1

        if brand_keywords and detailed_results:
            branded_res_exp = []
            non_branded_res_exp = []
            for r in detailed_results:
                qt = r.get('query_text', '') or ''
                if classify_query_branded(qt, brand_keywords):
                    branded_res_exp.append(r)
                else:
                    non_branded_res_exp.append(r)
            def _quick_mr(subset):
                if not subset:
                    return 0.0
                mentions = sum(1 for r in subset if r.get('brand_mentioned'))
                return round((mentions / len(subset)) * 100, 1) if len(subset) > 0 else 0.0
            cur_br_mr = _quick_mr(branded_res_exp)
            cur_nb_mr = _quick_mr(non_branded_res_exp)
        else:
            cur_br_mr = 0.0
            cur_nb_mr = 0.0

        prev_br_metrics = prev_branded_export.get('branded_metrics') or {}
        prev_nb_metrics = prev_branded_export.get('non_branded_metrics') or {}
        prev_br_mr = prev_br_metrics.get('mention_rate', 0)
        prev_nb_mr = prev_nb_metrics.get('mention_rate', 0)

        comparison_rows = [
            ("Branded MR (Current)", f"{cur_br_mr}%"),
            ("Branded MR (Previous)", f"{prev_br_mr}%"),
            ("Branded MR Change", f"{round(cur_br_mr - prev_br_mr, 1)} pp"),
            ("Non-Branded MR (Current)", f"{cur_nb_mr}%"),
            ("Non-Branded MR (Previous)", f"{prev_nb_mr}%"),
            ("Non-Branded MR Change", f"{round(cur_nb_mr - prev_nb_mr, 1)} pp"),
        ]
        for label, value in comparison_rows:
            ws1[f'A{row_offset}'] = label
            ws1[f'A{row_offset}'].font = Font(bold=True)
            ws1[f'B{row_offset}'] = str(value)
            row_offset += 1

        ws1.column_dimensions['A'].width = 35
        ws1.column_dimensions['B'].width = 50

        # ════════════════════════════════════════════════
        # SHEET 2: SHARE OF VOICE (brand vs competitors over time)
        # ════════════════════════════════════════════════
        ws2 = wb.create_sheet("Share of Voice")

        # Build SOV data from snapshots
        # Aggregate across LLMs per date for brand SOV
        sov_by_date = {}
        all_competitor_domains = set()

        for snap in snapshots:
            d = str(snap['snapshot_date'])
            if d not in sov_by_date:
                sov_by_date[d] = {
                    'brand_sov': [], 'brand_weighted_sov': [],
                    'brand_mentions': 0, 'total_competitor_mentions': 0,
                    'competitor_breakdown': {}, 'weighted_competitor_breakdown': {}
                }

            sov_by_date[d]['brand_sov'].append(float(snap['share_of_voice'] or 0))
            sov_by_date[d]['brand_weighted_sov'].append(float(snap['weighted_share_of_voice'] or 0))
            sov_by_date[d]['brand_mentions'] += (snap['total_mentions'] or 0)
            sov_by_date[d]['total_competitor_mentions'] += (snap['total_competitor_mentions'] or 0)

            # Merge competitor breakdowns
            for field, target in [
                ('competitor_breakdown', 'competitor_breakdown'),
                ('weighted_competitor_breakdown', 'weighted_competitor_breakdown')
            ]:
                breakdown = snap.get(field) or {}
                if isinstance(breakdown, str):
                    try:
                        breakdown = json_module.loads(breakdown)
                    except:
                        breakdown = {}
                for comp_key, val in breakdown.items():
                    all_competitor_domains.add(comp_key)
                    if comp_key not in sov_by_date[d][target]:
                        sov_by_date[d][target][comp_key] = 0
                    sov_by_date[d][target][comp_key] += (val if isinstance(val, (int, float)) else 0)

        sorted_competitor_domains = sorted(all_competitor_domains)

        # Build header: Date, Brand SoV (Weighted), Brand SoV (Standard), Brand Mentions,
        #               Competitor1 SoV, Competitor1 Mentions, Competitor2 SoV, ...
        sov_headers = [
            "Date",
            "Brand SoV Weighted (%)", "Brand SoV Standard (%)",
            "Brand Mentions", "Total Competitor Mentions"
        ]
        for comp_d in sorted_competitor_domains:
            sov_headers.append(f"{comp_d} - Mentions")
            sov_headers.append(f"{comp_d} - Weighted SoV")

        write_header_row(ws2, 1, sov_headers)

        row_idx = 2
        for d in sorted(sov_by_date.keys()):
            data = sov_by_date[d]
            avg_wsov = round(sum(data['brand_weighted_sov']) / len(data['brand_weighted_sov']), 2) if data['brand_weighted_sov'] else 0
            avg_sov = round(sum(data['brand_sov']) / len(data['brand_sov']), 2) if data['brand_sov'] else 0

            col = 1
            write_data_cell(ws2, row_idx, col, d); col += 1
            write_data_cell(ws2, row_idx, col, avg_wsov, '0.00'); col += 1
            write_data_cell(ws2, row_idx, col, avg_sov, '0.00'); col += 1
            write_data_cell(ws2, row_idx, col, data['brand_mentions']); col += 1
            write_data_cell(ws2, row_idx, col, data['total_competitor_mentions']); col += 1

            # Competitor columns
            total_all = data['brand_mentions'] + data['total_competitor_mentions']
            for comp_d in sorted_competitor_domains:
                mentions = data['competitor_breakdown'].get(comp_d, 0)
                weighted = data['weighted_competitor_breakdown'].get(comp_d, 0)
                write_data_cell(ws2, row_idx, col, mentions); col += 1
                write_data_cell(ws2, row_idx, col, round(float(weighted), 2) if weighted else 0, '0.00'); col += 1

            row_idx += 1

        auto_width(ws2)

        # ════════════════════════════════════════════════
        # SHEET 3: LLM COMPARISON (aggregated per provider)
        # ════════════════════════════════════════════════
        ws3 = wb.create_sheet("LLM Comparison")

        comp_headers = [
            "LLM Provider", "Total Queries", "Total Results", "Total Mentions",
            "Mention Rate (%)", "Prev. MR (%)", "\u0394 MR (%)",
            "Avg Position",
            "Positive", "Neutral", "Negative", "Avg Sentiment Score",
            "SoV Weighted (%)", "SoV Standard (%)",
            "Total Cost (USD)", "Total Tokens", "Avg Response Time (ms)"
        ]
        write_header_row(ws3, 1, comp_headers)

        # Also compute avg SOV from snapshots per provider
        sov_per_provider = {}
        for snap in snapshots:
            prov = snap['llm_provider']
            if prov not in sov_per_provider:
                sov_per_provider[prov] = {'sov': [], 'wsov': []}
            sov_per_provider[prov]['sov'].append(float(snap['share_of_voice'] or 0))
            sov_per_provider[prov]['wsov'].append(float(snap['weighted_share_of_voice'] or 0))

        for row_idx, m in enumerate(llm_metrics, 2):
            prov = m['llm_provider']
            prov_sov = sov_per_provider.get(prov, {'sov': [0], 'wsov': [0]})
            avg_wsov = round(sum(prov_sov['wsov']) / len(prov_sov['wsov']), 2) if prov_sov['wsov'] else 0
            avg_sov = round(sum(prov_sov['sov']) / len(prov_sov['sov']), 2) if prov_sov['sov'] else 0

            cur_mr = round(float(m['mention_rate_pct'] or 0), 1)
            prev_mr = round(prev_llm_mr_export.get(prov, 0), 1)
            delta_mr = round(cur_mr - prev_mr, 1)

            col = 1
            write_data_cell(ws3, row_idx, col, prov.upper()); col += 1
            write_data_cell(ws3, row_idx, col, m['total_queries'] or 0); col += 1
            write_data_cell(ws3, row_idx, col, m['total_results'] or 0); col += 1
            write_data_cell(ws3, row_idx, col, m['total_mentions'] or 0); col += 1
            write_data_cell(ws3, row_idx, col, cur_mr, '0.0'); col += 1
            write_data_cell(ws3, row_idx, col, prev_mr, '0.0'); col += 1
            write_data_cell(ws3, row_idx, col, delta_mr, '0.0'); col += 1
            write_data_cell(ws3, row_idx, col, round(float(m['avg_position'] or 0), 1) if m['avg_position'] else 'N/A', '0.0'); col += 1
            write_data_cell(ws3, row_idx, col, m['positive_count'] or 0); col += 1
            write_data_cell(ws3, row_idx, col, m['neutral_count'] or 0); col += 1
            write_data_cell(ws3, row_idx, col, m['negative_count'] or 0); col += 1
            write_data_cell(ws3, row_idx, col, round(float(m['avg_sentiment_score'] or 0), 3) if m['avg_sentiment_score'] else 'N/A', '0.000'); col += 1
            write_data_cell(ws3, row_idx, col, avg_wsov, '0.00'); col += 1
            write_data_cell(ws3, row_idx, col, avg_sov, '0.00'); col += 1
            write_data_cell(ws3, row_idx, col, round(float(m['total_cost'] or 0), 4) if m['total_cost'] else 0, '0.0000'); col += 1
            write_data_cell(ws3, row_idx, col, m['total_tokens'] or 0); col += 1
            write_data_cell(ws3, row_idx, col, round(float(m['avg_response_time'] or 0), 0) if m['avg_response_time'] else 'N/A'); col += 1

        auto_width(ws3)

        # ════════════════════════════════════════════════
        # SHEET 4: DAILY METRICS (snapshots per date × LLM)
        # ════════════════════════════════════════════════
        ws4 = wb.create_sheet("Daily Metrics")

        daily_headers = [
            "Date", "LLM Provider",
            "Total Queries", "Total Mentions", "Mention Rate (%)",
            "Avg Position", "Top 3", "Top 5", "Top 10",
            "SoV Weighted (%)", "SoV Standard (%)",
            "Competitor Mentions",
            "Positive", "Neutral", "Negative", "Sentiment Score",
            "Cost (USD)", "Tokens", "Avg Response (ms)"
        ]
        write_header_row(ws4, 1, daily_headers)

        for row_idx, snap in enumerate(snapshots, 2):
            col = 1
            write_data_cell(ws4, row_idx, col, str(snap['snapshot_date'])); col += 1
            write_data_cell(ws4, row_idx, col, snap['llm_provider'].upper()); col += 1
            write_data_cell(ws4, row_idx, col, snap['total_queries'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, snap['total_mentions'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, round(float(snap['mention_rate'] or 0), 1), '0.0'); col += 1
            write_data_cell(ws4, row_idx, col, round(float(snap['avg_position'] or 0), 1) if snap['avg_position'] else 'N/A', '0.0'); col += 1
            write_data_cell(ws4, row_idx, col, snap['appeared_in_top3'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, snap['appeared_in_top5'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, snap['appeared_in_top10'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, round(float(snap['weighted_share_of_voice'] or 0), 2), '0.00'); col += 1
            write_data_cell(ws4, row_idx, col, round(float(snap['share_of_voice'] or 0), 2), '0.00'); col += 1
            write_data_cell(ws4, row_idx, col, snap['total_competitor_mentions'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, snap['positive_mentions'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, snap['neutral_mentions'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, snap['negative_mentions'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, round(float(snap['avg_sentiment_score'] or 0), 3) if snap['avg_sentiment_score'] else 'N/A', '0.000'); col += 1
            write_data_cell(ws4, row_idx, col, round(float(snap['total_cost_usd'] or 0), 4) if snap['total_cost_usd'] else 0, '0.0000'); col += 1
            write_data_cell(ws4, row_idx, col, snap['total_tokens'] or 0); col += 1
            write_data_cell(ws4, row_idx, col, snap['avg_response_time_ms'] or 0); col += 1

        auto_width(ws4)

        # ════════════════════════════════════════════════
        # SHEET 5: PROMPTS & QUERIES (enhanced)
        # ════════════════════════════════════════════════
        ws5 = wb.create_sheet("Prompts & Queries")

        export_country = project['country_code'] or 'Global'
        query_headers = [
            "Prompt", "Country", "Language", "Type",
            "LLMs Analyzed", "Total Results",
            "Total Mentions", "Text Mentions", "URL Citations",
            "Visibility (%)", "Avg Position",
            "Last Analysis", "Branded Query"
        ]
        write_header_row(ws5, 1, query_headers)

        for row_idx, q in enumerate(queries, 2):
            col = 1
            write_data_cell(ws5, row_idx, col, q['prompt']); col += 1
            write_data_cell(ws5, row_idx, col, export_country); col += 1
            write_data_cell(ws5, row_idx, col, q['language'] or 'N/A'); col += 1
            write_data_cell(ws5, row_idx, col, q['query_type'] or 'general'); col += 1
            write_data_cell(ws5, row_idx, col, q['llms_analyzed'] or 0); col += 1
            write_data_cell(ws5, row_idx, col, q['total_results'] or 0); col += 1
            write_data_cell(ws5, row_idx, col, q['total_mentions'] or 0); col += 1
            write_data_cell(ws5, row_idx, col, q['text_mentions'] or 0); col += 1
            write_data_cell(ws5, row_idx, col, q['url_citations'] or 0); col += 1
            write_data_cell(ws5, row_idx, col, round(float(q['visibility_pct'] or 0), 1), '0.0'); col += 1
            write_data_cell(ws5, row_idx, col, round(float(q['avg_position'] or 0), 1) if q['avg_position'] else 'N/A', '0.0'); col += 1
            write_data_cell(ws5, row_idx, col, str(q['last_analysis']) if q['last_analysis'] else 'N/A'); col += 1
            is_branded = classify_query_branded(q.get('prompt', '') or q.get('query_text', ''), brand_keywords)
            write_data_cell(ws5, row_idx, col, "Yes" if is_branded else "No"); col += 1

        ws5.column_dimensions['A'].width = 60
        auto_width(ws5, min_width=12)
        ws5.column_dimensions['A'].width = 60  # Override for prompt column

        # ════════════════════════════════════════════════
        # SHEET 6: URL RANKINGS
        # ════════════════════════════════════════════════
        ws6 = wb.create_sheet("URL Rankings")

        url_headers = ["Rank", "URL", "Total Mentions", "Percentage (%)", "Providers"]
        # Add per-LLM breakdown columns
        llm_provider_names = ['openai', 'anthropic', 'google', 'perplexity']
        for llm_name in llm_provider_names:
            url_headers.append(f"{llm_name.upper()} Mentions")

        write_header_row(ws6, 1, url_headers)

        for row_idx, url_data in enumerate(urls_ranking, 2):
            col = 1
            write_data_cell(ws6, row_idx, col, url_data.get('rank', row_idx - 1)); col += 1
            write_data_cell(ws6, row_idx, col, url_data.get('url', 'N/A')); col += 1
            write_data_cell(ws6, row_idx, col, url_data.get('mentions', 0)); col += 1
            write_data_cell(ws6, row_idx, col, round(float(url_data.get('percentage', 0)), 1), '0.0'); col += 1
            write_data_cell(ws6, row_idx, col, ', '.join(url_data.get('providers', []))); col += 1

            llm_breakdown = url_data.get('llm_breakdown', {})
            for llm_name in llm_provider_names:
                write_data_cell(ws6, row_idx, col, llm_breakdown.get(llm_name, 0)); col += 1

        ws6.column_dimensions['B'].width = 70
        auto_width(ws6, min_width=12)
        ws6.column_dimensions['B'].width = 70  # Override for URL column

        # ════════════════════════════════════════════════
        # SHEET 7: SENTIMENT ANALYSIS
        # ════════════════════════════════════════════════
        ws7 = wb.create_sheet("Sentiment Analysis")

        # Section A: Sentiment by LLM Provider (aggregated)
        ws7['A1'] = "Sentiment Distribution by LLM Provider"
        ws7['A1'].font = section_font

        sent_headers = [
            "LLM Provider", "Positive", "Neutral", "Negative", "Total",
            "Positive %", "Neutral %", "Negative %", "Avg Sentiment Score"
        ]
        write_header_row(ws7, 3, sent_headers)

        row_idx = 4
        for m in llm_metrics:
            pos = m['positive_count'] or 0
            neu = m['neutral_count'] or 0
            neg = m['negative_count'] or 0
            total = pos + neu + neg

            col = 1
            write_data_cell(ws7, row_idx, col, m['llm_provider'].upper()); col += 1
            c = write_data_cell(ws7, row_idx, col, pos); c.fill = positive_fill; col += 1
            write_data_cell(ws7, row_idx, col, neu); col += 1
            c = write_data_cell(ws7, row_idx, col, neg); c.fill = negative_fill; col += 1
            write_data_cell(ws7, row_idx, col, total); col += 1
            write_data_cell(ws7, row_idx, col, round(pos / total * 100, 1) if total > 0 else 0, '0.0'); col += 1
            write_data_cell(ws7, row_idx, col, round(neu / total * 100, 1) if total > 0 else 0, '0.0'); col += 1
            write_data_cell(ws7, row_idx, col, round(neg / total * 100, 1) if total > 0 else 0, '0.0'); col += 1
            write_data_cell(ws7, row_idx, col, round(float(m['avg_sentiment_score'] or 0), 3), '0.000'); col += 1
            row_idx += 1

        # Section B: Sentiment over time (from snapshots)
        row_idx += 2
        ws7.cell(row=row_idx, column=1, value="Sentiment Over Time (Daily)").font = section_font
        row_idx += 1

        sent_time_headers = ["Date", "LLM Provider", "Positive", "Neutral", "Negative", "Sentiment Score"]
        write_header_row(ws7, row_idx, sent_time_headers)
        row_idx += 1

        for snap in snapshots:
            col = 1
            write_data_cell(ws7, row_idx, col, str(snap['snapshot_date'])); col += 1
            write_data_cell(ws7, row_idx, col, snap['llm_provider'].upper()); col += 1
            write_data_cell(ws7, row_idx, col, snap['positive_mentions'] or 0); col += 1
            write_data_cell(ws7, row_idx, col, snap['neutral_mentions'] or 0); col += 1
            write_data_cell(ws7, row_idx, col, snap['negative_mentions'] or 0); col += 1
            write_data_cell(ws7, row_idx, col, round(float(snap['avg_sentiment_score'] or 0), 3) if snap['avg_sentiment_score'] else 'N/A', '0.000'); col += 1
            row_idx += 1

        auto_width(ws7)

        # ════════════════════════════════════════════════
        # SHEET 8: DETAILED RESULTS (per query × LLM × date)
        # ════════════════════════════════════════════════
        ws8 = wb.create_sheet("Detailed Results")

        detail_headers = [
            "Date", "LLM Provider", "Model", "Query/Prompt", "Branded Query",
            "Brand Mentioned", "Mention Count", "Position", "Total in List",
            "Position Source", "Sentiment", "Sentiment Score",
            "Competitors Mentioned", "URLs Cited",
            "Tokens", "Cost (USD)", "Response Time (ms)", "Response Length"
        ]
        write_header_row(ws8, 1, detail_headers)

        # Limit to 5000 rows to avoid memory issues
        max_detail_rows = min(len(detailed_results), 5000)

        for row_idx, r in enumerate(detailed_results[:max_detail_rows], 2):
            col = 1
            write_data_cell(ws8, row_idx, col, str(r['analysis_date'])); col += 1
            write_data_cell(ws8, row_idx, col, r['llm_provider'].upper()); col += 1
            write_data_cell(ws8, row_idx, col, r['model_used'] or 'N/A'); col += 1
            write_data_cell(ws8, row_idx, col, r['query_text'] or 'N/A'); col += 1
            is_branded_detail = classify_query_branded(r.get('query_text', '') or '', brand_keywords)
            write_data_cell(ws8, row_idx, col, "Yes" if is_branded_detail else "No"); col += 1
            write_data_cell(ws8, row_idx, col, "Yes" if r['brand_mentioned'] else "No"); col += 1
            write_data_cell(ws8, row_idx, col, r['mention_count'] or 0); col += 1
            write_data_cell(ws8, row_idx, col, r['position_in_list'] if r['position_in_list'] else 'N/A'); col += 1
            write_data_cell(ws8, row_idx, col, r['total_items_in_list'] if r['total_items_in_list'] else 'N/A'); col += 1
            write_data_cell(ws8, row_idx, col, r['position_source'] or 'N/A'); col += 1
            write_data_cell(ws8, row_idx, col, r['sentiment'] or 'N/A'); col += 1
            write_data_cell(ws8, row_idx, col, round(float(r['sentiment_score'] or 0), 3) if r['sentiment_score'] else 'N/A', '0.000'); col += 1

            # Competitors mentioned (JSON → readable string)
            comps = r.get('competitors_mentioned') or {}
            if isinstance(comps, str):
                try:
                    comps = json_module.loads(comps)
                except:
                    comps = {}
            if isinstance(comps, dict) and comps:
                comps_str = ', '.join([f"{k}: {v}" for k, v in comps.items()])
            else:
                comps_str = 'None'
            write_data_cell(ws8, row_idx, col, comps_str); col += 1

            # Sources/URLs (JSON → readable string)
            sources = r.get('sources') or []
            if isinstance(sources, str):
                try:
                    sources = json_module.loads(sources)
                except:
                    sources = []
            if isinstance(sources, list) and sources:
                urls_str = ', '.join([s.get('url', '') for s in sources if isinstance(s, dict) and s.get('url')])
            else:
                urls_str = 'None'
            write_data_cell(ws8, row_idx, col, urls_str); col += 1

            write_data_cell(ws8, row_idx, col, r['tokens_used'] or 0); col += 1
            write_data_cell(ws8, row_idx, col, round(float(r['cost_usd'] or 0), 6) if r['cost_usd'] else 0, '0.000000'); col += 1
            write_data_cell(ws8, row_idx, col, r['response_time_ms'] or 0); col += 1
            write_data_cell(ws8, row_idx, col, r['response_length'] or 0); col += 1

        # Add note if truncated
        if len(detailed_results) > max_detail_rows:
            note_row = max_detail_rows + 2
            ws8.cell(row=note_row, column=1,
                     value=f"⚠ Showing {max_detail_rows} of {len(detailed_results)} results. Full data available in the application.").font = Font(italic=True, color="666666")

        ws8.column_dimensions['D'].width = 60  # Query column
        ws8.column_dimensions['M'].width = 30  # Competitors
        ws8.column_dimensions['N'].width = 50  # URLs
        auto_width(ws8, min_width=12)
        ws8.column_dimensions['D'].width = 60
        ws8.column_dimensions['M'].width = 30
        ws8.column_dimensions['N'].width = 50

        # ════════════════════════════════════════════════
        # SHEET 9: BRANDED vs NON-BRANDED ANALYSIS
        # ════════════════════════════════════════════════
        ws9 = wb.create_sheet("Branded vs Non-Branded")

        # Title
        ws9.merge_cells('A1:J1')
        ws9['A1'] = 'Branded vs Non-Branded Analysis'
        ws9['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        ws9['A1'].fill = PatternFill(start_color='161616', end_color='161616', fill_type='solid')
        ws9['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws9.row_dimensions[1].height = 35

        # Classify all detailed results
        branded_results = []
        non_branded_results = []
        for r in detailed_results:
            qt = r.get('query_text', '') or ''
            if classify_query_branded(qt, brand_keywords):
                branded_results.append(r)
            else:
                non_branded_results.append(r)

        # Helper to compute metrics for a subset
        def compute_subset_metrics(subset, label, prev_mr=0.0):
            total = len(subset)
            mentions = sum(1 for r in subset if r.get('brand_mentioned'))
            mention_rate = round((mentions / total) * 100, 1) if total > 0 else 0.0
            delta_mr = round(mention_rate - prev_mr, 1)
            positions = [r['position_in_list'] for r in subset if r.get('position_in_list') is not None]
            avg_pos = round(sum(positions) / len(positions), 1) if positions else None
            pos_count = sum(1 for r in subset if r.get('sentiment') == 'positive')
            neu_count = sum(1 for r in subset if r.get('sentiment') == 'neutral')
            neg_count = sum(1 for r in subset if r.get('sentiment') == 'negative')
            return [label, total, mentions, mention_rate, round(prev_mr, 1), delta_mr,
                    avg_pos or 'N/A',
                    round(pos_count / total * 100, 1) if total else 0,
                    round(neu_count / total * 100, 1) if total else 0,
                    round(neg_count / total * 100, 1) if total else 0]

        # Headers
        bvnb_headers = ['Query Type', 'Total Results', 'Total Mentions', 'Mention Rate (%)',
                        'Prev. MR (%)', '\u0394 MR',
                        'Avg Position', 'Positive %', 'Neutral %', 'Negative %']
        for col_idx, h in enumerate(bvnb_headers, 1):
            cell = ws9.cell(row=3, column=col_idx, value=h)
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Previous period MRs for branded / non-branded
        prev_all_mr = 0.0
        prev_b_mr_exp = (prev_branded_export.get('branded_metrics') or {}).get('mention_rate', 0)
        prev_nb_mr_exp = (prev_branded_export.get('non_branded_metrics') or {}).get('mention_rate', 0)
        if prev_b_mr_exp or prev_nb_mr_exp:
            # Weighted average for "all" based on result counts
            prev_b_total = (prev_branded_export.get('branded_metrics') or {}).get('total_results', 0)
            prev_nb_total = (prev_branded_export.get('non_branded_metrics') or {}).get('total_results', 0)
            prev_all_total = prev_b_total + prev_nb_total
            if prev_all_total > 0:
                prev_all_mr = (prev_b_mr_exp * prev_b_total + prev_nb_mr_exp * prev_nb_total) / prev_all_total

        # Data rows
        all_metrics = compute_subset_metrics(detailed_results, 'All Queries', prev_all_mr)
        branded_m = compute_subset_metrics(branded_results, 'Branded', prev_b_mr_exp)
        non_branded_m = compute_subset_metrics(non_branded_results, 'Non-Branded', prev_nb_mr_exp)

        for row_idx, row_data in enumerate([all_metrics, non_branded_m, branded_m], 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws9.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                if row_idx == 5:  # Non-branded row highlight
                    cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                elif row_idx == 6:  # Branded row highlight
                    cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

        # Auto-width
        for col in ws9.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max(max_length + 2, 12), 30)
            ws9.column_dimensions[col[0].column_letter].width = adjusted_width

        # ──────────────────────────────────────────────────
        # SAVE & RETURN
        # ──────────────────────────────────────────────────

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = project['name'].replace(' ', '-').replace('/', '-')[:50]
        filename = f"llm-monitoring-{safe_name}-{days}d-{datetime.now().strftime('%Y%m%d')}.xlsx"

        logger.info(f"✅ Comprehensive Excel exported for project {project_id} ({len(snapshots)} snapshots, {len(queries)} queries, {len(detailed_results)} results, {len(urls_ranking)} URLs)")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"❌ Error exportando Excel para proyecto {project_id}: {e}", exc_info=True)
        return jsonify({'error': f'Error exportando Excel: {str(e)}'}), 500


@llm_monitoring_bp.route('/projects/<int:project_id>/export/pdf', methods=['GET'])
@login_required
@validate_project_ownership
def export_project_pdf(project_id):
    """
    Exportar datos del proyecto a PDF - Multi-page professional report

    Query params:
        days: int - Periodo de dias (default: 30)
    """
    from io import BytesIO
    from flask import send_file

    logger.info(f"Starting PDF export for project {project_id}")

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, KeepTogether
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        logger.info("reportlab imported successfully for PDF export")
    except ImportError as e:
        logger.error(f"reportlab no esta instalado: {e}")
        return jsonify({'error': 'PDF export not available. Missing reportlab library.'}), 500

    days = _normalize_days_param(request.args.get('days'), default=30)

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexion a BD'}), 500

    try:
        cur = conn.cursor()

        # ── 1. Project data ──
        cur.execute("""
            SELECT name, industry, brand_domain, brand_keywords, language,
                   country_code, enabled_llms, selected_competitors
            FROM llm_monitoring_projects
            WHERE id = %s
        """, (project_id,))
        project = cur.fetchone()

        if not project:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        prev_start = start_date - timedelta(days=days)
        enabled_llms_filter = project.get('enabled_llms') or []
        brand_keywords_pdf = project.get('brand_keywords') or []

        # ── Fetch current LLM models with knowledge cutoff ──
        cur.execute("""
            SELECT llm_provider, model_id, model_display_name,
                   knowledge_cutoff, knowledge_cutoff_date
            FROM llm_model_registry
            WHERE is_current = TRUE AND is_available = TRUE
            ORDER BY llm_provider
        """)
        current_models_pdf = cur.fetchall()

        # Helper to append LLM filter
        def _llm_filter(query, params):
            if enabled_llms_filter:
                query += " AND llm_provider = ANY(%s)"
                params.append(enabled_llms_filter)
            return query, params

        # ── 2. LLM results metrics (per provider) ──
        pdf_q = """
            SELECT llm_provider,
                   COUNT(DISTINCT query_id) as total_queries,
                   SUM(CASE WHEN brand_mentioned THEN 1 ELSE 0 END) as total_mentions,
                   ROUND(AVG(CASE WHEN brand_mentioned THEN 100.0 ELSE 0 END), 1) as mention_rate_pct
            FROM llm_monitoring_results
            WHERE project_id = %s AND analysis_date >= %s AND analysis_date <= %s
        """
        pdf_p = [project_id, start_date, end_date]
        pdf_q, pdf_p = _llm_filter(pdf_q, pdf_p)
        pdf_q += " GROUP BY llm_provider ORDER BY mention_rate_pct DESC"
        cur.execute(pdf_q, pdf_p)
        metrics = cur.fetchall()

        # ── 3. Snapshot aggregates (SOV, sentiment, position) ──
        snap_q = """
            SELECT llm_provider,
                AVG(mention_rate) as avg_mr,
                AVG(share_of_voice) as avg_sov,
                AVG(avg_position) as avg_pos,
                AVG(avg_sentiment_score) as avg_sentiment,
                SUM(positive_mentions) as total_positive,
                SUM(neutral_mentions) as total_neutral,
                SUM(negative_mentions) as total_negative,
                SUM(total_queries) as total_queries
            FROM llm_monitoring_snapshots
            WHERE project_id = %s AND snapshot_date >= %s AND snapshot_date <= %s
        """
        snap_p = [project_id, start_date.date(), end_date.date()]
        snap_q, snap_p = _llm_filter(snap_q, snap_p)
        snap_q += " GROUP BY llm_provider ORDER BY avg_mr DESC"
        cur.execute(snap_q, snap_p)
        snapshot_metrics = {row['llm_provider']: row for row in cur.fetchall()}

        # Previous period snapshot aggregates for comparison
        prev_snap_q = """
            SELECT AVG(mention_rate) as avg_mr,
                   AVG(share_of_voice) as avg_sov
            FROM llm_monitoring_snapshots
            WHERE project_id = %s AND snapshot_date >= %s AND snapshot_date < %s
        """
        prev_snap_p = [project_id, prev_start.date(), start_date.date()]
        prev_snap_q, prev_snap_p = _llm_filter(prev_snap_q, prev_snap_p)
        cur.execute(prev_snap_q, prev_snap_p)
        prev_snap_agg = cur.fetchone() or {}

        # ── 4. Branded vs Non-Branded results ──
        bvnb_q = """
            SELECT q.query_text, r.brand_mentioned, r.sentiment
            FROM llm_monitoring_results r
            JOIN llm_monitoring_queries q ON r.query_id = q.id
            WHERE r.project_id = %s AND r.analysis_date >= %s AND r.analysis_date <= %s
        """
        bvnb_p = [project_id, start_date, end_date]
        bvnb_q, bvnb_p = _llm_filter(bvnb_q, bvnb_p)
        cur.execute(bvnb_q, bvnb_p)
        bvnb_results = cur.fetchall()

        # Previous period branded/non-branded
        prev_bvnb_q = """
            SELECT q.query_text, r.brand_mentioned, r.sentiment
            FROM llm_monitoring_results r
            JOIN llm_monitoring_queries q ON r.query_id = q.id
            WHERE r.project_id = %s AND r.analysis_date >= %s AND r.analysis_date < %s
        """
        prev_bvnb_p = [project_id, prev_start, start_date]
        prev_bvnb_q, prev_bvnb_p = _llm_filter(prev_bvnb_q, prev_bvnb_p)
        prev_branded_pdf = []
        prev_non_branded_pdf = []
        try:
            cur.execute(prev_bvnb_q, prev_bvnb_p)
            for r in cur.fetchall():
                qt = r.get('query_text', '') or ''
                if classify_query_branded(qt, brand_keywords_pdf):
                    prev_branded_pdf.append(r)
                else:
                    prev_non_branded_pdf.append(r)
        except Exception:
            logger.warning("Could not compute prev branded for PDF")

        # Previous period LLM MR
        prev_llm_mr = {}
        try:
            plmr_q = """
                SELECT llm_provider,
                       ROUND(AVG(CASE WHEN brand_mentioned THEN 100.0 ELSE 0 END), 1) as mention_rate_pct
                FROM llm_monitoring_results
                WHERE project_id = %s AND analysis_date >= %s AND analysis_date < %s
            """
            plmr_p = [project_id, prev_start, start_date]
            plmr_q, plmr_p = _llm_filter(plmr_q, plmr_p)
            plmr_q += " GROUP BY llm_provider"
            cur.execute(plmr_q, plmr_p)
            for row in cur.fetchall():
                prev_llm_mr[row['llm_provider']] = float(row['mention_rate_pct'] or 0)
        except Exception:
            logger.warning("Could not compute prev LLM MR for PDF")

        # ── 5. Prompt / query performance ──
        prompt_q = """
            SELECT q.query_text,
                COUNT(r.id) as total_results,
                SUM(CASE WHEN r.brand_mentioned THEN 1 ELSE 0 END) as mentions,
                AVG(CASE WHEN r.brand_mentioned THEN r.position_in_list ELSE NULL END) as avg_position
            FROM llm_monitoring_results r
            JOIN llm_monitoring_queries q ON r.query_id = q.id
            WHERE r.project_id = %s AND r.analysis_date >= %s AND r.analysis_date <= %s
        """
        prompt_p = [project_id, start_date, end_date]
        prompt_q, prompt_p = _llm_filter(prompt_q, prompt_p)
        prompt_q += " GROUP BY q.query_text ORDER BY mentions DESC, total_results DESC LIMIT 20"
        cur.execute(prompt_q, prompt_p)
        prompt_data = cur.fetchall()

        # ── 6. Top URLs (via service) ──
        urls_data = []
        try:
            urls_data = LLMMonitoringStatsService.get_project_urls_ranking(
                project_id=project_id,
                days=days,
                enabled_llms=enabled_llms_filter if enabled_llms_filter else None,
                limit=15
            )
        except Exception as url_err:
            logger.warning(f"Could not fetch URL rankings for PDF: {url_err}")

        # ── 7. Competitor snapshot breakdown ──
        comp_snapshot_q = """
            SELECT snapshot_date, llm_provider,
                   total_mentions, total_competitor_mentions,
                   competitor_breakdown
            FROM llm_monitoring_snapshots
            WHERE project_id = %s AND snapshot_date >= %s AND snapshot_date <= %s
        """
        comp_snap_p = [project_id, start_date.date(), end_date.date()]
        comp_snapshot_q, comp_snap_p = _llm_filter(comp_snapshot_q, comp_snap_p)
        comp_snapshot_q += " ORDER BY snapshot_date"
        cur.execute(comp_snapshot_q, comp_snap_p)
        comp_snapshots = cur.fetchall()

        cur.close()
        conn.close()

        # ── Aggregate competitor mentions ──
        brand_total_mentions = 0
        competitor_mention_totals = {}
        num_days_with_data = set()
        for snap in comp_snapshots:
            brand_total_mentions += int(snap.get('total_mentions') or 0)
            num_days_with_data.add(str(snap.get('snapshot_date', '')))
            breakdown = snap.get('competitor_breakdown') or {}
            if isinstance(breakdown, str):
                import json as json_mod
                try:
                    breakdown = json_mod.loads(breakdown)
                except Exception:
                    breakdown = {}
            for comp_key, count in breakdown.items():
                comp_key_lower = comp_key.lower().strip()
                competitor_mention_totals[comp_key_lower] = competitor_mention_totals.get(comp_key_lower, 0) + int(count or 0)

        days_count = max(len(num_days_with_data), 1)

        # Map competitor_breakdown keys to selected_competitors domains
        comp_display_map = {}  # normalized_key -> display_name
        for comp in (project.get('selected_competitors') or []):
            domain = (comp.get('domain') or '').lower().strip()
            keywords = comp.get('keywords') or []
            display = domain or (keywords[0] if keywords else 'Unknown')
            # Map domain and keywords to this display name
            for variant in [domain] + [k.lower().strip() for k in keywords]:
                if variant:
                    comp_display_map[variant] = display
                    # Also strip TLD
                    for tld in ['.com', '.es', '.mx', '.org', '.net', '.io', '.co']:
                        if variant.endswith(tld):
                            comp_display_map[variant.replace(tld, '')] = display

        # Consolidate competitor mentions by display name
        comp_consolidated = {}
        for key, count in competitor_mention_totals.items():
            display = comp_display_map.get(key, key)
            comp_consolidated[display] = comp_consolidated.get(display, 0) + count

        # ── Compute aggregate KPIs ──
        total_positive = sum(float(s.get('total_positive') or 0) for s in snapshot_metrics.values())
        total_neutral = sum(float(s.get('total_neutral') or 0) for s in snapshot_metrics.values())
        total_negative = sum(float(s.get('total_negative') or 0) for s in snapshot_metrics.values())
        sentiment_total = total_positive + total_neutral + total_negative

        if sentiment_total > 0:
            best_sent = max(
                [('Positive', total_positive), ('Neutral', total_neutral), ('Negative', total_negative)],
                key=lambda x: x[1]
            )
            dominant_sentiment = best_sent[0]
        else:
            dominant_sentiment = 'N/A'

        # Weighted averages across providers
        agg_mr_values = [float(s.get('avg_mr') or 0) for s in snapshot_metrics.values() if s.get('avg_mr') is not None]
        agg_sov_values = [float(s.get('avg_sov') or 0) for s in snapshot_metrics.values() if s.get('avg_sov') is not None]
        agg_pos_values = [float(s.get('avg_pos') or 0) for s in snapshot_metrics.values() if s.get('avg_pos') is not None and float(s.get('avg_pos') or 0) > 0]

        overall_mr = round(sum(agg_mr_values) / len(agg_mr_values), 1) if agg_mr_values else 0
        overall_sov = round(sum(agg_sov_values) / len(agg_sov_values), 1) if agg_sov_values else 0
        overall_pos = round(sum(agg_pos_values) / len(agg_pos_values), 1) if agg_pos_values else 0

        prev_mr_agg = round(float(prev_snap_agg.get('avg_mr') or 0), 1)
        prev_sov_agg = round(float(prev_snap_agg.get('avg_sov') or 0), 1)

        # ── Classify branded / non-branded ──
        branded_pdf = []
        non_branded_pdf = []
        for r in bvnb_results:
            qt = r.get('query_text', '') or ''
            if classify_query_branded(qt, brand_keywords_pdf):
                branded_pdf.append(r)
            else:
                non_branded_pdf.append(r)

        # ── Competitors list ──
        selected_competitors = project.get('selected_competitors') or []
        competitor_names = []
        for comp in selected_competitors:
            domain = comp.get('domain', '')
            if domain:
                competitor_names.append(domain)

        # =====================================================================
        # BUILD PDF
        # =====================================================================
        output = BytesIO()

        # Color palette
        CLR_DARK = colors.HexColor('#161616')
        CLR_WHITE = colors.white
        CLR_ACCENT = colors.HexColor('#D8F9B8')
        CLR_SUBHEADER = colors.HexColor('#F3F4F6')
        CLR_GREEN_CELL = colors.HexColor('#D1FAE5')
        CLR_YELLOW_CELL = colors.HexColor('#FEF3C7')
        CLR_RED_CELL = colors.HexColor('#FEE2E2')
        CLR_BODY = colors.HexColor('#374151')
        CLR_BORDER = colors.HexColor('#E5E7EB')
        CLR_LIGHT_GRAY = colors.HexColor('#9CA3AF')
        CLR_ROW_ALT = colors.HexColor('#F9FAFB')

        page_width, page_height = A4
        usable_width = page_width - 4 * cm  # 2cm margins each side

        # Page header / footer callbacks
        total_pages_holder = [0]

        # Logo path for watermark
        import os as _os
        _logo_path = _os.path.join(_os.path.dirname(__file__), 'static', 'images', 'logos', 'logo clicandseo.png')
        _logo_exists = _os.path.exists(_logo_path)

        def _draw_page_chrome(canvas_obj, doc_obj):
            """Shared header + footer + watermark for all pages."""
            canvas_obj.saveState()
            # Dark header bar
            canvas_obj.setFillColor(CLR_DARK)
            canvas_obj.rect(0, page_height - 1.2 * cm, page_width, 1.2 * cm, fill=1, stroke=0)
            canvas_obj.setFillColor(CLR_WHITE)
            canvas_obj.setFont('Helvetica-Bold', 10)
            canvas_obj.drawString(2 * cm, page_height - 0.85 * cm, "LLM Visibility Monitor Report")
            # Footer text
            canvas_obj.setFillColor(CLR_LIGHT_GRAY)
            canvas_obj.setFont('Helvetica', 8)
            footer_text = f"Generated by ClicAndSEO LLM Visibility Monitor | Page {doc_obj.page}"
            canvas_obj.drawCentredString(page_width / 2, 1 * cm, footer_text)
            # Watermark logo bottom-right
            if _logo_exists:
                try:
                    canvas_obj.drawImage(
                        _logo_path,
                        page_width - 4.5 * cm, 0.4 * cm,
                        width=3.5 * cm, height=1 * cm,
                        preserveAspectRatio=True, mask='auto'
                    )
                except Exception:
                    pass  # Silently skip if image can't be drawn
            canvas_obj.restoreState()

        def _on_first_page(canvas_obj, doc_obj):
            _draw_page_chrome(canvas_obj, doc_obj)

        def _on_later_pages(canvas_obj, doc_obj):
            _draw_page_chrome(canvas_obj, doc_obj)

        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2.2 * cm,
            bottomMargin=1.8 * cm
        )

        styles = getSampleStyleSheet()

        # ── Custom styles ──
        st_title = ParagraphStyle('PDFTitle', parent=styles['Heading1'],
                                  fontSize=22, fontName='Helvetica-Bold',
                                  textColor=CLR_DARK, spaceAfter=6, spaceBefore=0)
        st_project = ParagraphStyle('PDFProject', parent=styles['Heading1'],
                                    fontSize=18, fontName='Helvetica-Bold',
                                    textColor=CLR_DARK, spaceAfter=4)
        st_period = ParagraphStyle('PDFPeriod', parent=styles['Normal'],
                                   fontSize=10, fontName='Helvetica',
                                   textColor=CLR_LIGHT_GRAY, spaceAfter=14)
        st_section = ParagraphStyle('PDFSection', parent=styles['Heading2'],
                                    fontSize=14, fontName='Helvetica-Bold',
                                    textColor=CLR_DARK, spaceAfter=8, spaceBefore=12)
        st_subsection = ParagraphStyle('PDFSubsection', parent=styles['Heading3'],
                                       fontSize=11, fontName='Helvetica-Bold',
                                       textColor=CLR_BODY, spaceAfter=6, spaceBefore=8)
        st_body = ParagraphStyle('PDFBody', parent=styles['Normal'],
                                 fontSize=9, fontName='Helvetica',
                                 textColor=CLR_BODY, spaceAfter=4)
        st_kpi_label = ParagraphStyle('KPILabel', parent=styles['Normal'],
                                      fontSize=8, fontName='Helvetica',
                                      textColor=CLR_LIGHT_GRAY, spaceAfter=0, alignment=TA_CENTER)
        st_kpi_value = ParagraphStyle('KPIValue', parent=styles['Normal'],
                                      fontSize=16, fontName='Helvetica-Bold',
                                      textColor=CLR_DARK, spaceAfter=0, alignment=TA_CENTER)
        st_kpi_delta = ParagraphStyle('KPIDelta', parent=styles['Normal'],
                                      fontSize=8, fontName='Helvetica',
                                      textColor=CLR_LIGHT_GRAY, spaceAfter=0, alignment=TA_CENTER)
        st_no_data = ParagraphStyle('NoData', parent=styles['Normal'],
                                    fontSize=10, fontName='Helvetica',
                                    textColor=CLR_LIGHT_GRAY, spaceAfter=8,
                                    alignment=TA_CENTER)

        # Reusable table style builder
        def _base_table_style(num_rows):
            """Returns a standard TableStyle list for dark-header tables."""
            style_cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), CLR_DARK),
                ('TEXTCOLOR', (0, 0), (-1, 0), CLR_WHITE),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('TEXTCOLOR', (0, 1), (-1, -1), CLR_BODY),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, CLR_BORDER),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            # Alternating row colors
            for i in range(1, num_rows):
                if i % 2 == 0:
                    style_cmds.append(('BACKGROUND', (0, i), (-1, i), CLR_ROW_ALT))
            return style_cmds

        def _truncate(text, max_len):
            """Truncate text with ellipsis."""
            if not text:
                return 'N/A'
            text = str(text)
            if len(text) <= max_len:
                return text
            return text[:max_len - 3] + '...'

        def _delta_str(current, previous):
            """Format delta string for percentage point change."""
            d = round(current - previous, 1)
            if previous == 0 and current == 0:
                return 'N/A'
            sign = '+' if d > 0 else ''
            return f"{sign}{d} pp"

        def _sentiment_label(score):
            """Convert numeric sentiment score to label."""
            if score is None:
                return 'N/A'
            s = float(score)
            if s >= 0.3:
                return 'Positive'
            elif s <= -0.3:
                return 'Negative'
            return 'Neutral'

        # Delta color helper for KPI paragraphs
        CLR_DELTA_UP = colors.HexColor('#059669')
        CLR_DELTA_DOWN = colors.HexColor('#DC2626')
        st_kpi_delta_up = ParagraphStyle('KPIDeltaUp', parent=st_kpi_delta, textColor=CLR_DELTA_UP)
        st_kpi_delta_down = ParagraphStyle('KPIDeltaDown', parent=st_kpi_delta, textColor=CLR_DELTA_DOWN)

        def _delta_style(current, previous):
            """Return (text, style) for delta display with color."""
            d = round(current - previous, 1)
            if previous == 0 and current == 0:
                return 'N/A', st_kpi_delta
            if previous == 0 and current > 0:
                return f'+{current} pp', st_kpi_delta_up
            sign = '+' if d > 0 else ''
            style = st_kpi_delta_up if d > 0 else (st_kpi_delta_down if d < 0 else st_kpi_delta)
            return f"vs prev: {sign}{d} pp", style

        elements = []

        # =================================================================
        # PAGE 1: PROJECT DETAILS
        # =================================================================
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(Paragraph("LLM Visibility Monitor Report", st_title))
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph(project['name'], st_project))
        elements.append(Paragraph(
            f"Period: Last {days} days  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            st_period
        ))
        elements.append(Spacer(1, 0.8 * cm))

        # ── Project Details ──
        elements.append(Paragraph("Project Details", st_section))
        details_info = [
            f"<b>Industry:</b> {project.get('industry') or 'N/A'}",
            f"<b>Domain:</b> {project.get('brand_domain') or 'N/A'}",
            f"<b>Keywords:</b> {', '.join(brand_keywords_pdf) if brand_keywords_pdf else 'N/A'}",
            f"<b>Language:</b> {project.get('language') or 'N/A'}  |  <b>Country:</b> {project.get('country_code') or 'N/A'}",
        ]
        for info in details_info:
            elements.append(Paragraph(info, st_body))
        elements.append(Spacer(1, 0.4 * cm))

        # ── LLM Models used ──
        elements.append(Paragraph("LLM Models Used", st_subsection))
        elements.append(Spacer(1, 0.2 * cm))
        try:
            conn_models = get_db_connection()
            cur_models = conn_models.cursor()
            cur_models.execute("""
                SELECT llm_provider, model_id, model_display_name, knowledge_cutoff
                FROM llm_model_registry
                WHERE is_current = TRUE
                ORDER BY llm_provider
            """)
            current_models = cur_models.fetchall()
            cur_models.close()
            conn_models.close()
        except Exception:
            current_models = []

        if current_models:
            model_header = ['LLM Provider', 'Model', 'Knowledge Cutoff']
            model_rows = [model_header]
            provider_display = {'openai': 'ChatGPT', 'anthropic': 'Claude', 'google': 'Gemini', 'perplexity': 'Perplexity'}
            for m in current_models:
                prov = provider_display.get(m.get('llm_provider', ''), m.get('llm_provider', ''))
                model_name = m.get('model_display_name') or m.get('model_id', 'N/A')
                cutoff = m.get('knowledge_cutoff') or 'Unknown'
                model_rows.append([prov, model_name, cutoff])
            model_widths = [3.5 * cm, 5.5 * cm, 6 * cm]
            model_table = Table(model_rows, colWidths=model_widths)
            model_style = _base_table_style(len(model_rows))
            model_style.append(('ALIGN', (0, 1), (-1, -1), 'LEFT'))
            model_table.setStyle(TableStyle(model_style))
            elements.append(model_table)
        else:
            elements.append(Paragraph("Model information not available.", st_body))
        elements.append(Spacer(1, 0.5 * cm))

        # ── Competitors list ──
        if competitor_names:
            elements.append(Paragraph("Competitors", st_subsection))
            elements.append(Paragraph(', '.join(competitor_names), st_body))
            elements.append(Spacer(1, 0.5 * cm))

        # ── Competitor Details table ──
        if selected_competitors:
            elements.append(Paragraph("Competitor Details", st_subsection))
            elements.append(Spacer(1, 0.2 * cm))
            comp_detail_header = ['Competitor', 'Domain', 'Keywords']
            comp_detail_rows = [comp_detail_header]
            for i, comp in enumerate(selected_competitors):
                domain = comp.get('domain', 'N/A')
                keywords = ', '.join(comp.get('keywords', [])) or 'N/A'
                comp_detail_rows.append([
                    f"Competitor {i + 1}",
                    domain,
                    Paragraph(_truncate(keywords, 45), st_body),
                ])

            cd_widths = [3 * cm, 4.5 * cm, 7.5 * cm]
            cd_table = Table(comp_detail_rows, colWidths=cd_widths)
            cd_style = _base_table_style(len(comp_detail_rows))
            cd_style.append(('ALIGN', (0, 1), (0, -1), 'LEFT'))
            cd_style.append(('ALIGN', (1, 1), (1, -1), 'LEFT'))
            cd_style.append(('ALIGN', (2, 1), (2, -1), 'LEFT'))
            cd_table.setStyle(TableStyle(cd_style))
            elements.append(cd_table)

        # ── LLM Models Used ──
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(Paragraph("LLM Models Used", st_subsection))
        elements.append(Spacer(1, 0.2 * cm))

        if current_models_pdf:
            provider_labels = {
                'openai': 'ChatGPT', 'anthropic': 'Claude',
                'google': 'Gemini', 'perplexity': 'Perplexity'
            }
            model_rows = [['Provider', 'Model', 'Knowledge Cutoff']]
            for m in current_models_pdf:
                prov = m.get('llm_provider', '')
                if enabled_llms_filter and prov not in enabled_llms_filter:
                    continue
                label = provider_labels.get(prov, prov.title())
                model_name = m.get('model_display_name') or m.get('model_id', 'N/A')
                cutoff = m.get('knowledge_cutoff') or 'Unknown'
                model_rows.append([label, model_name, cutoff])

            if len(model_rows) > 1:
                m_widths = [3.5 * cm, 5 * cm, 6.5 * cm]
                m_table = Table(model_rows, colWidths=m_widths)
                m_style = _base_table_style(len(model_rows))
                m_style.append(('ALIGN', (0, 1), (-1, -1), 'LEFT'))
                m_table.setStyle(TableStyle(m_style))
                elements.append(m_table)
            else:
                elements.append(Paragraph("No model data available.", st_no_data))
        else:
            elements.append(Paragraph("No model data available.", st_no_data))

        # =================================================================
        # PAGE 2: EXECUTIVE SUMMARY + LLM PERFORMANCE
        # =================================================================
        elements.append(PageBreak())
        elements.append(Spacer(1, 0.3 * cm))

        # ── Executive Summary KPIs (2x2 grid with colored deltas) ──
        elements.append(Paragraph("Executive Summary", st_section))

        mr_delta_text, mr_delta_st = _delta_style(overall_mr, prev_mr_agg)
        sov_delta_text, sov_delta_st = _delta_style(overall_sov, prev_sov_agg)

        kpi_data = [
            [
                Paragraph(f"<b>{overall_mr}%</b>", st_kpi_value),
                Paragraph(f"<b>{overall_sov}%</b>", st_kpi_value),
            ],
            [
                Paragraph("Mention Rate", st_kpi_label),
                Paragraph("Share of Voice", st_kpi_label),
            ],
            [
                Paragraph(mr_delta_text, mr_delta_st),
                Paragraph(sov_delta_text, sov_delta_st),
            ],
            [Spacer(1, 0.2 * cm), Spacer(1, 0.2 * cm)],
            [
                Paragraph(f"<b>#{overall_pos}</b>" if overall_pos > 0 else "<b>N/A</b>", st_kpi_value),
                Paragraph(f"<b>{dominant_sentiment}</b>", st_kpi_value),
            ],
            [
                Paragraph("Avg Position (when mentioned)", st_kpi_label),
                Paragraph("Dominant Sentiment", st_kpi_label),
            ],
        ]

        kpi_col_w = usable_width / 2
        kpi_table = Table(kpi_data, colWidths=[kpi_col_w, kpi_col_w])
        kpi_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (0, 2), 1, CLR_BORDER),
            ('BOX', (1, 0), (1, 2), 1, CLR_BORDER),
            ('BOX', (0, 4), (0, 5), 1, CLR_BORDER),
            ('BOX', (1, 4), (1, 5), 1, CLR_BORDER),
            ('BACKGROUND', (0, 0), (0, 2), CLR_SUBHEADER),
            ('BACKGROUND', (1, 0), (1, 2), CLR_SUBHEADER),
            ('BACKGROUND', (0, 4), (0, 5), CLR_SUBHEADER),
            ('BACKGROUND', (1, 4), (1, 5), CLR_SUBHEADER),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.5 * cm))

        # ── LLM Performance Comparison ──
        elements.append(Paragraph("LLM Performance Comparison", st_section))

        if metrics:
            perf_header = ["LLM", "Prompts", "Mentions", "MR (%)", "SOV (%)", "Avg Pos", "Sentiment", "vs Prev MR"]
            perf_rows = [perf_header]
            row_meta = []  # (sentiment_label, delta_value) per data row

            for idx, m in enumerate(metrics):
                provider = m['llm_provider']
                cur_mr = round(float(m['mention_rate_pct'] or 0), 1)
                prev_mr_val = round(prev_llm_mr.get(provider, 0), 1)
                snap = snapshot_metrics.get(provider, {})
                sov_val = round(float(snap.get('avg_sov') or 0), 1)
                pos_val = round(float(snap.get('avg_pos') or 0), 1)
                sent_score = snap.get('avg_sentiment')
                sent_label = _sentiment_label(sent_score)

                delta_val = round(cur_mr - prev_mr_val, 1)
                delta_sign = '+' if delta_val > 0 else ''
                delta_text = f"{delta_sign}{delta_val} pp" if (prev_mr_val > 0 or cur_mr > 0) else 'new'

                row_meta.append((sent_label, delta_val))

                perf_rows.append([
                    provider.upper(),
                    str(m['total_queries'] or 0),
                    str(m['total_mentions'] or 0),
                    f"{cur_mr}%",
                    f"{sov_val}%",
                    f"#{pos_val}" if pos_val > 0 else 'N/A',
                    sent_label,
                    delta_text,
                ])

            perf_widths = [2 * cm, 1.6 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 2 * cm, 2 * cm]
            perf_table = Table(perf_rows, colWidths=perf_widths)
            style_cmds = _base_table_style(len(perf_rows))
            # Color sentiment and delta columns
            for row_idx, (sent, dv) in enumerate(row_meta):
                r = row_idx + 1
                if sent == 'Positive':
                    style_cmds.append(('BACKGROUND', (6, r), (6, r), CLR_GREEN_CELL))
                elif sent == 'Negative':
                    style_cmds.append(('BACKGROUND', (6, r), (6, r), CLR_RED_CELL))
                # Color delta column (col 7) green/red
                if dv > 0:
                    style_cmds.append(('TEXTCOLOR', (7, r), (7, r), CLR_DELTA_UP))
                elif dv < 0:
                    style_cmds.append(('TEXTCOLOR', (7, r), (7, r), CLR_DELTA_DOWN))
            perf_table.setStyle(TableStyle(style_cmds))
            elements.append(perf_table)
        else:
            elements.append(Paragraph("No LLM performance data available for this period.", st_no_data))

        elements.append(Spacer(1, 0.8 * cm))

        # ── Branded vs Non-Branded Analysis ──
        elements.append(Paragraph("Branded vs Non-Branded Analysis", st_section))

        bvnb_header = ['Prompt Type', 'Responses', 'Mentions', 'MR (%)', 'vs Prev']
        bvnb_rows = [bvnb_header]
        bvnb_deltas = []

        prev_subsets = {'Non-Branded': prev_non_branded_pdf, 'Branded': prev_branded_pdf}
        for label, subset in [('Non-Branded', non_branded_pdf), ('Branded', branded_pdf)]:
            total = len(subset)
            mentions = sum(1 for r in subset if r.get('brand_mentioned'))
            rate = round((mentions / total) * 100, 1) if total > 0 else 0
            prev_sub = prev_subsets.get(label, [])
            prev_total = len(prev_sub)
            prev_mentions = sum(1 for r in prev_sub if r.get('brand_mentioned'))
            prev_rate = round((prev_mentions / prev_total) * 100, 1) if prev_total > 0 else 0
            dv = round(rate - prev_rate, 1)
            bvnb_deltas.append(dv)
            bvnb_rows.append([label, str(total), str(mentions), f"{rate}%", _delta_str(rate, prev_rate)])

        bvnb_widths = [3.2 * cm, 2.2 * cm, 2.2 * cm, 2.5 * cm, 2.5 * cm]
        bvnb_table = Table(bvnb_rows, colWidths=bvnb_widths)
        bvnb_style = _base_table_style(len(bvnb_rows))
        # Non-branded row green, branded row yellow
        bvnb_style.append(('BACKGROUND', (0, 1), (-1, 1), CLR_GREEN_CELL))
        bvnb_style.append(('BACKGROUND', (0, 2), (-1, 2), CLR_YELLOW_CELL))
        # Color delta column
        for i, dv in enumerate(bvnb_deltas):
            r = i + 1
            if dv > 0:
                bvnb_style.append(('TEXTCOLOR', (4, r), (4, r), CLR_DELTA_UP))
            elif dv < 0:
                bvnb_style.append(('TEXTCOLOR', (4, r), (4, r), CLR_DELTA_DOWN))
        bvnb_table.setStyle(TableStyle(bvnb_style))
        elements.append(bvnb_table)

        # =================================================================
        # PAGE 3: COMPETITOR ANALYSIS
        # =================================================================
        elements.append(PageBreak())
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph("Competitor Analysis", st_section))
        elements.append(Paragraph("Brand vs competitors — Share of Voice comparison", st_body))
        elements.append(Spacer(1, 0.3 * cm))

        # Build brand vs competitors SOV table
        all_entity_mentions = brand_total_mentions + sum(comp_consolidated.values())
        sorted_comps = sorted(comp_consolidated.items(), key=lambda x: x[1], reverse=True) if comp_consolidated else []

        if all_entity_mentions > 0:
            sov_header = ['Brand / Competitor', 'Total Mentions', 'Mention Share (%)', 'Avg Mentions/Day']
            sov_rows = [sov_header]

            brand_share = round((brand_total_mentions / all_entity_mentions) * 100, 1)
            brand_avg_day = round(brand_total_mentions / days_count, 1)
            sov_rows.append([
                Paragraph(f"<b>{(project.get('brand_domain') or project['name']).upper()}</b>  (Your Brand)", st_body),
                str(brand_total_mentions),
                f"{brand_share}%",
                str(brand_avg_day),
            ])

            for comp_name, comp_mentions in sorted_comps:
                comp_share = round((comp_mentions / all_entity_mentions) * 100, 1)
                comp_avg_day = round(comp_mentions / days_count, 1)
                sov_rows.append([
                    comp_name.upper(),
                    str(comp_mentions),
                    f"{comp_share}%",
                    str(comp_avg_day),
                ])

            sov_widths = [5.5 * cm, 3 * cm, 3.5 * cm, 3 * cm]
            sov_table = Table(sov_rows, colWidths=sov_widths)
            sov_style = _base_table_style(len(sov_rows))
            sov_style.append(('BACKGROUND', (0, 1), (-1, 1), CLR_GREEN_CELL))
            sov_style.append(('ALIGN', (0, 1), (0, -1), 'LEFT'))
            sov_table.setStyle(TableStyle(sov_style))
            elements.append(sov_table)
        else:
            elements.append(Paragraph("No competitor mention data available for this period.", st_no_data))

        elements.append(Spacer(1, 0.5 * cm))

        # SOV Over Time line chart (like the UI)
        if comp_snapshots and all_entity_mentions > 0:
            from reportlab.graphics.shapes import Drawing, Line, String, Rect
            from reportlab.graphics.charts.lineplots import LinePlot
            from reportlab.graphics.widgets.markers import makeMarker
            from reportlab.graphics import renderPDF

            elements.append(Paragraph("Share of Voice Over Time", st_subsection))
            elements.append(Spacer(1, 0.2 * cm))

            # Aggregate snapshots by date: brand mentions + competitor mentions
            date_brand = {}  # date_str -> total brand mentions
            date_comp = {}   # date_str -> {comp_name: mentions}
            date_total = {}  # date_str -> total all mentions

            for snap in comp_snapshots:
                ds = str(snap.get('snapshot_date', ''))
                bm = int(snap.get('total_mentions') or 0)
                date_brand[ds] = date_brand.get(ds, 0) + bm

                breakdown = snap.get('competitor_breakdown') or {}
                if isinstance(breakdown, str):
                    import json as json_mod2
                    try:
                        breakdown = json_mod2.loads(breakdown)
                    except Exception:
                        breakdown = {}
                for ck, cv in breakdown.items():
                    ck_display = comp_display_map.get(ck.lower().strip(), ck.lower().strip())
                    if ds not in date_comp:
                        date_comp[ds] = {}
                    date_comp[ds][ck_display] = date_comp[ds].get(ck_display, 0) + int(cv or 0)

            sorted_dates = sorted(date_brand.keys())
            if len(sorted_dates) >= 2:
                # Calculate SOV % per date
                chart_colors = [
                    colors.HexColor('#3B82F6'),  # Brand: blue
                    colors.HexColor('#EF4444'),  # Comp 1: red
                    colors.HexColor('#F97316'),  # Comp 2: orange
                    colors.HexColor('#10B981'),  # Comp 3: green
                    colors.HexColor('#8B5CF6'),  # Comp 4: purple
                ]

                # Build SOV series for each entity
                comp_names_ordered = [cn for cn, _ in sorted_comps[:4]]
                all_series_names = [(project.get('brand_domain') or project['name'])] + comp_names_ordered

                # Create data tuples for LinePlot: list of [(x, y), ...]
                plot_data = []
                for si, series_name in enumerate(all_series_names):
                    series_points = []
                    for di, ds in enumerate(sorted_dates):
                        if si == 0:  # brand
                            mentions = date_brand.get(ds, 0)
                        else:
                            mentions = date_comp.get(ds, {}).get(series_name, 0)
                        # SOV for this date
                        total_day = date_brand.get(ds, 0) + sum(date_comp.get(ds, {}).values())
                        sov_pct = round((mentions / total_day) * 100, 1) if total_day > 0 else 0
                        series_points.append((di, sov_pct))
                    plot_data.append(series_points)

                # Create Drawing
                chart_w = usable_width
                chart_h = 6.5 * cm
                drawing = Drawing(float(chart_w), float(chart_h))

                # Background
                drawing.add(Rect(0, 0, float(chart_w), float(chart_h),
                                 fillColor=colors.HexColor('#FAFAFA'), strokeColor=None))

                lp = LinePlot()
                lp.x = 40
                lp.y = 30
                lp.width = float(chart_w) - 60
                lp.height = float(chart_h) - 55
                lp.data = plot_data

                # Style each line
                for si in range(len(all_series_names)):
                    lp.lines[si].strokeColor = chart_colors[si % len(chart_colors)]
                    lp.lines[si].strokeWidth = 2.5 if si == 0 else 1.5
                    lp.lines[si].symbol = makeMarker('Circle')
                    lp.lines[si].symbol.size = 4 if si == 0 else 3

                # Axes
                lp.xValueAxis.valueMin = 0
                lp.xValueAxis.valueMax = len(sorted_dates) - 1
                lp.xValueAxis.valueSteps = list(range(len(sorted_dates)))
                lp.xValueAxis.labelTextFormat = lambda v: sorted_dates[int(v)][-5:] if 0 <= int(v) < len(sorted_dates) else ''
                lp.xValueAxis.labels.fontSize = 7
                lp.xValueAxis.labels.angle = 0
                lp.xValueAxis.strokeColor = colors.HexColor('#E5E7EB')

                lp.yValueAxis.valueMin = 0
                lp.yValueAxis.valueMax = 100
                lp.yValueAxis.valueStep = 20
                lp.yValueAxis.labelTextFormat = '%d%%'
                lp.yValueAxis.labels.fontSize = 7
                lp.yValueAxis.strokeColor = colors.HexColor('#E5E7EB')
                lp.yValueAxis.gridStrokeColor = colors.HexColor('#F3F4F6')
                lp.yValueAxis.visibleGrid = 1

                drawing.add(lp)

                # Legend text at bottom
                legend_x = 40
                for si, sn in enumerate(all_series_names):
                    color = chart_colors[si % len(chart_colors)]
                    drawing.add(Rect(legend_x, 5, 8, 8, fillColor=color, strokeColor=None))
                    label = sn[:15].upper()
                    drawing.add(String(legend_x + 11, 6, label, fontSize=6,
                                       fillColor=colors.HexColor('#6B7280')))
                    legend_x += len(label) * 4 + 22

                elements.append(drawing)
            else:
                elements.append(Paragraph("Not enough data points for SOV trend chart.", st_no_data))

        elements.append(Spacer(1, 0.5 * cm))

        # Key insight
        if all_entity_mentions > 0 and sorted_comps:
            top_comp_name, top_comp_mentions = sorted_comps[0]
            insight_text = (
                f"In the analyzed {days}-day period, <b>{project.get('brand_domain') or project['name']}</b> "
                f"received <b>{brand_total_mentions}</b> total mentions across all LLMs, "
                f"representing <b>{brand_share}%</b> of all entity mentions. "
                f"The top competitor by mentions was <b>{top_comp_name}</b> "
                f"with <b>{top_comp_mentions}</b> mentions."
            )
            elements.append(Paragraph(insight_text, st_body))

        # =================================================================
        # PAGE 4: PROMPT PERFORMANCE
        # =================================================================
        elements.append(PageBreak())
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph("Prompt Performance", st_section))
        elements.append(Paragraph("Top 20 prompts by visibility", st_body))
        elements.append(Spacer(1, 0.3 * cm))

        if prompt_data:
            pr_header = ["Prompt", "Type", "Brand Mentions", "Visibility %", "Avg Pos"]
            pr_rows = [pr_header]
            pr_row_types = []  # 'branded', 'non-branded' for coloring
            for p in prompt_data:
                qt = p.get('query_text', '') or ''
                is_branded = classify_query_branded(qt, brand_keywords_pdf)
                type_label = '🏷️ Branded' if is_branded else '🌿 Generic'
                total_r = int(p.get('total_results') or 0)
                ment = int(p.get('mentions') or 0)
                vis_pct = round((ment / total_r) * 100, 1) if total_r > 0 else 0
                avg_p = round(float(p.get('avg_position') or 0), 1)
                pr_row_types.append('branded' if is_branded else 'generic')
                pr_rows.append([
                    Paragraph(_truncate(qt, 55), st_body),
                    type_label,
                    str(ment),
                    f"{vis_pct}%",
                    f"#{avg_p}" if avg_p > 0 else 'N/A',
                ])

            pr_widths = [6 * cm, 2.2 * cm, 2.2 * cm, 2 * cm, 1.6 * cm]
            pr_table = Table(pr_rows, colWidths=pr_widths)
            pr_style = _base_table_style(len(pr_rows))
            pr_style.append(('ALIGN', (0, 1), (0, -1), 'LEFT'))
            # Color rows: green for mentioned, red for 0 mentions
            for ri, rtype in enumerate(pr_row_types):
                r = ri + 1
                ment_val = int(pr_rows[r][2])
                if ment_val > 0:
                    pr_style.append(('TEXTCOLOR', (2, r), (2, r), CLR_DELTA_UP))
                else:
                    pr_style.append(('TEXTCOLOR', (2, r), (2, r), CLR_DELTA_DOWN))
                # Visibility color
                vis_str = pr_rows[r][3]
                vis_val = float(vis_str.replace('%', '')) if '%' in str(vis_str) else 0
                if vis_val >= 50:
                    pr_style.append(('TEXTCOLOR', (3, r), (3, r), CLR_DELTA_UP))
                elif vis_val == 0:
                    pr_style.append(('TEXTCOLOR', (3, r), (3, r), CLR_DELTA_DOWN))
            pr_table.setStyle(TableStyle(pr_style))
            elements.append(pr_table)
        else:
            elements.append(Paragraph("No prompt data available for this period.", st_no_data))

        # =================================================================
        # PAGE 5: TOP CITED URLs
        # =================================================================
        elements.append(PageBreak())
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph("Most Cited URLs by LLMs", st_section))
        elements.append(Paragraph("Top 15 URLs by total mentions", st_body))
        elements.append(Spacer(1, 0.3 * cm))

        # Style for clickable URL links
        st_url_link = ParagraphStyle('URLLink', parent=st_body,
                                      textColor=colors.HexColor('#2563EB'), fontSize=8)
        brand_domain = (project.get('brand_domain') or '').lower().strip()

        if urls_data:
            url_header = ["Rank", "URL", "Mentions", "% of Total"]
            url_rows = [url_header]
            url_is_brand = []
            for u in urls_data[:15]:
                raw_url = u.get('url', '') or ''
                display_url = _truncate(raw_url, 55)
                # Make clickable link
                if raw_url.startswith('http'):
                    url_para = Paragraph(f'<a href="{raw_url}" color="#2563EB">{display_url}</a>', st_url_link)
                else:
                    url_para = Paragraph(display_url, st_body)
                is_brand_url = brand_domain and brand_domain in raw_url.lower()
                url_is_brand.append(is_brand_url)
                url_rows.append([
                    str(u.get('rank', '')),
                    url_para,
                    str(u.get('mentions', 0)),
                    f"{round(float(u.get('percentage', 0)), 1)}%",
                ])

            url_widths = [1.2 * cm, 8.5 * cm, 2 * cm, 2.3 * cm]
            url_table = Table(url_rows, colWidths=url_widths)
            url_style = _base_table_style(len(url_rows))
            url_style.append(('ALIGN', (1, 1), (1, -1), 'LEFT'))
            # Highlight brand URLs in green
            for ri, is_brand in enumerate(url_is_brand):
                if is_brand:
                    url_style.append(('BACKGROUND', (0, ri + 1), (-1, ri + 1), CLR_GREEN_CELL))
            url_table.setStyle(TableStyle(url_style))
            elements.append(url_table)
        else:
            elements.append(Paragraph("No cited URL data available for this period.", st_no_data))

        # ── Build PDF ──
        doc.build(elements, onFirstPage=_on_first_page, onLaterPages=_on_later_pages)
        output.seek(0)

        safe_name = re.sub(r'[^a-zA-Z0-9_-]', '-', project['name'])
        filename = f"llm-monitoring-{safe_name}-{datetime.now().strftime('%Y%m%d')}.pdf"

        logger.info(f"PDF exported successfully for project {project_id}")

        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Error exporting PDF for project {project_id}: {e}", exc_info=True)
        return jsonify({'error': f'Error exportando PDF: {str(e)}'}), 500


# ============================================================================
# CRON: Model Discovery v2 - Smart Chat-Only con Aprobación por Email
# ============================================================================

# Filtros para identificar modelos de CHAT (no reasoning, imagen, video, embedding, etc.)
CHAT_MODEL_FILTERS = {
    'openai': {
        'include_patterns': ['gpt-5', 'gpt-4o', 'gpt-4.1'],
        'exclude_patterns': [
            'o1', 'o3', 'o4',           # Reasoning/thinking models
            'dall-e', 'gpt-image',       # Image generation
            'whisper', 'tts', 'audio',   # Audio models
            'embedding', 'moderation',   # Utility models
            'realtime',                  # Realtime API models
            'davinci', 'babbage',        # Legacy completion models
            'mini-audio', 'mini-realtime',
        ],
    },
    'google': {
        'include_patterns': ['gemini'],
        'exclude_patterns': [
            'embedding', 'imagen', 'veo',   # Image/video generation
            'chirp', 'codey', 'medlm',      # Specialized models
            'gemma',                          # Open-weight (not API chat)
            'aqa',                            # Attributed QA only
        ],
    },
    'anthropic': {
        'include_patterns': ['claude-sonnet', 'claude-opus', 'claude-haiku'],
        'exclude_patterns': ['claude-instant'],
    },
    'perplexity': {
        'include_patterns': ['sonar', 'sonar-pro'],
        'exclude_patterns': [
            'sonar-reasoning', 'sonar-deep-research',  # Reasoning/research models
        ],
    },
}

# Pricing estimado por tier de modelo (input, output por 1M tokens)
DEFAULT_PRICING_BY_TIER = {
    'openai': {
        'gpt-5': (2.50, 10.00), 'gpt-4o': (2.50, 10.00), 'gpt-4.1': (2.00, 8.00),
    },
    'anthropic': {
        'claude-sonnet': (3.00, 15.00), 'claude-opus': (15.00, 75.00), 'claude-haiku': (0.25, 1.25),
    },
    'google': {
        'gemini-3-flash': (0.50, 3.00), 'gemini-3-pro': (1.25, 5.00),
        'gemini-2': (0.30, 1.25), 'gemini-1': (0.15, 0.60),
    },
    'perplexity': {
        'sonar': (1.00, 1.00), 'sonar-pro': (3.00, 15.00),
    },
}


def is_chat_model(provider: str, model_id: str) -> bool:
    """
    Determina si un modelo es de tipo CHAT (apto para monitorización de marcas).
    Rechaza modelos de reasoning, imagen, video, embedding, audio, etc.
    """
    model_lower = model_id.lower()
    filters = CHAT_MODEL_FILTERS.get(provider, {})

    # Verificar exclusiones primero (mayor prioridad)
    for pattern in filters.get('exclude_patterns', []):
        if pattern.lower() in model_lower:
            return False

    # Verificar inclusiones
    for pattern in filters.get('include_patterns', []):
        if pattern.lower() in model_lower:
            return True

    return False


def estimate_pricing_for_model(provider: str, model_id: str) -> tuple:
    """
    Estima pricing para un modelo nuevo basado en su tier/familia.
    Returns: (cost_per_1m_input, cost_per_1m_output) o (None, None) si no se puede estimar.
    """
    model_lower = model_id.lower()
    tier_pricing = DEFAULT_PRICING_BY_TIER.get(provider, {})

    # Buscar el tier más específico que coincida
    best_match = None
    best_match_len = 0
    for tier_key, pricing in tier_pricing.items():
        if tier_key.lower() in model_lower and len(tier_key) > best_match_len:
            best_match = pricing
            best_match_len = len(tier_key)

    return best_match if best_match else (None, None)


def validate_model_before_switch(provider: str, model_id: str) -> dict:
    """
    Ejecuta una query de prueba al modelo para verificar que funciona
    antes de activarlo como modelo actual.

    Returns: dict con 'success', 'response_length', 'error'
    """
    test_prompt = "What is SEO? Answer in one sentence."
    try:
        from services.llm_providers.provider_factory import LLMProviderFactory
        provider_instance = LLMProviderFactory.create_provider(provider, model_id=model_id)
        if not provider_instance:
            return {'success': False, 'error': f'Could not create provider for {provider}/{model_id}'}

        result = provider_instance.execute_query(test_prompt)
        response_text = result.get('response', '') if result else ''

        if len(response_text) > 10:
            return {'success': True, 'response_length': len(response_text), 'error': None}
        else:
            return {'success': False, 'response_length': len(response_text), 'error': 'Response too short'}

    except Exception as e:
        logger.error(f"Pre-switch validation failed for {provider}/{model_id}: {e}")
        return {'success': False, 'response_length': 0, 'error': str(e)[:200]}


def _log_model_change(cur, provider, old_model_id, new_model_id, old_display, new_display,
                      change_type, changed_by, reason=None, metadata=None):
    """Registra un cambio de modelo en llm_model_changelog."""
    import json
    cur.execute("""
        INSERT INTO llm_model_changelog
            (llm_provider, old_model_id, new_model_id, old_display_name, new_display_name,
             change_type, changed_by, reason, metadata)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (provider, old_model_id, new_model_id, old_display, new_display,
          change_type, changed_by, reason, json.dumps(metadata or {})))


def get_model_version_score(model_id: str) -> tuple:
    """
    Calcula un score de versión para comparar modelos.
    Retorna una tupla (major_version, sub_version, date_score, model_id) para ordenar.
    """
    import re

    model_lower = model_id.lower()

    major = 0
    if 'gpt-5' in model_lower or 'gpt5' in model_lower:
        major = 5
    elif 'gpt-4.1' in model_lower:
        major = 4.1
    elif 'gpt-4' in model_lower or 'gpt4' in model_lower:
        major = 4
    elif 'o3' in model_lower:
        major = 6
    elif 'o1' in model_lower:
        major = 5.5
    elif 'gemini-3' in model_lower:
        major = 3
    elif 'gemini-2' in model_lower:
        major = 2
    elif 'gemini-1' in model_lower:
        major = 1
    elif 'claude-sonnet-4' in model_lower or 'claude-4' in model_lower:
        major = 4
    elif 'claude-3' in model_lower:
        major = 3
    elif 'sonar-pro' in model_lower:
        major = 2
    elif 'sonar-reasoning' in model_lower:
        major = 3
    elif 'sonar' in model_lower:
        major = 1

    sub_version = 0
    sub_match = re.search(r'(?:gpt|gemini)-(\d+)\.(\d+)', model_lower)
    if not sub_match:
        sub_match = re.search(r'claude-(?:sonnet|opus|haiku)-(\d+)-(\d+)', model_lower)
    if sub_match:
        sub_version = int(sub_match.group(2))

    date_score = 0
    date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', model_id)
    if date_match:
        date_score = int(f"{date_match.group(1)}{date_match.group(2)}{date_match.group(3)}")
    else:
        if 'preview' in model_lower or 'latest' in model_lower:
            date_score = 99999998
        else:
            date_score = 99999999

    return (major, sub_version, date_score, model_id)


def is_model_newer(new_model_id: str, current_model_id: str) -> bool:
    """Determina si new_model_id es más nuevo que current_model_id."""
    new_score = get_model_version_score(new_model_id)
    current_score = get_model_version_score(current_model_id)
    return (new_score[0], new_score[1], new_score[2]) > (current_score[0], current_score[1], current_score[2])


@llm_monitoring_bp.route('/cron/model-discovery', methods=['POST'])
@cron_or_auth_required
def cron_model_discovery():
    """
    Endpoint para CRON de descubrimiento de modelos LLM v2.

    Mejoras v2:
    - Filtra modelos NO-CHAT (reasoning, imagen, video, embedding, etc.)
    - Flujo de aprobación por email con tokens seguros
    - Validación pre-switch antes de activar un modelo
    - Estimación automática de pricing
    - Registro en llm_model_changelog

    Query params:
        - notify_email: Email para notificación (default: info@soycarlosgonzalez.com)
        - auto_update: true/false - Aprobar automáticamente (con validación pre-switch)
    """
    auth_error = _ensure_cron_token_or_admin()
    if auth_error:
        return auth_error

    import openai
    import google.generativeai as genai

    notify_email = request.args.get('notify_email', 'info@soycarlosgonzalez.com')
    auto_update = request.args.get('auto_update', 'false').lower() == 'true'

    logger.info("=" * 60)
    logger.info("🔍 CRON: Smart Model Discovery v2...")
    logger.info(f"   Notify email: {notify_email}")
    logger.info(f"   Auto-update: {auto_update}")
    logger.info("=" * 60)

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500

    try:
        cur = conn.cursor()

        # 1. Obtener modelos actuales de BD
        cur.execute("""
            SELECT llm_provider, model_id, model_display_name, is_current
            FROM llm_model_registry
            ORDER BY llm_provider, is_current DESC
        """)
        db_models = cur.fetchall()

        current_models = {}
        current_display_names = {}
        known_model_ids = set()
        for m in db_models:
            known_model_ids.add(m['model_id'])
            if m['is_current']:
                current_models[m['llm_provider']] = m['model_id']
                current_display_names[m['llm_provider']] = m['model_display_name'] or m['model_id']

        logger.info(f"📊 Modelos actuales en BD: {current_models}")

        # 2. Descubrir modelos de cada proveedor (SOLO CHAT)
        discovered_models = []
        newer_chat_models = []
        skipped_non_chat = []
        older_models = []
        same_or_known = []
        errors = []

        # --- OpenAI ---
        try:
            openai_key = os.getenv('OPENAI_API_KEY')
            if openai_key:
                client = openai.OpenAI(api_key=openai_key)
                models = client.models.list()
                current_openai = current_models.get('openai', 'gpt-5.4')

                for m in models.data:
                    discovered_models.append({'provider': 'openai', 'model_id': m.id})

                    if m.id in known_model_ids:
                        same_or_known.append({'provider': 'openai', 'model_id': m.id})
                        continue

                    if not is_chat_model('openai', m.id):
                        skipped_non_chat.append({'provider': 'openai', 'model_id': m.id, 'reason': 'non-chat'})
                        continue

                    model_info = {'provider': 'openai', 'model_id': m.id, 'display_name': m.id}
                    if is_model_newer(m.id, current_openai):
                        model_info['status'] = 'NEWER'
                        newer_chat_models.append(model_info)
                    else:
                        older_models.append(model_info)

                logger.info(f"✅ OpenAI: Consultado correctamente")
        except Exception as e:
            errors.append(f"OpenAI: {str(e)[:100]}")
            logger.error(f"❌ OpenAI error: {e}")

        # --- Google ---
        try:
            google_key = os.getenv('GOOGLE_AI_API_KEY') or os.getenv('GOOGLE_API_KEY')
            if google_key:
                genai.configure(api_key=google_key)
                models = genai.list_models()
                current_google = current_models.get('google', 'gemini-3-flash-preview')

                for m in models:
                    model_id = m.name.split('/')[-1] if '/' in m.name else m.name
                    display = getattr(m, 'display_name', model_id)
                    discovered_models.append({'provider': 'google', 'model_id': model_id})

                    if model_id in known_model_ids:
                        same_or_known.append({'provider': 'google', 'model_id': model_id})
                        continue

                    if not is_chat_model('google', model_id):
                        skipped_non_chat.append({'provider': 'google', 'model_id': model_id, 'reason': 'non-chat'})
                        continue

                    model_info = {'provider': 'google', 'model_id': model_id, 'display_name': display}
                    if is_model_newer(model_id, current_google):
                        model_info['status'] = 'NEWER'
                        newer_chat_models.append(model_info)
                    else:
                        older_models.append(model_info)

                logger.info(f"✅ Google: Consultado correctamente")
        except Exception as e:
            errors.append(f"Google: {str(e)[:100]}")
            logger.error(f"❌ Google error: {e}")

        # --- Perplexity (lista estática, solo chat) ---
        perplexity_models = ['sonar', 'sonar-pro']
        current_perplexity = current_models.get('perplexity', 'sonar-pro')
        for model_id in perplexity_models:
            discovered_models.append({'provider': 'perplexity', 'model_id': model_id})
            if model_id not in known_model_ids and is_chat_model('perplexity', model_id):
                model_info = {'provider': 'perplexity', 'model_id': model_id,
                              'display_name': f'Perplexity {model_id.replace("-", " ").title()}'}
                if is_model_newer(model_id, current_perplexity):
                    model_info['status'] = 'NEWER'
                    newer_chat_models.append(model_info)
                else:
                    older_models.append(model_info)

        logger.info(f"📊 Resultados: {len(newer_chat_models)} chat models nuevos, "
                    f"{len(skipped_non_chat)} no-chat descartados")

        # 3. Procesar modelos nuevos de chat
        models_added = []
        models_auto_approved = []
        approval_tokens_generated = []
        public_base_url = os.getenv('PUBLIC_BASE_URL', 'https://app.clicandseo.com')

        for model in newer_chat_models:
            try:
                # Estimar pricing
                est_input, est_output = estimate_pricing_for_model(model['provider'], model['model_id'])

                # Generar token de aprobación
                approval_token = secrets.token_urlsafe(64)
                token_expires = datetime.now() + timedelta(days=7)

                if auto_update:
                    # Auto-update: validar y activar directamente
                    validation = validate_model_before_switch(model['provider'], model['model_id'])

                    if validation['success']:
                        # Insertar como current
                        cur.execute("""
                            INSERT INTO llm_model_registry
                                (llm_provider, model_id, model_display_name, is_current, is_available,
                                 model_category, cost_per_1m_input_tokens, cost_per_1m_output_tokens,
                                 pre_switch_validated)
                            VALUES (%s, %s, %s, TRUE, TRUE, 'chat', %s, %s, TRUE)
                            ON CONFLICT (llm_provider, model_id) DO UPDATE SET
                                is_current = TRUE, pre_switch_validated = TRUE, updated_at = NOW()
                        """, (model['provider'], model['model_id'], model['display_name'],
                              est_input, est_output))

                        # Desactivar modelo anterior
                        cur.execute("""
                            UPDATE llm_model_registry SET is_current = FALSE, updated_at = NOW()
                            WHERE llm_provider = %s AND model_id != %s AND is_current = TRUE
                        """, (model['provider'], model['model_id']))

                        # Registrar en changelog
                        old_model = current_models.get(model['provider'])
                        old_display = current_display_names.get(model['provider'])
                        _log_model_change(cur, model['provider'], old_model, model['model_id'],
                                         old_display, model['display_name'], 'auto_approved', 'cron',
                                         'Auto-approved after successful validation',
                                         {'validation': validation, 'estimated_pricing': {'input': est_input, 'output': est_output}})

                        models_auto_approved.append(model)
                        models_added.append(model)
                        logger.info(f"   ✅ Auto-aprobado: {model['provider']} / {model['model_id']}")
                    else:
                        logger.warning(f"   ⚠️ Validación fallida para {model['model_id']}: {validation['error']}")
                        errors.append(f"Validation failed for {model['model_id']}: {validation['error']}")
                else:
                    # Insertar como pendiente de aprobación
                    cur.execute("""
                        INSERT INTO llm_model_registry
                            (llm_provider, model_id, model_display_name, is_current, is_available,
                             model_category, cost_per_1m_input_tokens, cost_per_1m_output_tokens,
                             pending_approval, approval_token, approval_token_expires_at)
                        VALUES (%s, %s, %s, FALSE, TRUE, 'chat', %s, %s, TRUE, %s, %s)
                        ON CONFLICT (llm_provider, model_id) DO UPDATE SET
                            pending_approval = TRUE, approval_token = EXCLUDED.approval_token,
                            approval_token_expires_at = EXCLUDED.approval_token_expires_at,
                            updated_at = NOW()
                    """, (model['provider'], model['model_id'], model['display_name'],
                          est_input, est_output, approval_token, token_expires))

                    # Registrar en changelog como discovery
                    old_model = current_models.get(model['provider'])
                    old_display = current_display_names.get(model['provider'])
                    _log_model_change(cur, model['provider'], old_model, model['model_id'],
                                     old_display, model['display_name'], 'discovery', 'cron',
                                     'New chat model discovered, pending email approval')

                    model['approval_token'] = approval_token
                    model['estimated_pricing'] = {'input': est_input, 'output': est_output}
                    approval_tokens_generated.append(model)
                    models_added.append(model)
                    logger.info(f"   ⏳ Pendiente aprobación: {model['provider']} / {model['model_id']}")

            except Exception as e:
                logger.error(f"   ❌ Error procesando {model['model_id']}: {e}")
                errors.append(f"Processing {model['model_id']}: {str(e)[:100]}")

        conn.commit()

        # 4. Obtener estado final
        cur.execute("""
            SELECT llm_provider, model_id, model_display_name
            FROM llm_model_registry
            WHERE is_current = TRUE
            ORDER BY llm_provider
        """)
        final_models = {row['llm_provider']: row['model_id'] for row in cur.fetchall()}

        # 5. Enviar email con botones de aprobación
        email_sent = False
        if notify_email:
            try:
                from email_service import send_email

                if newer_chat_models and not auto_update:
                    subject = f"🆕 {len(newer_chat_models)} nuevo(s) modelo(s) de Chat detectados - Aprobación requerida"
                elif models_auto_approved:
                    subject = f"✅ {len(models_auto_approved)} modelo(s) auto-aprobados y activados"
                else:
                    subject = "🤖 LLM Model Discovery - Todo actualizado ✅"

                html_body = f"""
                <div style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 650px; margin: 0 auto;">
                    <h1 style="color: #161616; border-bottom: 3px solid #D8F9B8; padding-bottom: 10px;">
                        🔍 LLM Model Discovery Report v2
                    </h1>
                    <p style="color: #666;">Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')} UTC</p>
                """

                # Modelos actuales
                html_body += """
                    <h2 style="color: #161616;">📊 Modelos Actuales</h2>
                    <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px;">
                        <tr style="background: #161616; color: #D8F9B8;">
                            <th style="padding: 10px; text-align: left;">Proveedor</th>
                            <th style="padding: 10px; text-align: left;">Modelo Activo</th>
                        </tr>
                """
                for provider, mid in final_models.items():
                    html_body += f"""
                        <tr style="border-bottom: 1px solid #eee;">
                            <td style="padding: 10px;">{provider.upper()}</td>
                            <td style="padding: 10px;"><code>{mid}</code></td>
                        </tr>
                    """
                html_body += "</table>"

                # Modelos auto-aprobados
                if models_auto_approved:
                    html_body += f"""
                    <h2 style="color: #22c55e;">✅ Modelos Auto-Aprobados ({len(models_auto_approved)})</h2>
                    <p style="color: #666; font-size: 14px;">Validados y activados automáticamente:</p>
                    <ul style="background: #f0fdf4; padding: 15px 15px 15px 35px; border-radius: 8px; border-left: 4px solid #22c55e;">
                    """
                    for m in models_auto_approved:
                        html_body += f"<li><strong>{m['provider'].upper()}</strong>: <code>{m['model_id']}</code> ✅ Validación exitosa</li>"
                    html_body += "</ul>"

                # Modelos pendientes de aprobación (con botones)
                if approval_tokens_generated:
                    html_body += f"""
                    <h2 style="color: #f59e0b;">⏳ Modelos Pendientes de Aprobación ({len(approval_tokens_generated)})</h2>
                    <p style="color: #666; font-size: 14px;">Nuevos modelos de Chat detectados. Aprueba o rechaza cada uno:</p>
                    """
                    for m in approval_tokens_generated:
                        approve_url = f"{public_base_url}/api/llm-monitoring/models/approve?token={m['approval_token']}"
                        reject_url = f"{public_base_url}/api/llm-monitoring/models/reject?token={m['approval_token']}"
                        pricing_text = ""
                        if m.get('estimated_pricing', {}).get('input'):
                            pricing_text = f"Pricing estimado: ${m['estimated_pricing']['input']}/{m['estimated_pricing']['output']} por 1M tokens"

                        html_body += f"""
                        <div style="background: #fffbeb; padding: 20px; border-radius: 12px; border-left: 4px solid #f59e0b; margin-bottom: 16px;">
                            <h3 style="margin: 0 0 8px 0; color: #161616;">
                                {m['provider'].upper()}: <code>{m['model_id']}</code>
                            </h3>
                            <p style="color: #666; font-size: 13px; margin: 4px 0;">
                                Reemplazaría a: <code>{current_models.get(m['provider'], 'N/A')}</code>
                            </p>
                            <p style="color: #666; font-size: 13px; margin: 4px 0;">
                                Tipo: Chat | {pricing_text}
                            </p>
                            <p style="color: #92400e; font-size: 12px; margin: 4px 0;">
                                Token expira en 7 días
                            </p>
                            <div style="margin-top: 16px; display: flex; gap: 12px;">
                                <a href="{approve_url}" style="display: inline-block; background: #22c55e; color: white; padding: 12px 24px; border-radius: 8px; text-decoration: none; font-weight: 600; font-size: 14px;">
                                    ✅ Aprobar y Activar
                                </a>
                                <a href="{reject_url}" style="display: inline-block; background: #ef4444; color: white; padding: 12px 24px; border-radius: 8px; text-decoration: none; font-weight: 600; font-size: 14px;">
                                    ❌ Rechazar
                                </a>
                            </div>
                        </div>
                        """

                # No hay nuevos modelos
                if not newer_chat_models:
                    html_body += """
                    <div style="background: #f0fdf4; padding: 20px; border-radius: 8px; text-align: center; border-left: 4px solid #22c55e;">
                        <p style="color: #166534; margin: 0; font-weight: 600;">
                            ✅ ¡Tu APP está usando los modelos de chat más recientes!
                        </p>
                    </div>
                    """

                # Non-chat models descartados
                if skipped_non_chat:
                    html_body += f"""
                    <details style="margin-top: 20px;">
                        <summary style="cursor: pointer; color: #666; font-size: 14px;">
                            🚫 {len(skipped_non_chat)} modelos no-chat descartados (reasoning, imagen, etc.)
                        </summary>
                        <ul style="background: #f9fafb; padding: 15px 15px 15px 35px; border-radius: 8px; color: #666; font-size: 13px; margin-top: 10px;">
                    """
                    for m in skipped_non_chat[:15]:
                        html_body += f"<li>{m['provider']}: {m['model_id']}</li>"
                    if len(skipped_non_chat) > 15:
                        html_body += f"<li>... y {len(skipped_non_chat) - 15} más</li>"
                    html_body += "</ul></details>"

                if errors:
                    html_body += f"""
                    <h3 style="color: #dc2626; margin-top: 20px;">⚠️ Errores</h3>
                    <ul style="color: #666; font-size: 14px;">
                    """
                    for err in errors:
                        html_body += f"<li>{err}</li>"
                    html_body += "</ul>"

                html_body += f"""
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #eee;">
                    <p style="color: #999; font-size: 12px; text-align: center;">
                        ClicAndSEO - LLM Visibility Monitor v2<br>
                        Smart Model Discovery con filtro de Chat y aprobación por email.<br>
                        <small>Total escaneados: {len(discovered_models)} | Chat nuevos: {len(newer_chat_models)} | No-chat descartados: {len(skipped_non_chat)}</small>
                    </p>
                </div>
                """

                send_email(notify_email, subject, html_body)
                email_sent = True
                logger.info(f"📧 Email enviado a {notify_email}")

            except Exception as e:
                logger.error(f"❌ Error enviando email: {e}")

        logger.info("=" * 60)
        logger.info(f"✅ Smart Discovery v2 completado")
        logger.info(f"   Modelos escaneados: {len(discovered_models)}")
        logger.info(f"   Nuevos modelos de Chat: {len(newer_chat_models)}")
        logger.info(f"   No-chat descartados: {len(skipped_non_chat)}")
        logger.info(f"   Auto-aprobados: {len(models_auto_approved)}")
        logger.info(f"   Pendientes aprobación: {len(approval_tokens_generated)}")
        logger.info("=" * 60)

        return jsonify({
            'success': True,
            'message': 'Smart Model Discovery v2 completed',
            'discovered_count': len(discovered_models),
            'newer_chat_models': [{'provider': m['provider'], 'model_id': m['model_id'], 'display_name': m.get('display_name')} for m in newer_chat_models],
            'skipped_non_chat_count': len(skipped_non_chat),
            'older_models_count': len(older_models),
            'models_added': [{'provider': m['provider'], 'model_id': m['model_id']} for m in models_added],
            'auto_approved': [{'provider': m['provider'], 'model_id': m['model_id']} for m in models_auto_approved],
            'pending_approval': len(approval_tokens_generated),
            'current_models': final_models,
            'email_sent': email_sent,
            'errors': errors,
            'summary': {
                'app_is_updated': len(newer_chat_models) == 0,
                'action_required': len(approval_tokens_generated) > 0
            }
        }), 200

    except Exception as e:
        logger.error(f"❌ Error en model discovery: {e}", exc_info=True)
        return jsonify({'error': f'Error en model discovery: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


# ============================================================================
# Model Approval/Rejection Endpoints (clickable from email)
# ============================================================================

@llm_monitoring_bp.route('/models/approve', methods=['GET'])
def approve_model_by_token():
    """
    Aprueba y activa un modelo pendiente usando el token del email.
    Ejecuta validación pre-switch antes de activar.
    Devuelve HTML para que el admin vea el resultado en el navegador.
    """
    token = request.args.get('token', '')
    if not token or len(token) < 20:
        return _render_approval_result('error', 'Token inválido o no proporcionado.')

    conn = get_db_connection()
    if not conn:
        return _render_approval_result('error', 'Error de conexión a base de datos.')

    try:
        cur = conn.cursor()

        # Buscar modelo con este token
        cur.execute("""
            SELECT id, llm_provider, model_id, model_display_name,
                   approval_token_expires_at, pending_approval
            FROM llm_model_registry
            WHERE approval_token = %s
        """, (token,))
        model = cur.fetchone()

        if not model:
            return _render_approval_result('error', 'Token no encontrado. Puede haber expirado o ya fue usado.')

        if not model['pending_approval']:
            return _render_approval_result('info', f"El modelo {model['model_id']} ya fue procesado anteriormente.")

        # Verificar expiración
        if model['approval_token_expires_at'] and datetime.now() > model['approval_token_expires_at']:
            return _render_approval_result('error',
                f"El token de aprobación para {model['model_id']} ha expirado. "
                f"Ejecuta un nuevo discovery para generar un nuevo token.")

        # Validación pre-switch
        logger.info(f"🔄 Validando modelo antes de activar: {model['llm_provider']}/{model['model_id']}")
        validation = validate_model_before_switch(model['llm_provider'], model['model_id'])

        if not validation['success']:
            # Marcar como no aprobado
            cur.execute("""
                UPDATE llm_model_registry
                SET pending_approval = FALSE, approval_token = NULL,
                    approval_token_expires_at = NULL, pre_switch_validated = FALSE,
                    updated_at = NOW()
                WHERE id = %s
            """, (model['id'],))
            conn.commit()

            return _render_approval_result('error',
                f"La validación pre-switch falló para {model['model_id']}. "
                f"Error: {validation['error']}. El modelo NO ha sido activado.")

        # Obtener modelo actual antes de desactivar
        cur.execute("""
            SELECT model_id, model_display_name
            FROM llm_model_registry
            WHERE llm_provider = %s AND is_current = TRUE
        """, (model['llm_provider'],))
        old_model = cur.fetchone()

        # Desactivar modelos anteriores del mismo provider
        cur.execute("""
            UPDATE llm_model_registry
            SET is_current = FALSE, updated_at = NOW()
            WHERE llm_provider = %s AND is_current = TRUE
        """, (model['llm_provider'],))

        # Activar nuevo modelo
        cur.execute("""
            UPDATE llm_model_registry
            SET is_current = TRUE, pending_approval = FALSE, approval_token = NULL,
                approval_token_expires_at = NULL, pre_switch_validated = TRUE,
                updated_at = NOW()
            WHERE id = %s
        """, (model['id'],))

        # Registrar en changelog
        _log_model_change(
            cur, model['llm_provider'],
            old_model['model_id'] if old_model else None,
            model['model_id'],
            old_model['model_display_name'] if old_model else None,
            model['model_display_name'],
            'email_approved', 'admin:email',
            'Approved via email link after successful pre-switch validation',
            {'validation': validation}
        )

        conn.commit()

        logger.info(f"✅ Modelo aprobado y activado: {model['llm_provider']}/{model['model_id']}")

        # Enviar email de confirmación
        try:
            from email_service import send_email
            notify_email = os.getenv('MODEL_DISCOVERY_EMAIL', 'info@soycarlosgonzalez.com')
            old_name = old_model['model_id'] if old_model else 'N/A'
            send_email(notify_email,
                      f"✅ Modelo activado: {model['llm_provider'].upper()} → {model['model_id']}",
                      f"""
                      <div style="font-family: -apple-system, sans-serif; max-width: 500px; margin: 0 auto;">
                          <h2 style="color: #22c55e;">✅ Modelo Activado Exitosamente</h2>
                          <p><strong>Proveedor:</strong> {model['llm_provider'].upper()}</p>
                          <p><strong>Nuevo modelo:</strong> <code>{model['model_id']}</code></p>
                          <p><strong>Modelo anterior:</strong> <code>{old_name}</code></p>
                          <p><strong>Validación pre-switch:</strong> ✅ Exitosa</p>
                          <p style="color: #666; font-size: 13px;">
                              El modal de modelos en la APP se actualizará automáticamente.
                          </p>
                      </div>
                      """)
        except Exception as e:
            logger.error(f"Error enviando email de confirmación: {e}")

        return _render_approval_result('success',
            f"Modelo {model['model_display_name'] or model['model_id']} aprobado y activado exitosamente "
            f"para {model['llm_provider'].upper()}. "
            f"Validación pre-switch: ✅ Exitosa.")

    except Exception as e:
        conn.rollback()
        logger.error(f"Error aprobando modelo: {e}", exc_info=True)
        return _render_approval_result('error', f'Error interno: {str(e)[:200]}')
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/models/reject', methods=['GET'])
def reject_model_by_token():
    """
    Rechaza un modelo pendiente usando el token del email.
    Devuelve HTML para el navegador.
    """
    token = request.args.get('token', '')
    if not token or len(token) < 20:
        return _render_approval_result('error', 'Token inválido.')

    conn = get_db_connection()
    if not conn:
        return _render_approval_result('error', 'Error de conexión a base de datos.')

    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT id, llm_provider, model_id, model_display_name, pending_approval
            FROM llm_model_registry
            WHERE approval_token = %s
        """, (token,))
        model = cur.fetchone()

        if not model:
            return _render_approval_result('error', 'Token no encontrado.')

        if not model['pending_approval']:
            return _render_approval_result('info', f"El modelo {model['model_id']} ya fue procesado.")

        # Rechazar: limpiar token y desactivar approval
        cur.execute("""
            UPDATE llm_model_registry
            SET pending_approval = FALSE, approval_token = NULL,
                approval_token_expires_at = NULL, is_available = FALSE,
                updated_at = NOW()
            WHERE id = %s
        """, (model['id'],))

        # Registrar en changelog
        _log_model_change(
            cur, model['llm_provider'], None, model['model_id'],
            None, model['model_display_name'],
            'rejected', 'admin:email',
            'Rejected via email link'
        )

        conn.commit()
        logger.info(f"❌ Modelo rechazado: {model['llm_provider']}/{model['model_id']}")

        return _render_approval_result('rejected',
            f"Modelo {model['model_display_name'] or model['model_id']} rechazado para "
            f"{model['llm_provider'].upper()}. No se realizaron cambios.")

    except Exception as e:
        conn.rollback()
        logger.error(f"Error rechazando modelo: {e}", exc_info=True)
        return _render_approval_result('error', f'Error interno: {str(e)[:200]}')
    finally:
        cur.close()
        conn.close()


def _render_approval_result(status: str, message: str) -> str:
    """Renderiza una página HTML simple con el resultado de la aprobación/rechazo."""
    colors = {
        'success': ('#22c55e', '#f0fdf4', '✅'),
        'error': ('#ef4444', '#fef2f2', '❌'),
        'rejected': ('#f59e0b', '#fffbeb', '🚫'),
        'info': ('#3b82f6', '#eff6ff', 'ℹ️'),
    }
    color, bg, icon = colors.get(status, colors['info'])

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>Model {status.title()} - ClicAndSEO</title>
    </head>
    <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                 display: flex; justify-content: center; align-items: center;
                 min-height: 100vh; margin: 0; background: #f9fafb;">
        <div style="max-width: 500px; background: white; border-radius: 16px; padding: 40px;
                    box-shadow: 0 4px 20px rgba(0,0,0,0.1); text-align: center;">
            <div style="font-size: 48px; margin-bottom: 16px;">{icon}</div>
            <h1 style="color: {color}; margin: 0 0 16px 0; font-size: 24px;">
                Model {status.replace('_', ' ').title()}
            </h1>
            <p style="color: #374151; font-size: 16px; line-height: 1.6; margin: 0 0 24px 0;">
                {message}
            </p>
            <div style="padding: 16px; background: {bg}; border-radius: 8px; border-left: 4px solid {color};">
                <p style="color: #666; font-size: 13px; margin: 0;">
                    Puedes cerrar esta pestaña. Los cambios ya se han aplicado.
                </p>
            </div>
            <p style="color: #999; font-size: 12px; margin-top: 24px;">
                ClicAndSEO - LLM Visibility Monitor
            </p>
        </div>
    </body>
    </html>
    """
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ============================================================================
# Model Changelog Endpoint
# ============================================================================

@llm_monitoring_bp.route('/models/changelog', methods=['GET'])
@login_required
def get_model_changelog():
    """
    Obtiene el historial de cambios de modelos.

    Query params:
        - limit: número máximo de entradas (default 50, max 200)
        - provider: filtrar por proveedor (opcional)
    """
    limit = min(int(request.args.get('limit', 50)), 200)
    provider = request.args.get('provider')

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión a BD'}), 500

    try:
        cur = conn.cursor()

        query = """
            SELECT id, llm_provider, old_model_id, new_model_id,
                   old_display_name, new_display_name, change_type,
                   changed_by, reason, metadata, created_at
            FROM llm_model_changelog
        """
        params = []
        if provider:
            query += " WHERE llm_provider = %s"
            params.append(provider)
        query += " ORDER BY created_at DESC LIMIT %s"
        params.append(limit)

        cur.execute(query, params)
        rows = cur.fetchall()

        changelog = []
        for row in rows:
            entry = {
                'id': row['id'],
                'llm_provider': row['llm_provider'],
                'old_model_id': row['old_model_id'],
                'new_model_id': row['new_model_id'],
                'old_display_name': row['old_display_name'],
                'new_display_name': row['new_display_name'],
                'change_type': row['change_type'],
                'changed_by': row['changed_by'],
                'reason': row['reason'],
                'metadata': row['metadata'] if row['metadata'] else {},
                'created_at': row['created_at'].isoformat() if row['created_at'] else None,
            }
            changelog.append(entry)

        return jsonify({
            'success': True,
            'changelog': changelog,
            'count': len(changelog)
        }), 200

    except Exception as e:
        logger.error(f"Error obteniendo changelog: {e}", exc_info=True)
        return jsonify({'error': f'Error: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


@llm_monitoring_bp.route('/cron/alert', methods=['POST'])
@cron_or_auth_required
def cron_alert():
    """
    Endpoint para alertas de cron (fallos o warnings).
    Reutiliza el sistema de email existente.
    
    Payload JSON opcional:
        - notify_email
        - job_name
        - status (failed|warning|ok)
        - message
        - endpoint
        - response_status
        - response_body
        - run_at
    """
    auth_error = _ensure_cron_token_or_admin()
    if auth_error:
        return auth_error

    data = request.get_json(silent=True) or {}
    
    notify_email = (
        data.get('notify_email')
        or request.args.get('notify_email')
        or os.getenv('CRON_ALERT_EMAIL')
        or os.getenv('MODEL_DISCOVERY_EMAIL')
        or 'info@soycarlosgonzalez.com'
    )
    
    job_name = data.get('job_name') or 'Cron'
    status = (data.get('status') or 'failed').lower()
    message = data.get('message') or 'Fallo no especificado'
    endpoint = data.get('endpoint') or ''
    response_status = data.get('response_status')
    response_body = data.get('response_body') or ''
    run_at = data.get('run_at') or datetime.utcnow().isoformat()
    
    def _truncate(value: str, max_len: int = 2000) -> str:
        if value is None:
            return ''
        text = str(value)
        if len(text) <= max_len:
            return text
        return text[:max_len] + "…"
    
    status_icon = "🚨" if status == "failed" else "⚠️" if status == "warning" else "✅"
    subject = f"{status_icon} Cron {status.upper()}: {job_name}"
    
    html_body = f"""
    <div style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 600px; margin: 0 auto;">
        <h1 style="color: #161616; border-bottom: 3px solid #D8F9B8; padding-bottom: 10px;">
            {status_icon} Alerta de Cron
        </h1>
        <p style="color: #666;">Fecha: {run_at} UTC</p>
        <table style="width: 100%; border-collapse: collapse; margin-bottom: 16px;">
            <tr style="background: #161616; color: #D8F9B8;">
                <th style="padding: 10px; text-align: left;">Campo</th>
                <th style="padding: 10px; text-align: left;">Valor</th>
            </tr>
            <tr style="border-bottom: 1px solid #eee;">
                <td style="padding: 10px;">Job</td>
                <td style="padding: 10px;">{job_name}</td>
            </tr>
            <tr style="border-bottom: 1px solid #eee;">
                <td style="padding: 10px;">Estado</td>
                <td style="padding: 10px;">{status.upper()}</td>
            </tr>
            <tr style="border-bottom: 1px solid #eee;">
                <td style="padding: 10px;">Endpoint</td>
                <td style="padding: 10px;">{_truncate(endpoint, 300)}</td>
            </tr>
            <tr style="border-bottom: 1px solid #eee;">
                <td style="padding: 10px;">HTTP Status</td>
                <td style="padding: 10px;">{response_status}</td>
            </tr>
        </table>
        <h3 style="color: #161616;">Mensaje</h3>
        <pre style="background: #f9fafb; padding: 12px; border-radius: 8px; white-space: pre-wrap; word-break: break-word;">{_truncate(message)}</pre>
        <h3 style="color: #161616;">Respuesta</h3>
        <pre style="background: #f9fafb; padding: 12px; border-radius: 8px; white-space: pre-wrap; word-break: break-word;">{_truncate(response_body)}</pre>
        <hr style="margin: 24px 0; border: none; border-top: 1px solid #eee;">
        <p style="color: #999; font-size: 12px; text-align: center;">
            ClicAndSEO - Alerta automática de cron
        </p>
    </div>
    """
    
    email_sent = False
    if notify_email:
        try:
            from email_service import send_email
            email_sent = send_email(notify_email, subject, html_body)
            logger.info(f"📧 Alerta de cron enviada a {notify_email}")
        except Exception as e:
            logger.error(f"❌ Error enviando alerta de cron: {e}")
    
    return jsonify({
        'success': True,
        'email_sent': email_sent,
        'notify_email': notify_email,
        'job_name': job_name,
        'status': status
    }), 200


logger.info("✅ LLM Monitoring Blueprint loaded successfully")
